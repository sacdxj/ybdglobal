from pathlib import Path
import re
import pandas as pd

SRC = Path(r"C:\Users\Administrator\Downloads\USB charging cable.xlsx")
OUTDIR = Path("outputs/usb-seo")
OUT = OUTDIR / "usb_keyword_sitemap_v1.xlsx"

SEEDS = """USB cable|USB charging cable|USB data cable|USB cable manufacturer|USB cable supplier|USB cable factory|USB cable wholesale|USB cable OEM|USB cable ODM|custom USB cable|USB C cable|Type C cable|USB C charging cable|USB C data cable|USB C to USB C cable|USB C to USB A cable|USB C fast charging cable|USB C PD cable|100W USB C cable|240W USB C cable|USB4 cable|USB4 40Gbps cable|USB4 80Gbps cable|fast charging cable|PD fast charging cable|100W fast charge cable|140W charging cable|240W charging cable|GaN charger cable|high power charging cable|high speed data cable|10Gbps USB cable|20Gbps USB cable|40Gbps USB4 cable|data transfer cable|sync cable|HDMI cable|HDMI cable manufacturer|HDMI cable supplier|HDMI cable factory|HDMI cable wholesale|HDMI OEM cable|custom HDMI cable|HDMI 2.0 cable|HDMI 2.1 cable|8K HDMI cable|4K HDMI cable|10K HDMI cable|8K 60Hz HDMI cable|4K 120Hz HDMI cable|HDMI ultra high speed cable|gaming HDMI cable|PS5 HDMI cable|Xbox HDMI cable|TV HDMI cable|monitor HDMI cable|projector HDMI cable|fiber HDMI cable|HDMI AOC cable|active optical HDMI cable|optical HDMI cable|long distance HDMI cable|50m HDMI cable|100m HDMI cable|DisplayPort cable|DP cable|DP cable manufacturer|DP cable supplier|DP cable factory|DisplayPort wholesale|DP 1.4 cable|DP 2.0 cable|DP 2.1 cable|8K DisplayPort cable|16K DP cable|DisplayPort 8K 60Hz|DisplayPort 1.4 8K cable|gaming DP cable|144Hz DP cable|165Hz DisplayPort cable|240Hz DP cable|high refresh rate cable|monitor cable|GPU DisplayPort cable|USB hub|USB C hub|USB C docking station|USB docking station|laptop docking station|USB hub manufacturer|USB hub supplier|USB hub factory|USB hub wholesale|7 in 1 USB C hub|8 in 1 USB C hub|10 in 1 docking station|12 in 1 USB C hub|multiport adapter|USB C adapter|Type C adapter|HDMI USB C hub|USB C hub with HDMI|USB C hub with Ethernet|USB C hub PD charging|USB C hub 4K|USB C hub dual monitor|triple display docking station|USB charger|PD charger|GaN charger|fast charger|wall charger|USB C charger|20W charger|30W charger|65W charger|100W charger|140W charger|240W charger|GaN power adapter|charger manufacturer|charger supplier|electronics wholesaler|computer accessories distributor|mobile accessories distributor|IT accessories importer|electronics importer|USB cable importer|HDMI cable importer|consumer electronics brand|computer accessories brand|mobile accessories brand|OEM electronics company|ODM electronics supplier|private label electronics|electronics retailer|online electronics store|Amazon seller|Shopify electronics store|computer accessories shop""".split("|")

def norm(s):
    return re.sub(r"\s+", " ", str(s).lower().replace("–", "-").replace("—", "-")).strip()

def page(menu, pillar, name, url, funnel, role, cta, parent):
    return dict(menu=menu, pillar_page=pillar, page_name=name, assigned_url=url,
                funnel=funnel, page_role=role, cta=cta, parent_url=parent)

def classify(k):
    # Strong exclusions: consumer troubleshooting, unrelated cables/vehicles, local retail intent.
    noise = r"(^| )(repair|replacement|near me|walmart|best buy|target|dollar tree|tesla|iphone|ipad|apple|lightning|micro usb|mini usb|ethernet cable|coaxial|aux cable|vga|dvi|sata|printer cable|extension cord|jumper cable|car battery|no signal|not working|problem|issue|driver|manual|review|reviews|reddit)( |$)"
    if re.search(noise, k):
        return "discard_noise", "与连接/充电配件B2B采购主题不一致"
    if re.search(r"\b(anker|ugreen|belkin|startech|insignia|rocketfish|satechi|hyperdrive|vava|novoo|cable matters|onn|amazon basics)\b", k):
        return "discard_noise", "竞品/零售品牌词，暂不进入自有产品架构"
    if re.search(r"^(best|what|how|why|does|is|can|where)\b|\bfor macbook\b|\bfor laptop\b", k):
        return "discard_noise", "消费研究或信息查询，暂不作为第一阶段商业页"
    if re.search(r"hdmi (to|and) (usb|displayport|dp|optical)|displayport to hdmi|hdmi to displayport|usb c to (hdmi|displayport)", k):
        return "discard_noise", "转换器/转接线意图与当前核心产品页不同"
    b2b = r"manufacturer|supplier|factory|wholesale|wholesaler|bulk|oem|odm|custom|private label|importer|distributor"
    # Buyer solution pages first.
    if re.search(r"amazon seller|amazon fba", k):
        return "core_keep", page("Solutions", "Buyer Solutions", "For Amazon Sellers", "/solutions/amazon-sellers/", "Mid", "Buyer solution", "Discuss Private Label", "/solutions/")
    if re.search(r"wholesaler|wholesale electronics|electronics wholesale|bulk electronics", k):
        return "core_keep", page("Solutions", "Buyer Solutions", "For Wholesalers", "/solutions/wholesalers/", "Mid", "Buyer solution", "Get Bulk Pricing", "/solutions/")
    if re.search(r"importer|distributor", k) and re.search(r"electronics|accessories|usb|hdmi|cable|charger", k):
        return "core_keep", page("Solutions", "Buyer Solutions", "For Importers & Distributors", "/solutions/importers-distributors/", "Mid", "Buyer solution", "Discuss Supply", "/solutions/")
    if re.search(r"private label", k) and re.search(r"electronics|cable|charger|hub", k):
        return "core_keep", page("OEM / ODM", "Customization", "Private Label Electronics", "/oem-odm/private-label-electronics/", "Bottom", "OEM landing", "Start Private Label Project", "/oem-odm/")

    # Product family detection.
    family = None
    if re.search(r"displayport|\bdp\b", k): family = "dp"
    elif re.search(r"hdmi", k): family = "hdmi"
    elif re.search(r"docking station|usb[- ]?c hub|usb hub|multiport adapter", k): family = "hub"
    elif re.search(r"charger|power adapter|wall adapter", k): family = "charger"
    elif re.search(r"usb4|usb cable|usb[- ]?c cable|type[- ]?c cable|charging cable|data cable|sync cable", k): family = "usb"
    if not family:
        return "discard_noise", "不属于目标产品线或采购客户主题"

    specs = bool(re.search(r"\b(20|30|45|60|65|67|100|120|140|200|240)w\b|\b(5|10|20|40|80)gbps\b|\b(4k|8k|10k|16k)\b|\b(60|120|144|165|240)hz\b|\b(7|8|10|12)[- ]?in[- ]?1\b|2\.0|2\.1|1\.4|usb4|fiber|optical|aoc|dual monitor|triple display", k))
    procurement = bool(re.search(b2b, k))

    if family == "usb":
        if re.search(r"usb4", k):
            p = page("Products", "USB Cables", "USB4 Cables", "/products/usb-cables/usb4-cables/", "Bottom", "Specification category", "Request Quote", "/products/usb-cables/")
        elif re.search(r"240w", k):
            p = page("Products", "USB Cables", "240W USB-C Cables", "/products/usb-cables/240w-usb-c-cables/", "Bottom", "Specification category", "Request Sample", "/products/usb-cables/")
        elif re.search(r"100w", k):
            p = page("Products", "USB Cables", "100W USB-C Cables", "/products/usb-cables/100w-usb-c-cables/", "Bottom", "Specification category", "Request Sample", "/products/usb-cables/")
        elif procurement:
            p = page("Products", "USB Cables", "USB Cable Manufacturer", "/usb-cable-manufacturer/", "Bottom", "B2B landing", "Get a Quote", "/products/usb-cables/")
        elif re.search(r"data|gbps|sync", k):
            p = page("Products", "USB Cables", "USB-C Data Cables", "/products/usb-cables/usb-c-data-cables/", "Bottom", "Product category", "Request Quote", "/products/usb-cables/")
        elif re.search(r"charging|fast charge|pd", k):
            p = page("Products", "USB Cables", "USB-C Charging Cables", "/products/usb-cables/usb-c-charging-cables/", "Bottom", "Product category", "Request Quote", "/products/usb-cables/")
        else:
            p = page("Products", "USB Cables", "USB Cables", "/products/usb-cables/", "Bottom", "Pillar category", "Get a Quote", "/products/")
    elif family == "hdmi":
        if re.search(r"fiber|optical|aoc|long distance|\b(50|100)m\b", k): name, url = "Fiber & AOC HDMI Cables", "/products/hdmi-cables/fiber-aoc-hdmi-cables/"
        elif re.search(r"2\.1|8k|10k|120hz|ultra high speed", k): name, url = "HDMI 2.1 & 8K Cables", "/products/hdmi-cables/hdmi-2-1-cables/"
        elif procurement: name, url = "HDMI Cable Manufacturer", "/hdmi-cable-manufacturer/"
        else: name, url = "HDMI Cables", "/products/hdmi-cables/"
        p = page("Products", "HDMI Cables", name, url, "Bottom", "B2B landing" if procurement else "Product category", "Get a Quote", "/products/hdmi-cables/")
    elif family == "dp":
        if re.search(r"2\.1|16k", k): name, url = "DisplayPort 2.1 Cables", "/products/displayport-cables/dp-2-1-cables/"
        elif re.search(r"1\.4|8k", k): name, url = "DisplayPort 1.4 & 8K Cables", "/products/displayport-cables/dp-1-4-cables/"
        elif re.search(r"gaming|144hz|165hz|240hz|refresh|gpu", k): name, url = "Gaming DisplayPort Cables", "/products/displayport-cables/gaming-dp-cables/"
        elif procurement: name, url = "DisplayPort Cable Manufacturer", "/displayport-cable-manufacturer/"
        else: name, url = "DisplayPort Cables", "/products/displayport-cables/"
        p = page("Products", "DisplayPort Cables", name, url, "Bottom", "B2B landing" if procurement else "Product category", "Get a Quote", "/products/displayport-cables/")
    elif family == "hub":
        if procurement: name, url = "USB Hub Manufacturer", "/usb-hub-manufacturer/"
        elif re.search(r"docking station|dual monitor|triple display", k): name, url = "Laptop Docking Stations", "/products/hubs-docking/laptop-docking-stations/"
        elif re.search(r"7[- ]?in[- ]?1|8[- ]?in[- ]?1|10[- ]?in[- ]?1|12[- ]?in[- ]?1|multiport", k): name, url = "Multiport USB-C Hubs", "/products/hubs-docking/multiport-usb-c-hubs/"
        else: name, url = "USB Hubs & Docking Stations", "/products/hubs-docking/"
        p = page("Products", "Hubs & Docking", name, url, "Bottom", "B2B landing" if procurement else "Product category", "Configure & Quote", "/products/hubs-docking/")
    else:
        if re.search(r"gan", k): name, url = "GaN Chargers", "/products/chargers/gan-chargers/"
        elif re.search(r"65w|100w|140w|240w", k): name, url = "High-Power USB-C Chargers", "/products/chargers/high-power-usb-c-chargers/"
        elif procurement: name, url = "Charger Manufacturer", "/charger-manufacturer/"
        else: name, url = "USB-C & PD Chargers", "/products/chargers/"
        p = page("Products", "Chargers", name, url, "Bottom", "B2B landing" if procurement else "Product category", "Get a Quote", "/products/chargers/")

    label = "core_keep" if procurement or specs else "expand"
    return label, p

def main():
    OUTDIR.mkdir(parents=True, exist_ok=True)
    df = pd.read_excel(SRC, sheet_name="Sheet1")
    df = df.dropna(subset=["Keyword"]).copy()
    df["keyword_norm"] = df["Keyword"].map(norm)
    df = df.sort_values("Volume", ascending=False).drop_duplicates("keyword_norm")
    existing=set(df.keyword_norm)
    seed_rows=[]
    for seed in SEEDS:
        if norm(seed) not in existing:
            seed_rows.append({"Keyword":seed,"Intent":"Manual seed; Commercial","Volume":0,"Keyword Difficulty":pd.NA,"CPC (USD)":pd.NA,"keyword_norm":norm(seed)})
    if seed_rows:
        df=pd.concat([df,pd.DataFrame(seed_rows)],ignore_index=True)
    rows=[]
    for _, r in df.iterrows():
        label, info = classify(r.keyword_norm)
        base = {"keyword":r.Keyword,"search_volume":r.Volume,"intent":r.Intent,"difficulty":r["Keyword Difficulty"],"cpc_usd":r["CPC (USD)"],"label":label}
        if label == "discard_noise":
            base.update({"reason":info,"coverage_role":"do_not_place","serp_validation_required":False})
        else:
            base.update(info); base["reason"]="目标产品/规格/采购意图相关"; base["coverage_role"]="must_cover" if label=="core_keep" else "can_cover"
            base["serp_validation_required"] = bool(r.Volume >= 1000 or (pd.notna(r["Keyword Difficulty"]) and r["Keyword Difficulty"] >= 50))
        rows.append(base)
    review=pd.DataFrame(rows)
    kept=review[review.label.isin(["core_keep","expand"])].copy()
    kept["url_primary_keyword"] = kept.groupby("assigned_url")["search_volume"].transform("max")
    kept["priority_score"] = (kept.search_volume.fillna(0).map(lambda x: min(x,50000))/500) + kept.label.eq("core_keep")*35 + kept.intent.astype(str).str.contains("Commercial|Transactional",case=False)*20 - kept.difficulty.fillna(50)/5
    kept["priority"] = pd.cut(kept.priority_score,[-999,20,45,999],labels=["P3","P2","P1"])
    groups=[]
    for url,g in kept.groupby("assigned_url"):
        top=g.sort_values(["priority_score","search_volume"],ascending=False).iloc[0]
        groups.append({"assigned_url":url,"menu":top.menu,"pillar_page":top.pillar_page,"page_name":top.page_name,"parent_url":top.parent_url,"funnel":top.funnel,"page_role":top.page_role,"cta":top.cta,"keyword_count":len(g),"total_volume":int(g.search_volume.sum()),"primary_keyword":top.keyword,"max_difficulty":g.difficulty.max(),"priority":top.priority,"sample_keywords":" | ".join(g.sort_values("search_volume",ascending=False).keyword.head(15))})
    pages=pd.DataFrame(groups).sort_values(["priority","total_volume"],ascending=[True,False])
    sitemap=pd.concat([pd.DataFrame([
        {"nav_group":"Core","menu":"Home","pillar_page":"Home","page_name":"Home","funnel":"Bottom","page_role":"RFQ homepage","url":"/","parent_url":"","page_goal":"Explain capability and route buyers","cta":"Get a Quote"},
        {"nav_group":"Core","menu":"Products","pillar_page":"Products","page_name":"Products","funnel":"Mid","page_role":"Product directory","url":"/products/","parent_url":"/","page_goal":"Choose product family","cta":"Explore Products"},
        {"nav_group":"Core","menu":"Solutions","pillar_page":"Buyer Solutions","page_name":"Buyer Solutions","funnel":"Mid","page_role":"Buyer directory","url":"/solutions/","parent_url":"/","page_goal":"Choose buyer scenario","cta":"Discuss Your Requirements"},
        {"nav_group":"Core","menu":"OEM / ODM","pillar_page":"Customization","page_name":"OEM / ODM","funnel":"Bottom","page_role":"Customization landing","url":"/oem-odm/","parent_url":"/","page_goal":"Capture custom projects","cta":"Start Customization"},
        {"nav_group":"Trust","menu":"Quality & Factory","pillar_page":"Factory","page_name":"Quality & Factory","funnel":"Mid","page_role":"Trust page","url":"/quality-factory/","parent_url":"/","page_goal":"Prove manufacturing capability","cta":"Request Factory Profile"},
        {"nav_group":"Conversion","menu":"Contact","pillar_page":"RFQ","page_name":"Get a Quote","funnel":"Bottom","page_role":"RFQ form","url":"/get-a-quote/","parent_url":"/","page_goal":"Capture qualified inquiry","cta":"Submit RFQ"},
    ]), pages.rename(columns={"assigned_url":"url"}).assign(nav_group=lambda x:x.menu.map(lambda y:"Product" if y=="Products" else "Commercial"))[["nav_group","menu","pillar_page","page_name","funnel","page_role","url","parent_url","cta"]].assign(page_goal="Rank and convert its keyword cluster")],ignore_index=True)
    summary=pd.DataFrame([{"metric":"source_unique_keywords","value":len(review)},{"metric":"retained_keywords","value":len(kept)},{"metric":"discarded_keywords","value":len(review)-len(kept)},{"metric":"recommended_keyword_pages","value":len(pages)},{"metric":"P1_pages","value":int((pages.priority.astype(str)=="P1").sum())}])
    coverage=kept[["assigned_url","keyword","coverage_role","search_volume","intent","difficulty","priority","serp_validation_required","reason"]].rename(columns={"assigned_url":"page_url"})
    with pd.ExcelWriter(OUT,engine="openpyxl") as w:
        summary.to_excel(w, sheet_name="Summary", index=False)
        kept.to_excel(w, sheet_name="SEO_Keyword_Library", index=False)
        pages.to_excel(w, sheet_name="URL_Class_Summary", index=False)
        review.to_excel(w, sheet_name="Keyword_Review", index=False)
        pages.to_excel(w, sheet_name="Page_Clusters", index=False)
        sitemap.to_excel(w, sheet_name="Final_Sitemap", index=False)
        coverage.to_excel(w, sheet_name="Page_Keyword_Coverage", index=False)
        review[review.label.eq("discard_noise")].to_excel(w, sheet_name="Discarded_Keywords", index=False)
    from openpyxl import load_workbook
    wb=load_workbook(OUT)
    for ws in wb.worksheets:
        ws.freeze_panes="A2"; ws.auto_filter.ref=ws.dimensions
        ws.sheet_view.showGridLines=False
        for cell in ws[1]: cell.font=cell.font.copy(bold=True,color="FFFFFF"); cell.fill=cell.fill.copy(fill_type="solid",fgColor="1F4E78")
        for col in ws.columns:
            letter=col[0].column_letter; ws.column_dimensions[letter].width=min(48,max(11,max(len(str(c.value or "")) for c in list(col)[:300])+2))
    wb.save(OUT)
    print(OUT.resolve()); print(summary.to_string(index=False)); print(pages.head(20).to_string(index=False))

if __name__ == "__main__": main()
