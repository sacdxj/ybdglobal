from __future__ import annotations

from pathlib import Path
import pandas as pd


ROOT = Path(__file__).resolve().parents[1]
OUTPUT = ROOT / "outputs" / "usb-seo" / "usb_serp_validation_batch1.xlsx"
CHECKED_DATE = "2026-08-11"


def main() -> None:
    pages = pd.DataFrame([
        {
            "Query": "usb cable", "Scope": "Core Page", "Current URL": "/products/usb-cables/",
            "Observed Intent": "Broad consumer product/category and educational intent",
            "Observed Result Types": "Product/category pages; standards or explanatory content; consumer discussions",
            "Provisional Decision": "Keep category; validate B2B procurement with separate manufacturer terms",
            "Standalone Page": "Yes - existing category", "Confidence": "Medium",
            "Evidence Status": "representative-live-review", "Publication Gate": "closed",
            "Source 1": "https://en.wikipedia.org/wiki/USB",
            "Source 2": "https://www.reddit.com/r/AskElectronics/comments/1sxewiy/why_do_some_usbc_devices_only_charge_with_usba_to/",
            "Reason": "The broad query does not by itself prove manufacturer/RFQ intent; retain the range page and use procurement keywords for the OEM page.",
        },
        {
            "Query": "hdmi cable", "Scope": "Core Page", "Current URL": "/products/hdmi-cables/",
            "Observed Intent": "Mixed product selection, official cable-type education and buying guides",
            "Observed Result Types": "Official standards guidance; buying guides; product recommendations",
            "Provisional Decision": "Keep category plus separate technical/buying-guide support",
            "Standalone Page": "Yes - existing category", "Confidence": "High",
            "Evidence Status": "representative-live-review", "Publication Gate": "closed",
            "Source 1": "https://www.hdmi.org/resource/cables",
            "Source 2": "https://www.crutchfield.com/learn/learningcenter/home/cables/hdmi.html",
            "Reason": "The result mix supports a commercial category and educational support, but certification/performance claims remain SKU-evidence dependent.",
        },
        {
            "Query": "displayport cable", "Scope": "Core Page", "Current URL": "/products/displayport-cables/",
            "Observed Intent": "Product selection with strong connector, certification and performance education",
            "Observed Result Types": "Official VESA guidance; technical guides; product pages",
            "Provisional Decision": "Keep category and technical guide; avoid unverified version/performance claims",
            "Standalone Page": "Yes - existing category", "Confidence": "High",
            "Evidence Status": "representative-live-review", "Publication Gate": "closed",
            "Source 1": "https://www.displayport.org/how-to-choose-a-displayport-cable-and-not-get-a-bad-one/",
            "Source 2": "https://www.displayport.org/faq/",
            "Reason": "Official results emphasize connector choice and certification, supporting category plus technical selection content.",
        },
        {
            "Query": "usb c charger", "Scope": "Core Page", "Current URL": "/products/chargers/",
            "Observed Intent": "Consumer commercial investigation and product comparison",
            "Observed Result Types": "Best-product roundups; charger product pages; shopping discussions",
            "Provisional Decision": "Keep product category; do not treat broad query as B2B procurement proof",
            "Standalone Page": "Yes - existing category", "Confidence": "High",
            "Evidence Status": "representative-live-review", "Publication Gate": "closed",
            "Source 1": "https://www.tomshardware.com/peripherals/usb/best-usb-chargers",
            "Source 2": "https://www.reddit.com/r/UsbCHardware/comments/1r9qnqa/can_i_have_a_good_recommendation_to_usb_c_charger/",
            "Reason": "The broad term supports a charger range page, while OEM/manufacturer terms must carry the RFQ-specific intent.",
        },
        {
            "Query": "usb hub docking station", "Scope": "Core Page", "Current URL": "/products/hubs-docking/",
            "Observed Intent": "Comparison and product selection between related but distinct device classes",
            "Observed Result Types": "Hub-vs-dock guides; product roundups; technical explanations",
            "Provisional Decision": "Keep shared parent; retain distinct hub and docking-station child clusters",
            "Standalone Page": "Yes - parent plus children", "Confidence": "High",
            "Evidence Status": "representative-live-review", "Publication Gate": "closed",
            "Source 1": "https://www.techradar.com/features/docking-stations-vs-usb-hubs-whats-the-difference",
            "Source 2": "https://www.anker.com/blogs/hubs-and-docks/whats-the-difference-between-a-hub-and-a-docking-station",
            "Reason": "Results explicitly distinguish hubs from docks, supporting a shared parent and separate product children rather than one undifferentiated page.",
        },
    ])

    pending = pd.DataFrame([
        ("hub para usb", 12100, "Product shopping in Spanish/Portuguese", "Reject from English architecture; retain for future multilingual research", "No", "High", "https://www.idealo.es/lista/123118791/hub-para-usb.html", "https://www.corsair.com/pt/pt/p/pc-components-accessories/cc-9310002-ww/corsair-internal-4-port-usb-2-hub", "Language/geography intent does not match the current English international architecture."),
        ("usb c laptop charger", 6600, "USB-C PD laptop charger product/comparison", "Content Support on charger range; no standalone page until product matrix and full Top 10 review", "No", "High", "https://www.pcworld.com/article/1915376/best-laptop-usb-c-pd-chargers.html", "https://www.chargetechlab.com/can-i-use-any-usb-c-charger-for-laptop", "Wattage and USB-PD compatibility are central; claims require SKU and compatibility evidence."),
        ("usb c phone charger", 3600, "Generic phone charger product intent", "Content Support on charger range; no standalone page", "No", "Medium", "https://www.tomshardware.com/peripherals/usb/best-usb-chargers", "", "The query is a device-use modifier of the charger category, not a clearly distinct B2B page task."),
        ("ipad charger cable", 2400, "Mixed Lightning, USB-C and retail product intent", "Keep Product Scope Hold", "No", "High", "https://www.anker.com/collections/charger-cable-for-ipad", "https://www.bestbuy.com/site/searchpage.jsp?id=pcat17071&st=ipad+charging+cables", "Results mix connector generations; final connector catalog and Lightning/MFi evidence are required."),
        ("usb c macbook charger", 1000, "MacBook-compatible USB-C PD charger product intent", "Content Support after compatibility evidence; no standalone page", "No", "High", "https://tech.yahoo.com/computing/articles/safe-third-party-usb-c-214700270.html", "https://www.walmart.com/c/kp/macbook-charger-usb-c", "Third-party USB-C charger demand exists, but model/wattage compatibility must be evidence-backed."),
        ("apple charger cable apple store", 4400, "Apple retail/navigation and exact-brand demand", "Market Intelligence Only", "No", "High", "https://www.apple.com/shop/accessories/all/charging-essentials", "", "Brand/store navigation should not be mapped into owned product pages."),
        ("apple cable and charger", 4400, "Ambiguous Apple-brand accessory shopping", "Market Intelligence Only", "No", "High", "https://www.apple.com/shop/accessories/all/charging-essentials", "", "The query is brand-led and does not establish a distinct owned-site content task."),
        ("20w apple charger", 1600, "Exact Apple-compatible/brand charger shopping", "Market Intelligence Only; retain certification boundary", "No", "High", "https://www.bestbuy.com/site/searchpage.jsp?st=apple+20w+usb-c+power+adapter", "", "Brand and safety/certification expectations make this unsuitable for automatic product-page mapping."),
    ], columns=["Query", "Search Volume", "Observed Intent", "Provisional Decision", "Standalone Page", "Confidence", "Source 1", "Source 2", "Reason"])
    pending.insert(1, "Scope", "Pending SERP")
    pending["Evidence Status"] = "representative-live-review"
    pending["Publication Gate"] = "closed"

    for frame in (pages, pending):
        frame["Checked Date"] = CHECKED_DATE
        frame["Review Method"] = "Representative live web-result inspection; not a complete localized Google Top 10 capture"

    summary = pd.DataFrame([
        ("Core pages reviewed", len(pages)),
        ("Pending keywords reviewed", len(pending)),
        ("Verified-live rows", 0),
        ("Publication-ready rows", 0),
        ("Method limitation", "Representative results only; complete localized Top 10/domain-overlap capture still required"),
    ], columns=["Metric", "Value"])

    OUTPUT.parent.mkdir(parents=True, exist_ok=True)
    with pd.ExcelWriter(OUTPUT, engine="openpyxl") as writer:
        summary.to_excel(writer, sheet_name="Summary", index=False)
        pages.to_excel(writer, sheet_name="Core_Page_SERP", index=False)
        pending.to_excel(writer, sheet_name="Pending_Keyword_SERP", index=False)

    from openpyxl import load_workbook
    from openpyxl.styles import Font, PatternFill
    workbook = load_workbook(OUTPUT)
    for sheet in workbook.worksheets:
        sheet.freeze_panes = "A2"
        sheet.auto_filter.ref = sheet.dimensions
        sheet.sheet_view.showGridLines = False
        for cell in sheet[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(fill_type="solid", fgColor="17324D")
        for column in sheet.columns:
            width = min(60, max(12, max(len(str(cell.value or "")) for cell in list(column)[:200]) + 2))
            sheet.column_dimensions[column[0].column_letter].width = width
    workbook.save(OUTPUT)
    print(OUTPUT)
    print(summary.to_string(index=False))


if __name__ == "__main__":
    main()
