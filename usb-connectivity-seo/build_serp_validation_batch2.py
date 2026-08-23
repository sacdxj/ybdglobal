from __future__ import annotations

from pathlib import Path
import re
import pandas as pd


ROOT = Path(__file__).resolve().parents[1]
AUDIT = ROOT / "outputs" / "usb-seo" / "usb_manual_review_audit_v1.xlsx"
BATCH1 = ROOT / "outputs" / "usb-seo" / "usb_serp_validation_batch1.xlsx"
OUTPUT = ROOT / "outputs" / "usb-seo" / "usb_serp_validation_batch2.xlsx"
COMBINED = ROOT / "outputs" / "usb-seo" / "usb_serp_validation_pending_complete.xlsx"
CHECKED_DATE = "2026-08-11"


def classify(keyword: str) -> dict[str, str]:
    k = keyword.lower()
    base = {
        "Standalone Page": "No", "Evidence Status": "representative-live-review",
        "Publication Gate": "closed", "Checked Date": CHECKED_DATE,
        "Review Method": "Representative live web-result inspection; not a complete localized Google Top 10 capture",
    }
    if re.search(r"\bapple\b", k) and not re.search(r"macbook|ipad", k):
        return base | {
            "Observed Intent": "Apple-brand accessory shopping/navigation",
            "Provisional Decision": "Market Intelligence Only",
            "Approved URL": "", "Confidence": "High",
            "Source 1": "https://www.apple.com/shop/accessories/all",
            "Source 2": "https://www.bestbuy.com/site/searchpage.jsp?id=pcat17071&st=apple+charger+cable",
            "Reason": "Results are brand/store-led and should not be mapped into owned product pages.",
            "Evidence Needed": "None",
        }
    if "ipad" in k:
        return base | {
            "Observed Intent": "Mixed iPad cable shopping across USB-C and Lightning generations",
            "Provisional Decision": "Keep Product Scope Hold",
            "Approved URL": "", "Confidence": "High",
            "Source 1": "https://www.anker.com/collections/charger-cable-for-ipad",
            "Source 2": "https://www.bestbuy.com/site/shop/apple-ipad-cable",
            "Reason": "Connector generation is ambiguous; Lightning/MFi and final connector catalog remain unresolved.",
            "Evidence Needed": "Product Catalog + Certification",
        }
    if re.search(r"hub.*(?:mac|macbook)|(?:mac|macbook).*hub", k):
        return base | {
            "Observed Intent": "Mac/MacBook-compatible USB-C hub product shopping",
            "Provisional Decision": "Content Support after compatibility evidence",
            "Approved URL": "/products/hubs-docking/", "Confidence": "High",
            "Source 1": "https://www.lention.com/products/lention-usb-c-hub-cb-ce18",
            "Source 2": "https://www.cdw.com/product/tripp-lite-usb-c-multiport-hub-4-port-x-x2-usb-a-ports-and-x2-usb-c-ports/4562770",
            "Reason": "Device modifier refines the hub use case but does not establish a distinct page task.",
            "Evidence Needed": "Port Matrix + Compatibility Test",
        }
    if re.search(r"macbook|\bmac\b|chromebook|laptop", k):
        source = "https://www.dell.com/en-us/shopping/usb-c-laptop-chargers"
        if "macbook" in k or re.search(r"\bmac\b", k):
            source = "https://www.macworld.com/article/819438/best-macbook-usb-c-charger.html"
        elif "chromebook" in k:
            source = "https://www.chargetechlab.com/best-usb-c-charger-for-chromebook"
        return base | {
            "Observed Intent": "Device-compatible USB-C PD charger product/comparison",
            "Provisional Decision": "Content Support after compatibility evidence",
            "Approved URL": "/products/chargers/", "Confidence": "High",
            "Source 1": source,
            "Source 2": "https://www.startech.com/en-us/computer-parts/wch1c",
            "Reason": "Wattage, PD revision and model compatibility are central; the modifier belongs on the charger range rather than a new URL.",
            "Evidence Needed": "SKU Power Matrix + Compatibility Test",
        }
    if "phone" in k:
        return base | {
            "Observed Intent": "Ambiguous phone charging cable/adapter shopping",
            "Provisional Decision": "Product Scope Hold; content support only after connector confirmation",
            "Approved URL": "", "Confidence": "Medium",
            "Source 1": "https://www.bestbuy.ca/en-ca/shop/cell-phones-plans-accessories/best-phone-charger-cable",
            "Source 2": "https://www.lenovo.com/buy/us/en/type-c-phone-charger-0amz00a",
            "Reason": "The wording mixes cable and power-adapter intent and does not specify USB-C versus other connectors.",
            "Evidence Needed": "Connector Catalog + SERP Top 10",
        }
    raise ValueError(f"No review rule for {keyword}")


def style(path: Path) -> None:
    from openpyxl import load_workbook
    from openpyxl.styles import Font, PatternFill
    workbook = load_workbook(path)
    for sheet in workbook.worksheets:
        sheet.freeze_panes = "A2"
        sheet.auto_filter.ref = sheet.dimensions
        sheet.sheet_view.showGridLines = False
        for cell in sheet[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(fill_type="solid", fgColor="17324D")
        for column in sheet.columns:
            width = min(60, max(12, max(len(str(cell.value or "")) for cell in list(column)[:200]) + 2))
            sheet.column_dimensions[column[0].column_letter].width = width
    workbook.save(path)


def main() -> None:
    all_pending = pd.read_excel(AUDIT, sheet_name="Pending_SERP")
    batch1 = pd.read_excel(BATCH1, sheet_name="Pending_Keyword_SERP")
    done = set(batch1["Query"].str.lower())
    remaining = all_pending[~all_pending["Keyword"].str.lower().isin(done)].copy()
    decisions = remaining["Keyword"].map(classify).apply(pd.Series)
    batch2 = pd.concat([
        remaining[["Keyword", "Volume", "Original Decision", "Original Intent"]].rename(columns={"Keyword": "Query", "Volume": "Search Volume"}).reset_index(drop=True),
        decisions.reset_index(drop=True),
    ], axis=1)
    summary = pd.DataFrame([
        ("Pending keywords reviewed", len(batch2)),
        ("Standalone pages approved", int(batch2["Standalone Page"].eq("Yes").sum())),
        ("Verified-live rows", int(batch2["Evidence Status"].eq("verified-live").sum())),
        ("Publication-ready rows", int(batch2["Publication Gate"].ne("closed").sum())),
    ], columns=["Metric", "Value"])
    with pd.ExcelWriter(OUTPUT, engine="openpyxl") as writer:
        summary.to_excel(writer, sheet_name="Summary", index=False)
        batch2.to_excel(writer, sheet_name="Pending_Keyword_SERP", index=False)
    style(OUTPUT)

    combined = pd.concat([batch1, batch2], ignore_index=True, sort=False)
    combined_summary = pd.DataFrame([
        ("Total original Pending SERP", len(all_pending)),
        ("Reviewed in batch 1", len(batch1)),
        ("Reviewed in batch 2", len(batch2)),
        ("Combined reviewed", len(combined)),
        ("Missing pending keywords", len(set(all_pending["Keyword"].str.lower()) - set(combined["Query"].str.lower()))),
        ("Standalone pages approved", int(combined["Standalone Page"].eq("Yes").sum())),
        ("Publication gate", "closed; full localized Top 10 validation still required"),
    ], columns=["Metric", "Value"])
    with pd.ExcelWriter(COMBINED, engine="openpyxl") as writer:
        combined_summary.to_excel(writer, sheet_name="Summary", index=False)
        combined.sort_values("Search Volume", ascending=False).to_excel(writer, sheet_name="All_Pending_SERP", index=False)
    style(COMBINED)
    print(OUTPUT)
    print(COMBINED)
    print(combined_summary.to_string(index=False))


if __name__ == "__main__":
    main()
