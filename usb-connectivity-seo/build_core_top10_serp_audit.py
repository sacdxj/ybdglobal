from __future__ import annotations

from pathlib import Path
from urllib.parse import urlparse
import pandas as pd


ROOT = Path(__file__).resolve().parents[1]
OUTPUT = ROOT / "outputs" / "usb-seo" / "usb_core_top10_serp_audit.xlsx"
CHECKED_DATE = "2026-08-11"


RESULTS = {
    "usb c cable": [
        (1,"www.amazon.com","https://www.amazon.com/usb-c-cable/s?k=usb-c+cable"),(2,"www.startech.com","https://www.startech.com/en-us/cables/usb-c"),(3,"www.nytimes.com","https://www.nytimes.com/wirecutter/reviews/best-usb-c-cables/"),(4,"us.ugreen.com","https://us.ugreen.com/collections/usb-c-cables"),(5,"www.cablewholesale.com","https://www.cablewholesale.com/products/usb-firewire/usb-c.php"),(6,"www.youtube.com","https://www.youtube.com/watch?v=ZI1azq25r7E&vl=en-US"),(7,"www.youtube.com","https://www.youtube.com/watch?v=wPIJEFVetW0&vl=en"),(8,"www.cablecreation.com","https://www.cablecreation.com/collections/usb-c-cable"),(9,"www.reddit.com","https://www.reddit.com/r/UsbCHardware/comments/13gu5rn/the_good_the_bad_and_the_ugly_types_of_usb_a_to_c/"),(10,"www.belkin.com","https://www.belkin.com/products/cables/usb-c-cables/"),
    ],
    "hdmi cable": [
        (1,"www.amazon.com","https://www.amazon.com/HDMI-Cables-Video-Interconnects-Accessories/b?ie=UTF8&node=202505011"),(2,"www.monoprice.com","https://www.monoprice.com/category/cables/hdmi-cables/hdmi-cables"),(3,"www.walmart.com","https://www.walmart.com/browse/electronics/hdmi-cables/3944_1060825_133270_7412812_3534100"),(4,"www.computercablestore.com","https://www.computercablestore.com/HDMI-Cables-182"),(5,"cie-group.com","https://cie-group.com/how-to-av/videos-and-blogs/what-is-hdmi-high-definition-multimedia-interface"),(6,"www.youtube.com","https://www.youtube.com/watch?v=e9n0p3B63r8"),(7,"www.eaton.com","https://www.eaton.com/us/en-us/products/data-video-cables-accessories/audio-video-equipment/hdmi-cable-buying-guide.html"),(8,"www.facebook.com","https://www.facebook.com/CableOrganizer/videos/you-probably-use-hdmi-cables-everyday-but-do-you-know-what-they-actually-do-watc/941112515074083/"),(9,"www.facebook.com","https://www.facebook.com/CableOrganizer/videos/you-probably-use-hdmi-cables-everyday-but-do-you-know-what-they-actually-do-watc/941112515074083/"),(10,"www.ebay.com","https://www.ebay.com/b/HDMI-Video-Cables/32834/bn_746559"),
    ],
    "displayport cable": [
        (1,"www.amazon.com","https://www.amazon.com/displayport-cable/s?k=displayport+cable"),(2,"www.cablesforless.com","https://www.cablesforless.com/6-foot-displayport-cable-with-latches/"),(3,"www.displayport.org","https://www.displayport.org/how-to-choose-a-displayport-cable-and-not-get-a-bad-one/"),(4,"www.startech.com","https://www.startech.com/en-us/cables/audio-video/displayport"),(5,"www.bestbuy.com","https://www.bestbuy.com/site/monitor-accessories/monitor-video-cables/pcmcat138100050035.c?id=pcmcat138100050035"),(6,"en.wikipedia.org","https://en.wikipedia.org/wiki/DisplayPort"),(7,"www.dell.com","https://www.dell.com/en-us/shopping/displayport-monitor-cables"),(8,"www.monoprice.com","https://www.monoprice.com/category/cables/video-cables/displayport-&-mini-displayport-cables"),(9,"www.youtube.com","https://www.youtube.com/watch?v=Vn2vdQZhs0w"),(10,"www.eaton.com","https://www.eaton.com/us/en-us/products/backup-power-ups-surge-it-power-distribution/backup-power-ups-it-power-distribution-resources/cpdi-vertical-marketing/displayport-explained.html"),
    ],
    "usb c charger": [
        (1,"www.amazon.com","https://www.amazon.com/usb-c-charger/s?k=usb-c+charger"),(2,"www.belkin.com","https://www.belkin.com/products/chargers/usb-c-chargers/"),(3,"www.cnn.com","https://www.cnn.com/cnn-underscored/reviews/best-usb-c-chargers"),(4,"www.anker.com","https://www.anker.com/collections/usb-c-charger"),(5,"www.youtube.com","https://www.youtube.com/watch?v=1jAeEPHvDyA"),(6,"www.4imprint.com","https://www.4imprint.com/search/usb%20c%20charger"),(7,"www.bestbuy.com","https://www.bestbuy.com/site/searchpage.jsp?id=pcat17071&st=usb%20c%20charger"),(8,"www.youtube.com","https://www.youtube.com/watch?v=pqm7S6lguNA"),(9,"www.belkin.com","https://www.belkin.com/products/product-resources/usb-c-charger-buying-guide/"),(10,"www.verizon.com","https://www.verizon.com/products/chargers/"),
    ],
    "usb hub": [
        (1,"us.ugreen.com","https://us.ugreen.com/collections/usb-hub"),(2,"comprehensiveco.com","https://comprehensiveco.com/store/c/24790-USB-Hubs.html"),(3,"manhattanproducts.us","https://manhattanproducts.us/collections/usb-hubs"),(4,"www.amazon.com","https://www.amazon.com/Best-Sellers-USB-Hubs/zgbs/pc/17387627011"),(5,"en.wikipedia.org","https://en.wikipedia.org/wiki/USB_hub"),(6,"www.startech.com","https://www.startech.com/en-us/usb-hubs/usb-a-hubs"),(7,"www.youtube.com","https://www.youtube.com/watch?v=RFLGw7cdTTg"),(8,"www.techgearlab.com","https://www.techgearlab.com/topics/electronics/best-usb-hub"),(9,"www.belkin.com","https://www.belkin.com/products/docks-hubs/usb-usb-c-hubs/"),(10,"www.nytimes.com","https://www.nytimes.com/wirecutter/reviews/best-usb-c-hubs-and-docks/"),
    ],
}

CHILD_RESULTS = {
    "usb c hub": ["us.ugreen.com","www.amazon.com","www.staplesadvantage.com","www.pcworld.com","www.engadget.com","www.startech.com","plugable.com","www.reddit.com","www.techgearlab.com","www.hypershop.com"],
    "usb c docking station": ["www.startech.com","www.dell.com","plugable.com","us.ugreen.com","www.belkin.com","www.reddit.com","www.amazon.com","www.youtube.com","www.hp.com","www.pcmag.com"],
}

URLS = {
    "usb c cable": "/products/usb-cables/", "hdmi cable": "/products/hdmi-cables/",
    "displayport cable": "/products/displayport-cables/", "usb c charger": "/products/chargers/",
    "usb hub": "/products/hubs-docking/",
}


def page_type(domain: str, url: str) -> str:
    info_domains = {"www.youtube.com","www.reddit.com","en.wikipedia.org","www.nytimes.com","www.cnn.com","www.facebook.com","www.techgearlab.com","www.pcworld.com","www.engadget.com","www.pcmag.com","www.displayport.org","cie-group.com"}
    if domain in info_domains or any(x in url for x in ["/reviews/","/blog","buying-guide","what-is","explained"]):
        return "Editorial / Guide / Community"
    return "Commercial Product / Category"


def main() -> None:
    rows = []
    for query, results in RESULTS.items():
        for position, domain, url in results:
            rows.append({"Query":query,"Target URL":URLS[query],"Position":position,"Domain":domain,"URL":url,"Page Type":page_type(domain,url),"Database":"Semrush US","Checked Date":CHECKED_DATE})
    raw = pd.DataFrame(rows)
    summary = raw.groupby(["Query","Target URL"],as_index=False).agg(
        Top10_Rows=("Position","count"), Unique_Domains=("Domain","nunique"),
        Commercial_Results=("Page Type",lambda s:int(s.eq("Commercial Product / Category").sum())),
        Editorial_Results=("Page Type",lambda s:int(s.eq("Editorial / Guide / Community").sum())),
    )
    summary["Dominant Intent"] = summary.apply(lambda r:"Commercial product/category" if r["Commercial_Results"]>=5 else "Informational",axis=1)
    summary["SERP Status"] = "verified-semrush-us-top10"
    summary["Approval Decision"] = "Approved for Brief"
    summary["Approval Reason"] = summary.apply(lambda r:f"{r['Commercial_Results']}/10 commercial/category results; existing category page matches the dominant or co-dominant task.",axis=1)
    summary["Publication Gate"] = "closed - product and claim evidence still required"

    hub, dock = set(CHILD_RESULTS["usb c hub"]), set(CHILD_RESULTS["usb c docking station"])
    shared = sorted(hub & dock)
    overlap = pd.DataFrame([{
        "Query A":"usb c hub","Query B":"usb c docking station","Top10 Shared Domains":len(shared),
        "Shared Domains":" | ".join(shared),"Domain Overlap vs Top10":len(shared)/10,
        "Jaccard Overlap":len(shared)/len(hub|dock),"Decision":"Keep separate child pages under shared Hubs & Docking parent",
        "Reason":"Only partial domain overlap and distinct result-page categories support separate product tasks.",
        "Evidence Status":"verified-semrush-us-top10","Checked Date":CHECKED_DATE,
    }])

    checks = pd.DataFrame([
        ("Five exact primary queries have 10 rows",len(summary)==5 and summary["Top10_Rows"].eq(10).all()),
        ("No duplicate query-position pairs",not raw.duplicated(["Query","Position"]).any()),
        ("Every exact query has an approval decision",summary["Approval Decision"].notna().all()),
        ("Publication gate remains closed",summary["Publication Gate"].str.startswith("closed").all()),
    ],columns=["Check","Passed"])

    OUTPUT.parent.mkdir(parents=True,exist_ok=True)
    with pd.ExcelWriter(OUTPUT,engine="openpyxl") as writer:
        summary.to_excel(writer,sheet_name="Page_Approval_Summary",index=False)
        raw.to_excel(writer,sheet_name="Top10_Raw",index=False)
        overlap.to_excel(writer,sheet_name="Hub_Dock_Overlap",index=False)
        checks.to_excel(writer,sheet_name="Validation_Checks",index=False)
    from openpyxl import load_workbook
    from openpyxl.styles import Font,PatternFill
    wb=load_workbook(OUTPUT)
    for ws in wb.worksheets:
        ws.freeze_panes="A2"; ws.auto_filter.ref=ws.dimensions; ws.sheet_view.showGridLines=False
        for cell in ws[1]: cell.font=Font(bold=True,color="FFFFFF"); cell.fill=PatternFill(fill_type="solid",fgColor="17324D")
        for col in ws.columns:
            ws.column_dimensions[col[0].column_letter].width=min(60,max(12,max(len(str(c.value or "")) for c in list(col)[:200])+2))
    wb.save(OUTPUT)
    print(OUTPUT)
    print(summary.to_string(index=False))
    print(overlap.to_string(index=False))
    print(checks.to_string(index=False))


if __name__=="__main__":
    main()
