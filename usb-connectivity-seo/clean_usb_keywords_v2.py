from pathlib import Path
import re
import runpy
import pandas as pd

ROOT = Path(__file__).resolve().parents[1]
SOURCE = Path(r"C:\Users\Administrator\Downloads\USB charging cable.xlsx")
OUTPUT = ROOT / "outputs" / "usb-seo" / "usb_keyword_cleaning_v2.xlsx"

BRANDS = {
    "anker", "ugreen", "belkin", "baseus",
    "satechi", "startech", "insignia", "rocketfish", "onn", "amazon basics", "asus",
    "dell", "hp", "lenovo", "samsung", "sony", "nintendo", "xbox", "ps4", "ps5",
    "vava", "novoo", "hyperdrive", "cable matters", "j5create", "targus", "caldigit",
    "acer", "alogic", "c2g", "griffin", "omnicharge", "unitek", "jackery", "eleks",
    "lg", "psvr2", "fire stick", "kindle",
    "vizio", "roku", "dewalt", "cvs", "jackery", "bonvoisin", "slimq",
    "wii", "wii u", "ps3", "gopro", "garmin", "fitbit", "nikon", "coolpix",
    "blue yeti", "oculus", "meta quest", "quest 2", "quest 3",
}
DEVICE_PLATFORMS = re.compile(
    r"\b(apple|iphone|ipad|macbook|mac|chromebook|pixel|laptop|notebook|tablet|phone)\b",
    re.I,
)

OUT_OF_SCOPE = re.compile(
    r"\b(lightning|micro usb|mini usb|ethernet|rj45 cable|coaxial|aux cable|vga|dvi|sata|"
    r"printer cable|extension cord|jumper cable|car battery|ev charger|tesla|solar|audio cable|"
    r"guitar cable|telephone cable|fiber internet|network cable|display cable replacement|"
    r"power bank|portable charger|car charger|wireless charger|charging station|cable organizer|"
    r"mini displayport|mini hdmi|micro hdmi|printer|digital camera|camera usb|rca|vga|dvi|"
    r"car charging|car radio|car stereo|car audio|car play|carplay|electric car|electric vehicle|cigarette lighter|"
    r"soundbar|surround sound|optical sound|sound cable|hdmi arc|arc cable|cable tester|screen mirror|3 in 1 usb cable|android auto|car|lab scale)\b",
    re.I,
)
RETAIL = re.compile(r"\b(near me|nearby|amazon|walmart|best buy|target|costco|dollar tree|dollar general|home depot|lowes|ebay)\b", re.I)
TROUBLE = re.compile(r"\b(not working|no signal|problem|issue|repair|replacement|fix|driver|manual|pinout|wiring diagram|reddit|bad|broken|firmware|update|setup|off|overclock|fold|untangle|hide cables?)\b", re.I)
SHOPPING = re.compile(r"\b(best|review|reviews|top rated|cheap|cheapest|deal|deals|coupon)\b", re.I)
CONVERSION = re.compile(
    r"\b(adapters?|converters?|kvm|switch(?:es|er|ing|able)?|splitters?|extenders?|capture cards?|receptacles?)\b|"
    r"(?:hdmi.{0,18}(?:display\s*port|dp)|(?:display\s*port|dp).{0,18}hdmi)|"
    r"(?:usb(?:\s*[- ]?c)?.{0,18}(?:hdmi|display\s*port)|(?:hdmi|display\s*port).{0,18}usb)|"
    r"(?:rca.{0,18}hdmi|hdmi.{0,18}rca)",
    re.I,
)
ACCESSORY = re.compile(r"\b(kvm|switch(?:es|er|ing|able)?|splitters?|extenders?|capture cards?|receptacles?)\b", re.I)
INFO = re.compile(r"^(what|why|how|does|do|can|is|are|which|where)\b|\b(vs|versus|meaning|explained|guide|review|reviews|best)\b", re.I)
PROCUREMENT = re.compile(r"\b(manufacturer|supplier|factory|wholesale|wholesaler|bulk|oem|odm|custom|private label|importer|distributor)\b", re.I)
SPEC = re.compile(
    r"\b(20|30|45|60|65|67|100|120|130|140|150|200|240)\s*(?:w|watt|watts)\b|"
    r"\b(5|10|20|40|80)\s*gbps\b|\b(4k|8k|10k|16k)\b|"
    r"\b(60|120|144|165|240)\s*hz\b|\b(7|8|10|12)\s*[- ]?in[- ]?1\b|"
    r"\b(hdmi\s*1\.4|hdmi\s*2\.0|hdmi\s*2\.1|dp\s*1\.2|dp\s*1\.4|dp\s*2\.0|dp\s*2\.1|displayport\s*1\.2|displayport\s*1\.4|displayport\s*2\.1|usb4|fiber|optical|aoc|dual monitor|triple display|pd charging|powered usb hub)\b",
    re.I,
)


def normalize(value: str) -> str:
    value = str(value).lower().replace("–", "-").replace("—", "-")
    value = re.sub(r"[^a-z0-9.+-]+", " ", value)
    return re.sub(r"\s+", " ", value).strip()


def has_brand(k: str) -> bool:
    return any(re.search(rf"\b{re.escape(brand)}\b", k) for brand in BRANDS)


def family(k: str) -> str:
    if re.search(r"\b(docking station|usb c hub|usb-c hub|usb hub|multiport adapter)\b", k):
        return "Hubs & Docking"
    if re.search(r"\b(displayport|dp cable|dp 1\.4|dp 2\.0|dp 2\.1)\b", k):
        return "DisplayPort Cables"
    if "hdmi" in k:
        return "HDMI Cables"
    if re.search(r"\b(usb c charger|usb-c charger|pd charger|gan charger|wall charger|fast charger|power adapter|charger manufacturer|charger supplier)\b", k):
        return "Chargers"
    if re.search(r"\b(usb4|usb 4|usb 2(?:\.0)?|usb 3(?:\.0|\.1|\.2)?|usb[ -]?a|usb[ -]?b|usb[ -]?c|type[ -]?a|type[ -]?b|type[ -]?c)\b.*\bcable\b", k) or re.search(r"\b(usb cable|usb extension cable|charging cable|data cable|sync cable)\b", k):
        return "USB Cables"
    if re.search(r"\b(electronics wholesaler|electronics importer|accessories distributor|electronics brand|amazon seller|shopify electronics|electronics retailer|private label electronics)\b", k):
        return "Buyer Solutions"
    return ""


def classify(k: str) -> tuple[str, str, str, str]:
    fam = family(k)
    if OUT_OF_SCOPE.search(k):
        return "discard_noise", "Discard", "", "outside phase-one product boundary"
    if fam == "Buyer Solutions":
        return "core_keep", "Buyer Solution", fam, "approved B2B buyer-type demand"
    if RETAIL.search(k):
        return "discard_noise", "Discard", fam, "retailer/local navigation intent"
    if TROUBLE.search(k):
        return "discard_noise", "Discard", fam, "consumer troubleshooting or replacement intent"
    if SHOPPING.search(k):
        return "content_support", "Content Support Only", fam, "consumer comparison demand retained for guides, FAQs and selection criteria; no standalone URL"
    if ACCESSORY.search(k):
        return "discard_noise", "Discard", fam, "switching or accessory hardware outside phase-one product scope"
    if CONVERSION.search(k) and fam != "Hubs & Docking" and not re.search(r"\b(multiport adapter|usb c adapter|type c adapter)\b", k):
        return "discard_noise", "Discard", fam, "conversion, switching or accessory intent outside straight product scope"
    if DEVICE_PLATFORMS.search(k):
        return "review", "Device Compatibility Review", fam, "device/platform compatibility intent requires product-scope and SERP review"
    if has_brand(k):
        return "review", "Brand/Competitor Review", fam, "brand/model query requires separate comparison policy"
    if not fam:
        return "discard_noise", "Discard", "", "no clear match to approved product or buyer scope"
    if fam == "USB Cables" and re.search(r"\b(usb extension cable|usb[ -]?[ab] cable|usb [23](?:\.0|\.1|\.2)? cable)\b", k):
        return "review", "Connector Normalization Review", fam, "recognized USB cable expression; exact connector and phase-one product scope require review"
    if INFO.search(k):
        return "expand", "Informational", fam, "relevant educational or comparison demand"
    if PROCUREMENT.search(k):
        return "core_keep", "Procurement", fam, "manufacturer, wholesale or customization demand"
    if SPEC.search(k):
        return "core_keep", "Specification", fam, "in-scope performance or construction requirement"
    if re.search(r"\b(for gaming|for tv|for monitor|for projector|for laptop|gaming|monitor|projector|long distance)\b", k):
        return "expand", "Application", fam, "in-scope application demand; page split requires review"
    return "core_keep", "Product", fam, "in-scope product demand"


def main():
    source = pd.read_excel(SOURCE, sheet_name="Sheet1")
    source = source.dropna(subset=["Keyword"]).copy()
    source["keyword_norm"] = source["Keyword"].map(normalize)
    source = source.sort_values(["keyword_norm", "Volume"], ascending=[True, False]).drop_duplicates("keyword_norm")
    source["source_type"] = "measured_source"

    seeds = runpy.run_path(str(ROOT / "build_usb_keyword_sitemap.py"))["SEEDS"]
    existing = set(source["keyword_norm"])
    manual = []
    for seed in seeds:
        key = normalize(seed)
        if key not in existing:
            manual.append({
                "Keyword": seed, "Intent": "Manual seed", "Volume": 0, "Trend": "",
                "Keyword Difficulty": pd.NA, "CPC (USD)": pd.NA, "Competitive Density": pd.NA,
                "SERP Features": "", "Number of Results": pd.NA,
                "keyword_norm": key, "source_type": "manual_seed",
            })
    if manual:
        source = pd.concat([source, pd.DataFrame(manual)], ignore_index=True)

    decisions = source["keyword_norm"].map(classify)
    source[["decision", "intent_class", "product_family", "decision_reason"]] = pd.DataFrame(decisions.tolist(), index=source.index)
    source.loc[(source["source_type"] == "manual_seed") & source["decision"].eq("core_keep"), "decision"] = "manual_seed"
    source["serp_validation_required"] = (
        source["decision"].eq("review")
        | source["intent_class"].isin(["Application", "Informational"])
        | source["Volume"].fillna(0).ge(1000)
    )

    retained = source[source["decision"].isin(["core_keep", "expand", "manual_seed"])].copy()
    review = source[source["decision"].eq("review")].copy()
    content_support = source[source["decision"].eq("content_support")].copy()
    discarded = source[source["decision"].eq("discard_noise")].copy()

    summary = pd.DataFrame([
        {"metric": "measured_source_rows", "value": int((source.source_type == "measured_source").sum())},
        {"metric": "manual_seed_rows", "value": int((source.source_type == "manual_seed").sum())},
        {"metric": "core_keep", "value": int((source.decision == "core_keep").sum())},
        {"metric": "manual_seed", "value": int((source.decision == "manual_seed").sum())},
        {"metric": "expand", "value": int((source.decision == "expand").sum())},
        {"metric": "review", "value": int((source.decision == "review").sum())},
        {"metric": "content_support", "value": int((source.decision == "content_support").sum())},
        {"metric": "discard_noise", "value": int((source.decision == "discard_noise").sum())},
        {"metric": "retained_total", "value": len(retained)},
    ])
    intent_summary = source.groupby(["decision", "intent_class"], dropna=False).agg(
        keyword_count=("Keyword", "size"), total_volume=("Volume", "sum")
    ).reset_index().sort_values(["decision", "total_volume"], ascending=[True, False])
    family_summary = retained.groupby(["product_family", "intent_class"], dropna=False).agg(
        keyword_count=("Keyword", "size"), total_volume=("Volume", "sum"),
        sample_keywords=("Keyword", lambda s: " | ".join(s.head(15).astype(str)))
    ).reset_index().sort_values(["product_family", "total_volume"], ascending=[True, False])

    OUTPUT.parent.mkdir(parents=True, exist_ok=True)
    with pd.ExcelWriter(OUTPUT, engine="openpyxl") as writer:
        summary.to_excel(writer, sheet_name="Summary", index=False)
        source.to_excel(writer, sheet_name="Keyword_Review", index=False)
        retained.to_excel(writer, sheet_name="Retained_Keywords", index=False)
        review.to_excel(writer, sheet_name="Brand_Ambiguous_Review", index=False)
        content_support.to_excel(writer, sheet_name="Content_Support", index=False)
        discarded.to_excel(writer, sheet_name="Discarded_Keywords", index=False)
        intent_summary.to_excel(writer, sheet_name="Intent_Summary", index=False)
        family_summary.to_excel(writer, sheet_name="Family_Intent_Summary", index=False)

    from openpyxl import load_workbook
    from openpyxl.styles import Font, PatternFill
    wb = load_workbook(OUTPUT)
    for ws in wb.worksheets:
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions
        ws.sheet_view.showGridLines = False
        for cell in ws[1]:
            cell.font = Font(name=cell.font.name, size=cell.font.size, bold=True, color="FFFFFF")
            cell.fill = PatternFill(fill_type="solid", fgColor="17324D")
        for col in ws.columns:
            width = min(48, max(11, max(len(str(c.value or "")) for c in list(col)[:400]) + 2))
            ws.column_dimensions[col[0].column_letter].width = width
    wb.save(OUTPUT)
    print(OUTPUT)
    print(summary.to_string(index=False))
    print(family_summary.to_string(index=False))


if __name__ == "__main__":
    main()
