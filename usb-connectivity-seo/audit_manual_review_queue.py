from __future__ import annotations

from datetime import date
from pathlib import Path
import re

import pandas as pd


ROOT = Path(__file__).resolve().parents[1]
SOURCE = ROOT / "outputs" / "usb-seo" / "usb_keyword_cleaning_v2.xlsx"
OUTPUT = ROOT / "outputs" / "usb-seo" / "usb_manual_review_audit_v1.xlsx"
REVIEW_DATE = date(2026, 8, 11).isoformat()


URLS = {
    "USB Cables": "/products/usb-cables/",
    "HDMI Cables": "/products/hdmi-cables/",
    "DisplayPort Cables": "/products/displayport-cables/",
    "Hubs & Docking": "/products/hubs-docking/",
    "Chargers": "/products/chargers/",
}


def contains(k: str, pattern: str) -> bool:
    return bool(re.search(pattern, k, re.I))


def proposed_url(k: str, family: str) -> str:
    if family == "USB Cables":
        if contains(k, r"(?:usb[ -]?a|type[ -]?a).{0,18}(?:usb[ -]?c|type[ -]?c)|(?:usb[ -]?c|type[ -]?c).{0,18}(?:usb[ -]?a|type[ -]?a)"):
            return "/products/usb-cables/usb-a-to-usb-c/"
        if contains(k, r"(?:usb[ -]?c|type[ -]?c).{0,18}(?:usb[ -]?c|type[ -]?c)"):
            return "/products/usb-cables/usb-c-to-usb-c/"
        if contains(k, r"data|sync|transfer"):
            return "/products/usb-cables/data-sync/"
        if contains(k, r"usb ?4|40\s*gbps|80\s*gbps"):
            return "/products/usb-cables/usb4/"
    return URLS.get(family, "")


def audit(row: pd.Series) -> dict[str, str]:
    k = str(row["keyword_norm"])
    original = str(row["decision"])
    intent = str(row.get("intent_class") or "")
    family = "" if pd.isna(row.get("product_family")) else str(row.get("product_family"))
    reason = str(row.get("decision_reason") or "")

    result = {
        "Manual Review Status": "Reviewed",
        "Manual Decision": "Reject",
        "Approved Intent": "Navigational / Excluded",
        "Approved Family": family,
        "Approved URL": "",
        "Standalone Page": "No",
        "Review Reason": reason,
        "Evidence Needed": "None",
        "Review Confidence": "High",
    }

    # High-confidence recoveries from false-discard patterns.
    if contains(k, r"\busb data (?:transfer|sync) cable\b"):
        result.update({
            "Manual Decision": "Keep", "Approved Intent": "Product",
            "Approved Family": "USB Cables", "Approved URL": "/products/usb-cables/data-sync/",
            "Review Reason": "Direct USB data-transfer cable demand; previous no-match decision was a false discard.",
        })
        return result
    if contains(k, r"^usb hubs?$"):
        result.update({
            "Manual Decision": "Keep", "Approved Intent": "Product",
            "Approved Family": "Hubs & Docking", "Approved URL": "/products/hubs-docking/",
            "Review Reason": "Unambiguous USB hub category demand; map to the existing hub range page.",
        })
        return result
    if contains(k, r"^types of usb cables$"):
        result.update({
            "Manual Decision": "Content Support", "Approved Intent": "Informational",
            "Approved Family": "USB Cables", "Approved URL": "/resources/usb-cable-technical-guide/",
            "Review Reason": "Useful selection/education topic; retain as guide coverage without a standalone URL.",
        })
        return result

    if intent == "Connector Normalization Review":
        clear_pair = contains(k, r"(?:usb[ -]?a|type[ -]?a).{0,18}(?:usb[ -]?c|type[ -]?c)|(?:usb[ -]?c|type[ -]?c).{0,18}(?:usb[ -]?a|type[ -]?a)|(?:usb[ -]?c|type[ -]?c).{0,18}(?:usb[ -]?c|type[ -]?c)")
        if clear_pair:
            result.update({
                "Manual Decision": "Keep", "Approved Intent": "Product", "Approved Family": "USB Cables",
                "Approved URL": proposed_url(k, "USB Cables"),
                "Review Reason": "Connector pair is identifiable after normalization; word order does not create a separate product page.",
            })
        else:
            result.update({
                "Manual Review Status": "Escalated", "Manual Decision": "Product Scope Hold",
                "Approved Intent": "Product", "Approved Family": "USB Cables",
                "Approved URL": proposed_url(k, "USB Cables"), "Standalone Page": "No",
                "Review Reason": "USB cable demand is relevant, but connector endpoints/protocol or supplied product scope is not explicit.",
                "Evidence Needed": "Product Catalog", "Review Confidence": "Medium",
            })
        return result

    if intent == "Device Compatibility Review":
        if contains(k, r"lightning|apple watch|iphone"):
            result.update({
                "Manual Review Status": "Escalated", "Manual Decision": "Product Scope Hold",
                "Approved Intent": "Application / Compatibility", "Approved Family": family,
                "Review Reason": "Apple-device wording may imply Lightning, MFi or model-specific compatibility outside the confirmed phase-one scope.",
                "Evidence Needed": "Product Catalog + Certification + Compatibility Test", "Review Confidence": "High",
            })
        elif family:
            result.update({
                "Manual Decision": "Content Support", "Approved Intent": "Application / Compatibility",
                "Approved Family": family, "Approved URL": proposed_url(k, family),
                "Review Reason": "Device name expresses a compatibility use case; retain for product-page guidance/FAQ, not a standalone page.",
                "Evidence Needed": "Compatibility Test + SERP", "Review Confidence": "Medium",
            })
        else:
            result.update({
                "Manual Review Status": "Escalated", "Manual Decision": "Pending SERP",
                "Approved Intent": "Application / Compatibility", "Standalone Page": "Pending SERP",
                "Review Reason": "Device intent is visible but the underlying product family is ambiguous.",
                "Evidence Needed": "SERP + Product Catalog", "Review Confidence": "Low",
            })
        return result

    if intent == "Brand/Competitor Review":
        result.update({
            "Manual Decision": "Market Intelligence Only", "Approved Intent": "Navigational / Competitor",
            "Approved URL": "", "Review Reason": "Exact third-party brand/model demand should not be mapped into owned product pages.",
            "Evidence Needed": "None", "Review Confidence": "High",
        })
        return result

    if original == "content_support":
        result.update({
            "Manual Decision": "Content Support", "Approved Intent": "Informational / Comparison",
            "Approved URL": proposed_url(k, family),
            "Review Reason": "Consumer comparison wording can inform buying guides, FAQs and selection criteria without creating a page.",
        })
        return result

    if original == "discard_noise":
        if "no clear match" in reason:
            result.update({
                "Manual Review Status": "Escalated", "Manual Decision": "Pending SERP",
                "Approved Intent": "Ambiguous", "Standalone Page": "Pending SERP",
                "Review Reason": "High-volume no-match term requires visible-result inspection before final rejection.",
                "Evidence Needed": "SERP", "Review Confidence": "Low",
            })
        else:
            result.update({
                "Manual Decision": "Reject", "Approved Intent": "Excluded",
                "Review Reason": f"Confirmed exclusion at first-pass review: {reason}.",
            })
        return result

    return result


def main() -> None:
    data = pd.read_excel(SOURCE, sheet_name="Keyword_Review")
    data["Volume"] = pd.to_numeric(data["Volume"], errors="coerce").fillna(0)

    connector = data[data["intent_class"].eq("Connector Normalization Review")]
    review_top = data[data["decision"].eq("review")].sort_values("Volume", ascending=False).head(200)
    high_reject = data[data["decision"].eq("discard_noise") & data["Volume"].ge(5000)]
    high_no_match = data[data["decision_reason"].eq("no clear match to approved product or buyer scope") & data["Volume"].ge(5000)]

    queue = pd.concat([connector, review_top, high_reject, high_no_match], ignore_index=True)
    queue = queue.sort_values("Volume", ascending=False).drop_duplicates("keyword_norm").reset_index(drop=True)
    selection_sets = {
        "All Connector Review": set(connector["keyword_norm"]),
        "Top 200 Manual Review": set(review_top["keyword_norm"]),
        "Rejected Volume >= 5000": set(high_reject["keyword_norm"]),
        "No Match Volume >= 5000": set(high_no_match["keyword_norm"]),
    }
    queue["Audit Selection"] = queue["keyword_norm"].map(
        lambda k: " | ".join(name for name, values in selection_sets.items() if k in values)
    )

    audited = queue.apply(audit, axis=1, result_type="expand")
    queue = pd.concat([queue, audited], axis=1)
    queue.insert(queue.columns.get_loc("Manual Review Status"), "Original Decision", queue["decision"])
    queue.insert(queue.columns.get_loc("Manual Review Status") + 1, "Original Intent", queue["intent_class"])
    queue["Reviewer"] = "Codex local rules + semantic first-pass"
    queue["Reviewed Date"] = REVIEW_DATE

    summary = queue.groupby(["Manual Review Status", "Manual Decision"], as_index=False).agg(
        Rows=("Keyword", "size"), Total_Volume=("Volume", "sum")
    ).sort_values(["Manual Review Status", "Total_Volume"], ascending=[True, False])
    selection_summary = pd.DataFrame([
        {"Selection": name, "Rows Before Deduplication": len(values)} for name, values in selection_sets.items()
    ])
    changes = queue[queue["Manual Decision"].isin(["Keep", "Content Support"])].copy()
    pending_serp = queue[queue["Manual Decision"].eq("Pending SERP")].copy()
    scope_hold = queue[queue["Manual Decision"].eq("Product Scope Hold")].copy()
    market_intel = queue[queue["Manual Decision"].eq("Market Intelligence Only")].copy()

    OUTPUT.parent.mkdir(parents=True, exist_ok=True)
    with pd.ExcelWriter(OUTPUT, engine="openpyxl") as writer:
        summary.to_excel(writer, sheet_name="Audit_Summary", index=False)
        selection_summary.to_excel(writer, sheet_name="Selection_Summary", index=False)
        queue.to_excel(writer, sheet_name="Audit_Queue", index=False)
        changes.to_excel(writer, sheet_name="Proposed_Changes", index=False)
        pending_serp.to_excel(writer, sheet_name="Pending_SERP", index=False)
        scope_hold.to_excel(writer, sheet_name="Product_Scope_Hold", index=False)
        market_intel.to_excel(writer, sheet_name="Market_Intelligence", index=False)

    from openpyxl import load_workbook
    from openpyxl.styles import Font, PatternFill

    workbook = load_workbook(OUTPUT)
    for sheet in workbook.worksheets:
        sheet.freeze_panes = "A2"
        sheet.auto_filter.ref = sheet.dimensions
        sheet.sheet_view.showGridLines = False
        for cell in sheet[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(fill_type="solid", fgColor="17324D")
        for column in sheet.columns:
            letter = column[0].column_letter
            width = min(55, max(12, max(len(str(cell.value or "")) for cell in list(column)[:500]) + 2))
            sheet.column_dimensions[letter].width = width
    workbook.save(OUTPUT)

    print(OUTPUT)
    print(f"audited={len(queue)} proposed_changes={len(changes)} pending_serp={len(pending_serp)} scope_hold={len(scope_hold)} market_intel={len(market_intel)}")
    print(summary.to_string(index=False))


if __name__ == "__main__":
    main()
