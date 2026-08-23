from __future__ import annotations

from pathlib import Path
import pandas as pd


ROOT = Path(__file__).resolve().parents[1]
BASE = ROOT / "outputs" / "usb-seo" / "usb_keyword_cleaning_v2.xlsx"
AUDIT = ROOT / "outputs" / "usb-seo" / "usb_manual_review_audit_v1.xlsx"
OUTPUT = ROOT / "outputs" / "usb-seo" / "usb_keyword_cleaning_reviewed_v1.xlsx"


def main() -> None:
    source = pd.read_excel(BASE, sheet_name="Keyword_Review")
    audit = pd.read_excel(AUDIT, sheet_name="Audit_Queue")
    approved = audit[audit["Manual Decision"].isin(["Keep", "Content Support"])].copy()
    if approved["keyword_norm"].duplicated().any():
        raise ValueError("Approved audit overlay contains duplicate normalized keywords")

    audit_columns = [
        "keyword_norm", "Manual Review Status", "Manual Decision", "Approved Intent",
        "Approved Family", "Approved URL", "Standalone Page", "Review Reason",
        "Evidence Needed", "Review Confidence", "Reviewer", "Reviewed Date",
    ]
    source = source.merge(approved[audit_columns], on="keyword_norm", how="left", validate="one_to_one")
    keep_mask = source["Manual Decision"].eq("Keep")
    support_mask = source["Manual Decision"].eq("Content Support")

    source.loc[keep_mask, "decision"] = "core_keep"
    source.loc[keep_mask, "intent_class"] = source.loc[keep_mask, "Approved Intent"]
    source.loc[keep_mask, "product_family"] = source.loc[keep_mask, "Approved Family"]
    source.loc[keep_mask, "decision_reason"] = "approved manual audit recovery: " + source.loc[keep_mask, "Review Reason"].fillna("")
    source.loc[keep_mask, "serp_validation_required"] = source.loc[keep_mask, "Volume"].fillna(0).ge(1000)

    source.loc[support_mask, "decision"] = "content_support"
    source.loc[support_mask, "intent_class"] = "Content Support Only"
    source.loc[support_mask, "product_family"] = source.loc[support_mask, "Approved Family"]
    source.loc[support_mask, "decision_reason"] = "approved manual audit content support: " + source.loc[support_mask, "Review Reason"].fillna("")
    source.loc[support_mask, "serp_validation_required"] = source.loc[support_mask, "Evidence Needed"].fillna("").str.contains("SERP")

    retained = source[source["decision"].isin(["core_keep", "expand", "manual_seed"])].copy()
    review = source[source["decision"].eq("review")].copy()
    content_support = source[source["decision"].eq("content_support")].copy()
    discarded = source[source["decision"].eq("discard_noise")].copy()

    summary = pd.DataFrame([
        {"metric": "source_rows", "value": len(source)},
        {"metric": "audit_overlay_rows", "value": len(approved)},
        {"metric": "audit_keep_applied", "value": int(keep_mask.sum())},
        {"metric": "audit_content_support_applied", "value": int(support_mask.sum())},
        {"metric": "retained_total", "value": len(retained)},
        {"metric": "manual_review_remaining", "value": len(review)},
        {"metric": "content_support_total", "value": len(content_support)},
        {"metric": "discarded_total", "value": len(discarded)},
    ])
    decision_summary = source.groupby(["decision", "intent_class"], dropna=False, as_index=False).agg(
        keyword_count=("Keyword", "size"), total_volume=("Volume", "sum")
    ).sort_values(["decision", "total_volume"], ascending=[True, False])
    family_summary = retained.groupby(["product_family", "intent_class"], dropna=False, as_index=False).agg(
        keyword_count=("Keyword", "size"), total_volume=("Volume", "sum")
    ).sort_values(["product_family", "total_volume"], ascending=[True, False])
    evidence_inventory = pd.DataFrame([
        ("Phase-one category boundary", "available", "POSITIONING_DECISION_V2.md", "USB/USB-C/USB4, HDMI, DisplayPort, hubs/docks and chargers are in working scope."),
        ("Final SKU catalog", "missing", "DATA_REQUIREMENTS.md", "Required before resolving connector- or SKU-specific Product Scope Hold rows."),
        ("Connector/port matrix", "missing", "PRODUCT_EVIDENCE_SCHEMA.md", "USB-A, USB-B and extension-cable supply cannot be confirmed."),
        ("Lightning/MFi evidence", "missing", "POSITIONING_DECISION_V2.md", "Lightning remains excluded unless supported by final catalog and certification evidence."),
        ("Device compatibility tests", "missing", "DATA_REQUIREMENTS.md", "MacBook, iPhone, laptop and phone compatibility claims remain evidence-gated."),
        ("Prototype product data", "provisional", "astro/src/data/products.ts", "Useful for taxonomy only; proofStatus fields explicitly require RFQ/test evidence."),
    ], columns=["Evidence Area", "Status", "Source", "Finding"])

    OUTPUT.parent.mkdir(parents=True, exist_ok=True)
    with pd.ExcelWriter(OUTPUT, engine="openpyxl") as writer:
        summary.to_excel(writer, sheet_name="Summary", index=False)
        source.to_excel(writer, sheet_name="Keyword_Review", index=False)
        retained.to_excel(writer, sheet_name="Retained_Keywords", index=False)
        review.to_excel(writer, sheet_name="Brand_Ambiguous_Review", index=False)
        content_support.to_excel(writer, sheet_name="Content_Support", index=False)
        discarded.to_excel(writer, sheet_name="Discarded_Keywords", index=False)
        decision_summary.to_excel(writer, sheet_name="Intent_Summary", index=False)
        family_summary.to_excel(writer, sheet_name="Family_Intent_Summary", index=False)
        evidence_inventory.to_excel(writer, sheet_name="Evidence_Inventory", index=False)
        approved[audit_columns].to_excel(writer, sheet_name="Applied_Audit_Overlay", index=False)

    from openpyxl import load_workbook
    from openpyxl.styles import Font, PatternFill
    workbook = load_workbook(OUTPUT)
    for sheet in workbook.worksheets:
        sheet.freeze_panes = "A2"
        sheet.auto_filter.ref = sheet.dimensions
        sheet.sheet_view.showGridLines = False
        for cell in sheet[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(fill_type="solid", fgColor="17324D")
        for column in sheet.columns:
            width = min(55, max(12, max(len(str(cell.value or "")) for cell in list(column)[:500]) + 2))
            sheet.column_dimensions[column[0].column_letter].width = width
    workbook.save(OUTPUT)
    print(OUTPUT)
    print(summary.to_string(index=False))


if __name__ == "__main__":
    main()
