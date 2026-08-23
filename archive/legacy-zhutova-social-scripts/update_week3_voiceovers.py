from pathlib import Path

from openpyxl import load_workbook


WORKBOOK = Path(r"D:\zhutova 2清洗 正确\社媒\Zhutova_社媒运营计划_简洁版.xlsx")

week3_rows = [
    (
        "第3周",
        "一站式跨境 B2B 平台是什么意思？",
        "Zhutova 是什么",
        "很多人以为，一站式跨境 B2B 平台，就是一个平台上什么都有。\n\n但真正重要的，不是服务多，而是把产品供应、工厂对接、交易流程、仓储物流和履约服务连成一条完整链路。\n\nZhutova 想做的，就是让跨境生意不再东拼西凑，而是更稳定地跑起来。",
        "品牌认知、平台定位",
    ),
    (
        "第3周",
        "全品类供应不等于所有产品都值得做",
        "产品与供应机会",
        "全品类供应，不代表所有产品都值得马上做。\n\n产品能不能做，还要看市场需求、利润空间、供应稳定性、售后难度和履约成本。\n\nZhutova 提供的是更多选择，但真正重要的是从这些选择里筛出值得长期做的产品。",
        "产品判断、品类内容",
    ),
    (
        "第3周",
        "智能履约解决什么问题？",
        "供应链与履约",
        "很多跨境订单的问题，不是卖不出去，而是后面跟不上。\n\n发货慢、库存乱、补货不稳、售后处理慢，都会直接影响客户体验和复购。\n\n智能履约要解决的，就是让订单、仓储、物流和交付协同起来，让整个交易更稳定、更可控。",
        "履约能力、平台服务",
    ),
    (
        "第3周",
        "中国 vs 巴西价格差真的代表利润吗？",
        "利润判断",
        "很多人看到中国价格和巴西售价差很大，就以为利润一定很高。\n\n但中间还有运费、税费、平台费、仓储成本、退货风险和资金占用。\n\n所以价格差只是表面，真正决定你赚不赚钱的，是把完整成本和到岸成本算清楚。",
        "TikTok/Reels流量内容",
    ),
    (
        "第3周",
        "中国批发市场、工厂、展会有什么区别？",
        "中国采购现场",
        "很多第一次来中国采购的人，会把批发市场、工厂和展会看成一回事。\n\n其实这三个地方解决的问题完全不同：市场看款式，工厂看生产，展会看趋势和资源。\n\n如果搞不清这个区别，采购效率会很低，判断也容易出错。",
        "高播放量、采购现场、新鲜感内容",
    ),
]


def main():
    wb = load_workbook(WORKBOOK)
    ws = wb["05_口播文案库"]

    existing_week3_rows = []
    for row_idx in range(2, ws.max_row + 1):
        if ws.cell(row_idx, 1).value == "第3周":
            existing_week3_rows.append(row_idx)

    if existing_week3_rows:
        start_row = existing_week3_rows[0]
        for i, row in enumerate(week3_rows):
            target_row = start_row + i
            for col_idx, value in enumerate(row, start=1):
                ws.cell(target_row, col_idx).value = value
            ws.row_dimensions[target_row].height = 142
    else:
        for row in week3_rows:
            ws.append(row)
            ws.row_dimensions[ws.max_row].height = 142

    wb.save(WORKBOOK)
    print(WORKBOOK)


if __name__ == "__main__":
    main()
