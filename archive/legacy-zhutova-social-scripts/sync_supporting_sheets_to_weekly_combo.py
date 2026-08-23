from pathlib import Path

from openpyxl import load_workbook


WORKBOOK = Path(r"D:\zhutova 2清洗 正确\社媒\Zhutova_社媒运营计划_简洁版.xlsx")


def update_sheet01(ws):
    ws["B5"] = "每周发布 5 条内容，另准备一组品牌基础资料内容作为补充素材。"
    ws["B6"] = "当前每周组合以你修改后的 03 表为准：第1条实时行业热点新闻，第2条产品精选，第3条供应链与履约，第4条跨境贸易风险与踩坑，第5条案例与商务合作 vlog。"
    ws["B8"] = "让用户知道 Zhutova 不只是讲进口风险，而是既懂中国供应链、跨境履约，也能通过真实内容和案例建立合作信任。"


def update_sheet02(ws):
    ws["C2"] = "建立平台身份，避免被理解为普通采购号或避坑号。当前更适合作为基础资料内容储备，而不是每周固定必发位。"
    ws["D2"] = "平台定位\n公司介绍\n一站式跨境 B2B 平台\n连接中国制造与全球市场\nZhutova 和普通采购服务的区别\n品牌基础资料"

    ws["C3"] = "解释 Zhutova 的平台流程和运作方式。当前更适合作为基础资料内容储备或穿插发布。"
    ws["D3"] = "工厂直供\n全球贸易\n本地仓储\n智能履约\n买家/卖家流程\n订单和物流管理\n平台基础流程"

    ws["C4"] = "展示 Zhutova 的供给能力和产品机会，对应当前每周第2条“产品精选”的主要选题池。"
    ws["D4"] = "产品精选\n品类机会\n产品趋势\n全品类供应\n中国制造优势\n适合跨境 B2B 的产品\n产品从供应到销售的路径\n行业热点带出的产品机会"

    ws["C5"] = "证明平台的执行能力和供应链能力，对应当前每周第3条固定内容位。工厂实景如果重点是生产能力、质检、打包和交付稳定性，归这个栏目。"

    ws["C6"] = "用于品牌背书、合作信任和咨询转化，对应当前每周第5条“案例与商务合作 vlog”内容位。"
    ws["D6"] = "客户案例\n模拟案例\n合作伙伴\n供应链金融\n定制采购\n卖家合作\n供应商合作\n渠道合作\n合作方式与咨询入口\nvlog 形式的商务合作记录"
    ws["E6"] = "Zhutova 连接的不只是产品，也连接真实合作机会。"

    ws["C7"] = "用利润和成本话题获得播放、收藏和评论，可作为当前第4条“跨境贸易风险与踩坑”的补充切角。"
    ws["D7"] = "值不值得进口？\n中国 vs 巴西价格差\n低价不等于利润\n成本拆解\n运费税费平台费吃掉利润\n真实到岸成本\n利润被忽略的环节"

    ws["B8"] = "流量型"
    ws["C8"] = "用跨境贸易风险、错误和踩坑内容获得传播，对应当前每周第4条固定流量位。"
    ws["D8"] = "跨境贸易风险\n本周踩坑\n卖家常见错误\n新手进口误区\n供应商红旗\nMOQ 与现金流\n爆品真相\n样品没看就下单"
    ws["E8"] = "跨境贸易里，很多问题不是不会卖，而是看错了风险。"

    ws["C9"] = "用中国采购场景制造新鲜感、真实感和流量入口。当前更适合作为热点补充或 vlog 画面素材来源，而不是每周固定必发位。工厂实景如果重点是现场感、反差感和第一次看中国工厂的感受，归这个栏目。"


def update_sheet04(ws):
    rows = [
        ("第1周", "实时行业热点新闻：本周跨境平台又出了什么变化？", "产品精选：哪些品类适合跨境 B2B？", "供应链与履约：从工厂直供到智能履约怎么走？", "跨境贸易风险与踩坑：中国低价不等于真实利润", "案例与商务合作 vlog：第一次来中国采购会看到什么？"),
        ("第2周", "实时行业热点新闻：巴西市场最近什么品类更值得关注？", "产品精选：中国制造如何进入巴西市场？", "供应链与履约：本地仓储有什么价值？", "跨境贸易风险与踩坑：便宜供应商为什么可能最贵？", "案例与商务合作 vlog：模拟案例，一个卖家如何判断产品？"),
        ("第3周", "实时行业热点新闻：跨境卖家最近最该关注什么趋势？", "产品精选：全品类供应不等于所有产品都值得做", "供应链与履约：智能履约解决什么问题？", "跨境贸易风险与踩坑：中国 vs 巴西价格差真的代表利润吗？", "案例与商务合作 vlog：中国批发市场、工厂、展会有什么区别？"),
        ("第4周", "实时行业热点新闻：最近哪些跨境变化会影响卖家判断？", "产品精选：一个产品从供应到销售要看什么？", "供应链与履约：订单和物流节点如何管理？", "跨境贸易风险与踩坑：MOQ 太高为什么危险？", "案例与商务合作 vlog：同一个产品在中国为什么有很多版本？"),
    ]
    for row_idx, row in enumerate(rows, start=2):
        for col_idx, value in enumerate(row, start=1):
            ws.cell(row_idx, col_idx).value = value


def update_sheet05(ws):
    # Keep existing week content but align the "适合场景" labels and week-3 themes to new weekly-combo naming.
    for row_idx in range(2, ws.max_row + 1):
        week = ws.cell(row_idx, 1).value
        theme = ws.cell(row_idx, 2).value
        category = ws.cell(row_idx, 3).value

        if category == "产品与供应机会":
            ws.cell(row_idx, 5).value = "产品精选"
        elif category == "供应链与履约":
            ws.cell(row_idx, 5).value = "供应链与履约"
        elif category == "风险与踩坑":
            ws.cell(row_idx, 5).value = "跨境贸易风险与踩坑"
        elif category == "案例与商务合作":
            ws.cell(row_idx, 5).value = "案例与商务合作 vlog"
        elif category == "中国采购现场":
            ws.cell(row_idx, 5).value = "案例与商务合作 vlog / 真实现场素材"
        elif category == "Zhutova 是什么":
            ws.cell(row_idx, 5).value = "基本资料 / 品牌认知"
        elif category == "利润判断":
            ws.cell(row_idx, 5).value = "跨境贸易风险与踩坑"

        # Align week 3 first item away from old "品牌位" phrasing and toward current combo logic.
        if week == "第3周" and theme == "一站式跨境 B2B 平台是什么意思？":
            ws.cell(row_idx, 5).value = "基本资料 / 品牌认知"


def main():
    wb = load_workbook(WORKBOOK)
    update_sheet01(wb["01_核心思路"])
    update_sheet02(wb["02_8大栏目主题池"])
    update_sheet04(wb["04_未来4周主题安排"])
    update_sheet05(wb["05_口播文案库"])
    wb.save(WORKBOOK)
    print(WORKBOOK)


if __name__ == "__main__":
    main()
