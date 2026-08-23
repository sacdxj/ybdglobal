from pathlib import Path

from openpyxl import load_workbook


WORKBOOK = Path(r"D:\zhutova 2清洗 正确\社媒\Zhutova_社媒运营计划_简洁版.xlsx")


def main():
    wb = load_workbook(WORKBOOK)
    ws = wb["02_8大栏目主题池"]

    updates = {
        5: {
            3: "证明平台的执行能力和供应链能力。工厂实景如果重点是生产能力、质检、打包和交付稳定性，归这个栏目。",
            4: "供应商管理\n工厂实景：生产线/组装/质检/打包出货\nQC\n仓储\n物流\n订单履约\n交付稳定性\n本地仓储价值",
            5: "工厂实景不是只看热闹，重点是看供应链能不能稳定交付。",
        },
        6: {
            3: "用于品牌背书、合作信任和咨询转化。",
            4: "客户案例\n模拟案例\n合作伙伴\n供应链金融\n定制采购\n卖家合作\n供应商合作\n渠道合作\n合作方式与咨询入口",
            5: "Zhutova 连接的不只是产品，也连接合作机会。",
        },
        8: {
            3: "用风险、错误和踩坑内容获得传播，对应每周第4条流量位轮换内容之一。",
        },
        9: {
            3: "用中国采购场景制造新鲜感、真实感和流量入口。工厂实景如果重点是现场感、反差感和第一次看中国工厂的感受，归这个栏目。",
            4: "第一次来中国采购会看到什么\n工厂实景：样品间/生产线/包装区的新鲜感\n中国批发市场/工厂/展会差异\n同一产品为什么有很多版本\n中国供应商报价为什么差距大\n巴西卖家容易误判的现场细节\n中国采购不是找便宜货，而是做筛选\n第一次去市场或工厂应该看什么",
            5: "第一次来中国采购，很多人都会被这些现场细节吸引。",
        },
    }

    for row_idx, cells in updates.items():
        for col_idx, value in cells.items():
            ws.cell(row_idx, col_idx).value = value

    wb.save(WORKBOOK)
    print(WORKBOOK)


if __name__ == "__main__":
    main()
