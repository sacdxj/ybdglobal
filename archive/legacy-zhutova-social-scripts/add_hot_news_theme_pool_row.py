from pathlib import Path

from openpyxl import load_workbook


WORKBOOK = Path(r"D:\zhutova 2清洗 正确\社媒\Zhutova_社媒运营计划_简洁版.xlsx")


def main():
    wb = load_workbook(WORKBOOK)

    if "02_8大栏目主题池" in wb.sheetnames:
        ws02 = wb["02_8大栏目主题池"]
        ws02.title = "02_9大栏目主题池"
    else:
        ws02 = wb["02_9大栏目主题池"]

    ws01 = wb["01_核心思路"]

    hot_news_exists = False
    for row_idx in range(2, ws02.max_row + 1):
        if ws02.cell(row_idx, 1).value == "实时行业热点新闻":
            hot_news_exists = True
            target_row = row_idx
            break
    else:
        target_row = 2
        ws02.insert_rows(target_row)

    hot_news_row = [
        "实时行业热点新闻",
        "流量型",
        "用跨境平台政策、行业变化、市场动态和卖家关注的新闻切入，快速获得注意力，对应当前每周第1条固定内容位。",
        "跨境平台政策变化\n巴西市场新趋势\n平台新规\n行业数据变化\n卖家正在讨论的话题\n热点新闻带出的产品机会\n热点新闻带出的风险提醒",
        "跨境圈这周最值得关注的一件事，可能会直接影响卖家判断。",
    ]

    for col_idx, value in enumerate(hot_news_row, start=1):
        ws02.cell(target_row, col_idx).value = value

    ws01["B4"] = "采用 9 个大栏目：5 个平台/品牌/业务支撑类 + 4 个流量型。当前每周发布以 03 表为准，其中“实时行业热点新闻”为第1条固定内容位。"

    wb.save(WORKBOOK)
    print(WORKBOOK)


if __name__ == "__main__":
    main()
