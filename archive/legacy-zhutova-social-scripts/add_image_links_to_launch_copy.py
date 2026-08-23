from pathlib import Path
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side


ROOT = Path(r"D:\zhutova 2清洗 正确")
XLSX = ROOT / "zhutova社媒规划_新版判断品牌计划.xlsx"

image_map = {
    "2026-05-28": r"D:\zhutova 2清洗 正确\社媒\instagram_warmup_2026_05_28_31_EN\2026-05-28_warmup_low-price.png",
    "2026-05-29": r"D:\zhutova 2清洗 正确\社媒\instagram_warmup_2026_05_28_31_EN\2026-05-29_warmup_supplier-check.png",
    "2026-05-30": r"D:\zhutova 2清洗 正确\社媒\instagram_warmup_2026_05_28_31_EN\2026-05-30_warmup_moq.png",
    "2026-05-31": r"D:\zhutova 2清洗 正确\社媒\instagram_warmup_2026_05_28_31_EN\2026-05-31_warmup_launch-tomorrow.png",
    "2026-06-01": r"D:\zhutova 2清洗 正确\社媒\instagram_launch_2026_06_01_EN",
}

wb = openpyxl.load_workbook(XLSX)
ws = wb["预热与上线文案"]

headers = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
if "Post Image / Asset" not in headers:
    insert_at = 3
    ws.insert_cols(insert_at)
    ws.cell(1, insert_at, "Post Image / Asset")
else:
    insert_at = headers.index("Post Image / Asset") + 1

for row in range(2, ws.max_row + 1):
    date_value = str(ws.cell(row, 1).value)
    asset = image_map.get(date_value, "")
    cell = ws.cell(row, insert_at, asset)
    if asset:
        cell.hyperlink = asset
        cell.style = "Hyperlink"

# Restyle header and sheet.
thin = Side(style="thin", color="D9E2EC")
for row in ws.iter_rows():
    for cell in row:
        cell.alignment = Alignment(vertical="top", wrap_text=True)
        cell.border = Border(bottom=thin)
for cell in ws[1]:
    cell.font = Font(bold=True, color="FFFFFF")
    cell.fill = PatternFill("solid", fgColor="0A1628")
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

widths = {
    "A": 18,
    "B": 34,
    "C": 72,
    "D": 86,
    "E": 54,
    "F": 60,
}
for col, width in widths.items():
    ws.column_dimensions[col].width = width

try:
    wb.save(XLSX)
    print(XLSX)
except PermissionError:
    fallback = ROOT / "zhutova社媒规划_新版判断品牌计划_带图片链接.xlsx"
    wb.save(fallback)
    print(fallback)
