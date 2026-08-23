from pathlib import Path
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side


ROOT = Path(r"D:\zhutova 2清洗 正确")
PLAN = ROOT / "社媒" / "Zhutova_社媒运营计划_简洁版.xlsx"


def style(ws, widths=None):
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A2"
    thin = Side(style="thin", color="D9E2EC")
    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.border = Border(bottom=thin)
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="0A1628")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    if widths:
        for col, width in widths.items():
            ws.column_dimensions[col].width = width


wb = openpyxl.load_workbook(PLAN)

# 1) Update weekly template to make style choices explicit.
ws = wb["03_每周5条发布模板"]
ws.delete_rows(1, ws.max_row)
headers = ["每周第几条", "建议栏目", "主风格", "可选流量切口", "内容目的", "选题方式"]
rows = [
    ("第1条", "Zhutova 是什么", "品牌/商务背书", "你以为跨境采购只是找供应商？真正难的是后面的交易、仓储和履约。", "建立平台身份，避免被理解成普通内容号。", "讲清 Zhutova 是一站式跨境 B2B 平台。"),
    ("第2条", "产品与供应机会", "产品机会/半流量", "这个产品看起来有机会，但不一定适合所有卖家。", "用产品/品类机会吸引卖家和B端客户。", "从一个品类或产品机会切入，讲供应与市场可能性。"),
    ("第3条", "供应链与履约", "品牌/商务背书", "找到工厂，不代表你真的能把货稳定卖出去。", "证明平台有执行能力，不只是讲概念。", "讲工厂直供、仓储、物流、履约、QC中的一个环节。"),
    ("第4条", "卖家增长与风险判断", "流量爆点", "很多卖家进口亏钱，不是因为产品不好，而是因为算错了成本。", "用风险内容获得播放和互动。", "从利润、MOQ、供应商风险、成本误区里选一个。"),
    ("第5条", "案例与商务合作", "案例信任/可流量", "这个产品不是不能做，是不能这样做。", "沉淀信任，承接商务合作和咨询。", "做客户/模拟案例，或讲合作伙伴、商务合作方式。"),
]
ws.append(headers)
for row in rows:
    ws.append(row)
style(ws, {"A": 14, "B": 28, "C": 20, "D": 56, "E": 46, "F": 62})
for r in range(2, ws.max_row + 1):
    ws.row_dimensions[r].height = 92


# 2) Replace voiceover sheet with both brand and traffic options.
ws = wb["05_口播文案"]
ws.delete_rows(1, ws.max_row)
headers = ["栏目", "品牌/商务背书版口播", "流量爆点版口播", "适合场景"]
rows = [
    (
        "Zhutova 是什么",
        "Zhutova 是连接中国制造与全球市场的一站式跨境 B2B 平台。\n\n它连接工厂直供、全球贸易、本地仓储和智能履约，帮助卖家和合作伙伴更高效地完成跨境交易。\n\nZhutova 的价值，不只是找到一个供应商，而是把供应、交易、仓储、履约和服务连接起来。",
        "你以为跨境采购只是找供应商？\n\n真正难的是后面的交易、仓储、物流和履约。\n\n很多卖家不是找不到货，而是找到了货却无法稳定卖、稳定交付。\n\nZhutova 做的，是把中国制造和全球市场连接成一个更完整的跨境 B2B 平台。",
        "品牌置顶、平台招商、Reels、LinkedIn；流量版适合 TikTok/Reels 开头。",
    ),
    (
        "平台如何运作",
        "Zhutova 的平台能力覆盖从工厂直供到智能履约。\n\nFactory Direct Supply、Global Trading、Local Warehousing、Smart Fulfillment，是跨境 B2B 交易中需要被管理的关键链路。\n\n这就是 Zhutova 作为一站式平台的价值。",
        "找到工厂，不代表采购结束。\n\n你还要处理交易、仓储、物流、履约和售后。\n\n如果只看一个报价，后面的风险很容易被忽略。\n\nZhutova 把这些环节连接起来，让跨境采购不只是单点找货。",
        "平台能力介绍、流程图视频、H5素材改编。",
    ),
    (
        "产品与供应机会",
        "Zhutova 关注的不只是产品本身，而是产品从供应到销售的完整机会。\n\n一个产品是否值得做，需要看供应稳定性、市场需求、履约成本、销售渠道和利润空间。\n\n这也是跨境 B2B 平台需要提供的判断能力。",
        "这个产品看起来有机会，但不一定适合所有卖家。\n\n便宜、好卖、价差大，只是第一层。\n\n真正要看的是供应稳不稳、运费高不高、平台费吃不吃利润、MOQ 会不会压库存。\n\n产品机会，必须放进完整链路里判断。",
        "产品机会、品类分析、China vs Brazil、卖家教育。",
    ),
    (
        "供应链与履约",
        "供应链稳定，往往比低价更重要。\n\nZhutova 强调工厂直供、本地仓储、物流管理和智能履约，是为了让跨境交易更可执行。\n\n对卖家来说，稳定交付和服务支持，才是长期经营的基础。",
        "找到便宜产品，不代表你能稳定赚钱。\n\n如果仓储、物流和履约跟不上，便宜产品也会变成延误、差评、退货和库存压力。\n\n供应链不是后台小事，它直接决定利润和复购。\n\n这就是 Zhutova 要做全链路的原因。",
        "供应链幕后、仓储物流、履约能力、商务背书。",
    ),
    (
        "卖家增长与风险判断",
        "Zhutova 帮助卖家在进口前判断产品、供应商、MOQ、到岸成本和履约风险。\n\n风险判断不是为了阻止采购，而是为了让采购更稳、更可控。\n\n当卖家看清完整成本，才知道产品是否真的值得做。",
        "很多卖家进口亏钱，不是因为产品不好。\n\n而是因为只看了采购价。\n\n运费、税费、平台费、MOQ、退货和履约风险加起来，利润可能直接消失。\n\n进口前，先把账算清楚。",
        "TikTok/Reels 流量内容、评论互动、WhatsApp产品分析引导。",
    ),
    (
        "案例与商务合作",
        "Zhutova 连接的不只是产品。\n\n它连接供应、交易、仓储、履约、服务和合作伙伴生态。\n\n无论是卖家、品牌方、供应商还是渠道伙伴，如果需要更完整的跨境 B2B 支持，都可以和 Zhutova 探讨合作。",
        "这个产品不是不能做，是不能这样做。\n\n有些产品小批量测试可以，但直接大单就很危险。\n\n有些供应商报价低，但后面履约风险很高。\n\nZhutova 的价值，就是在执行前先帮你把这些问题看清楚。",
        "客户案例、模拟案例、商务合作、B2B线索承接。",
    ),
]
ws.append(headers)
for row in rows:
    ws.append(row)
style(ws, {"A": 24, "B": 76, "C": 76, "D": 42})
for r in range(2, ws.max_row + 1):
    ws.row_dimensions[r].height = 168


# 3) Add a lightweight hook bank so the team can choose more traffic expressions.
sheet_name = "07_流量Hook库"
if sheet_name in wb.sheetnames:
    del wb[sheet_name]
ws = wb.create_sheet(sheet_name)
headers = ["对应栏目", "流量Hook", "适合用途"]
rows = [
    ("Zhutova 是什么", "你以为跨境采购只是找供应商？", "平台认知视频开头"),
    ("Zhutova 是什么", "很多卖家不是找不到货，而是找到了货也卖不稳。", "平台价值切入"),
    ("平台如何运作", "找到工厂，不代表采购结束。", "流程/供应链视频"),
    ("平台如何运作", "一个报价解决不了跨境交易的全部问题。", "品牌背书流量化"),
    ("产品与供应机会", "这个产品看起来有机会，但不一定适合你。", "产品机会视频"),
    ("产品与供应机会", "价格差很大，不代表你一定能赚。", "China vs Brazil"),
    ("供应链与履约", "便宜产品，也可能因为履约问题变成亏损。", "供应链内容"),
    ("供应链与履约", "供应链不是后台小事，它直接决定利润。", "供应链内容"),
    ("卖家增长与风险判断", "很多卖家进口亏钱，不是因为产品不好。", "爆点流量"),
    ("卖家增长与风险判断", "第一单买太多，可能直接压死现金流。", "MOQ内容"),
    ("案例与商务合作", "这个产品不是不能做，是不能这样做。", "案例内容"),
    ("案例与商务合作", "有些合作，不是缺产品，而是缺一条能执行的链路。", "商务合作"),
]
ws.append(headers)
for row in rows:
    ws.append(row)
style(ws, {"A": 26, "B": 70, "C": 34})
for r in range(2, ws.max_row + 1):
    ws.row_dimensions[r].height = 52

wb.save(PLAN)
print(PLAN)
