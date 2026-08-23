from pathlib import Path

from openpyxl import load_workbook


WORKBOOK = Path(r"D:\zhutova 2清洗 正确\社媒\Zhutova_社媒运营计划_简洁版.xlsx")
FALLBACK_WORKBOOK = Path(r"D:\zhutova 2清洗 正确\社媒\Zhutova_社媒运营计划_简洁版_第7周口播.xlsx")
WEEK = "第7周"

NEWS_RAW = """大家好，这里是 ZHUTOVA 一分钟跨境供应链周报。

第一条，亚马逊美国和英国站继续收紧 FBM 自配送规则。
平台不仅要求备货时间必须真实，还会根据历史发货表现进行校验，并陆续增加更多履约考核。
以后拼的不只是能发货，而是能不能稳定、准时地发货。

第二条，美线海运成本再次上涨。
最新数据显示，美西航线运价大幅走高，船公司也陆续上调旺季附加费，舱位变得更加紧张。
如果近期有美国市场出货计划，建议尽早锁定舱位，并重新核算报价和交期。

以上就是本周最值得关注的两条跨境资讯。

关注 ZHUTOVA，每周一分钟，带你了解最新跨境供应链动态。"""

TOPICS = [
    "实时行业热点新闻：ZHUTOVA 一分钟跨境供应链周报",
    "产品精选：美线运费上涨后，什么产品更值得做？",
    "供应链与履约：为什么备货时间不能随便写？",
    "跨境贸易风险与踩坑：海运涨价后，不重算报价有多危险？",
    "案例与商务合作 vlog：旺季出货前，锁舱和交期要看什么？",
]

VOICEOVERS = [
    (
        WEEK,
        TOPICS[0],
        "实时行业热点新闻",
        """这周两条跨境新闻，
都在提醒卖家一件事：

未来平台看重的，
不只是你能不能发货，
而是你能不能稳定、准时地发货。

这里是 ZHUTOVA 一分钟跨境供应链周报。

第一，亚马逊美国和英国站，
继续收紧 FBM 自配送规则。

平台不仅要求备货时间必须真实，
还会根据历史发货表现进行校验，
并陆续增加更多履约考核。

也就是说，
以后不是你写几天发货就算几天，
而是平台会看你过去到底能不能做到。

第二，美线海运成本再次上涨。

美西航线运价走高，
船公司也陆续上调旺季附加费，
舱位变得更加紧张。

如果近期有美国市场出货计划，
建议尽早锁定舱位，
并重新核算报价和交期。

关注 ZHUTOVA，
每周一分钟，
带你了解最新跨境供应链动态。""",
        "实时行业热点新闻 / 供应链周报 / 60秒",
    ),
    (
        WEEK,
        TOPICS[1],
        "产品精选",
        """美线运费上涨以后，
不是所有产品都还值得做。

尤其是体积大、重量高、客单价低的产品，
很容易被运费吃掉利润。

这时候选品要更看重三个点。

第一，产品体积要可控。
第二，货值不能太低。
第三，补货节奏要稳定。

比如小型电子配件、轻量工具、收纳类小件，
如果供应稳定、包装可控，
就比大件低货值产品更容易抗住运费波动。

所以 Zhutova 做产品精选，
不是只看产品有没有需求。

还要看在当前物流成本下，
这个产品还能不能保住利润。""",
        "产品精选 / 美线运费上涨选品 / 30-45秒",
    ),
    (
        WEEK,
        TOPICS[2],
        "供应链与履约",
        """备货时间，
真的不能随便写。

以前有些卖家会把备货时间写得很短，
先接订单，
后面再慢慢安排发货。

但现在平台越来越重视真实履约。

它不只看你承诺几天发货，
还会看你过去的发货表现，
能不能长期稳定做到。

所以履约不是前台写一个时间，
后台再想办法。

它需要库存、供应商、仓储和物流，
提前配合好。

这也是 Zhutova 看供应链时，
一直强调稳定交付的原因。""",
        "供应链与履约 / FBM备货时间 / 30-45秒",
    ),
    (
        WEEK,
        TOPICS[3],
        "跨境贸易风险与踩坑",
        """海运一涨价，
最危险的不是成本变高。

而是你还在用旧报价接单。

很多卖家利润变薄，
不是因为产品卖不动，
而是报价时没有把最新运费算进去。

特别是旺季前后，
运价、附加费、舱位，
都可能快速变化。

如果你的报价周期比较长，
一定要提前预留运费波动空间。

否则订单接到了，
货也出了，
最后才发现利润被物流成本吃掉。

所以海运涨价后，
第一件事不是急着接单，
而是重新核算报价和交期。""",
        "跨境贸易风险与踩坑 / 运费报价风险 / 30-45秒",
    ),
    (
        WEEK,
        TOPICS[4],
        "案例与商务合作",
        """旺季出货前，
不要只问还能不能发货。

更要提前看三件事。

第一，舱位能不能锁定。
第二，运费有效期有多长。
第三，交期有没有缓冲时间。

很多时候，
不是货做不出来，
而是舱位紧、运费变、船期不稳，
最后影响整批订单交付。

所以 Zhutova 看出货计划，
不是只看一个物流报价。

而是会把备货、锁舱、运费、交期，
一起放进供应链判断里。

这样客户在接单和报价时，
才不会被后面的物流变化打乱节奏。""",
        "案例与商务合作 vlog / 旺季出货与锁舱 / 30-45秒",
    ),
]


def upsert_news_input(wb):
    ws = wb["07_新闻热点输入"]
    target_row = None
    for row_idx in range(2, ws.max_row + 1):
        if ws.cell(row_idx, 1).value == WEEK:
            target_row = row_idx
            break
    if target_row is None:
        target_row = ws.max_row + 1

    values = [
        WEEK,
        "ZHUTOVA 一分钟跨境供应链周报",
        "亚马逊美国和英国站继续收紧 FBM 自配送规则，备货时间真实性和历史发货表现更重要。",
        "美线海运成本再次上涨，美西航线运价走高，旺季附加费和舱位压力增加。",
        "",
        NEWS_RAW,
        "用于每周第1条：实时行业热点新闻",
        "本周新闻由用户提供，口播已做短视频化整理。",
    ]
    for col_idx, value in enumerate(values, start=1):
        ws.cell(target_row, col_idx).value = value
    ws.row_dimensions[target_row].height = 120


def upsert_week_plan(wb):
    ws = wb["04_未来4周主题安排"]
    target_row = None
    for row_idx in range(2, ws.max_row + 1):
        if ws.cell(row_idx, 1).value == WEEK:
            target_row = row_idx
            break
    if target_row is None:
        target_row = ws.max_row + 1
    for col_idx, value in enumerate([WEEK, *TOPICS], start=1):
        ws.cell(target_row, col_idx).value = value
    ws.row_dimensions[target_row].height = 98


def upsert_voiceovers(wb):
    ws = wb["05_口播文案库"]
    existing = []
    for row_idx in range(2, ws.max_row + 1):
        if ws.cell(row_idx, 1).value == WEEK:
            existing.append(row_idx)

    if existing:
        start = existing[0]
        for i, row in enumerate(VOICEOVERS):
            target = start + i
            for col_idx, value in enumerate(row, start=1):
                ws.cell(target, col_idx).value = value
            ws.row_dimensions[target].height = 170 if i == 0 else 145
    else:
        for i, row in enumerate(VOICEOVERS):
            ws.append(row)
            ws.row_dimensions[ws.max_row].height = 170 if i == 0 else 145


def main():
    wb = load_workbook(WORKBOOK)
    upsert_news_input(wb)
    upsert_week_plan(wb)
    upsert_voiceovers(wb)
    try:
        wb.save(WORKBOOK)
        print(WORKBOOK)
    except PermissionError:
        wb.save(FALLBACK_WORKBOOK)
        print(FALLBACK_WORKBOOK)


if __name__ == "__main__":
    main()
