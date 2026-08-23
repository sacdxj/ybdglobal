from pathlib import Path
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side


ROOT = Path(r"D:\zhutova 2清洗 正确")
PLAN = ROOT / "社媒" / "Zhutova_社媒运营计划_简洁版.xlsx"

wb = openpyxl.load_workbook(PLAN)
sheet_name = "02_内容策略与栏目"
if sheet_name in wb.sheetnames:
    del wb[sheet_name]

ws = wb.create_sheet(sheet_name, 1)
ws.sheet_view.showGridLines = False
ws.freeze_panes = "A2"

headers = ["模块", "类别/栏目", "目标", "内容重点", "适合形式", "备注"]
rows = [
    (
        "内容策略",
        "品牌/商务背书内容",
        "建立平台实力、商务信任和合作价值",
        "Zhutova 是什么；Zhutova 作用；工厂直供到智能履约；平台运营环境；供应链幕后流程；客户案例/模拟案例；合作伙伴和生态；商务合作机会",
        "品牌短视频、LinkedIn/Facebook 图文、Instagram Reels、平台介绍轮播、商务合作帖",
        "语气更正式，重点是可信度、平台能力和合作价值。",
    ),
    (
        "内容策略",
        "流量型内容",
        "获得播放量、评论、收藏、分享",
        "中国低价不等于利润；便宜供应商可能是最贵的错误；MOQ 会压死现金流；爆品不等于赚钱；产品看起来有利润但未必值得进口；运费、税费、平台费会吃掉利润",
        "TikTok、Instagram Reels、YouTube Shorts、Facebook 讨论帖",
        "开头要有冲突和痛点，先吸引注意，再带出 Zhutova 的判断力。",
    ),
    (
        "核心内容栏目",
        "值不值得进口？",
        "建立 Zhutova 的产品判断心智",
        "拆产品机会、中国采购价、巴西售价、真实成本、MOQ、风险，最后给出是否值得进口的判断",
        "短视频、轮播、案例帖",
        "这是最重要的长期栏目，目标是让用户形成“进口前先问 Zhutova”。",
    ),
    (
        "核心内容栏目",
        "本周踩坑",
        "用风险内容制造传播和代入",
        "供应商风险、物流风险、MOQ 风险、税费误判、爆品误判、库存积压",
        "TikTok/Reels、Facebook 讨论帖",
        "标题要尖锐，例如：便宜供应商，可能是最贵的错误。",
    ),
    (
        "核心内容栏目",
        "中国 vs 巴西价格差",
        "用价格差吸引注意，再用真实成本教育用户",
        "中国采购价、巴西销售价、运费、税费、平台费、真实利润、是否值得做",
        "Instagram Carousel、Reels、短图文",
        "适合做收藏型内容，不要只展示价差，必须加入成本和风险。",
    ),
    (
        "核心内容栏目",
        "卖家常见错误",
        "触发自我代入，提升评论和收藏",
        "只看采购价、忽略 MOQ、忘记平台费、不做样品、不算退货、不看履约风险",
        "短视频、轮播、Facebook 长文",
        "每条只讲一个错误，避免信息过载。",
    ),
    (
        "核心内容栏目",
        "Zhutova 服务",
        "让用户理解 Zhutova 能具体帮什么",
        "产品分析、供应商判断、MOQ 建议、真实到岸成本估算、采购流程支持、仓储与履约、商务合作",
        "品牌背书视频、服务介绍轮播、FAQ、WhatsApp 引导帖",
        "不要写成功能清单，要用卖家场景表达。",
    ),
    (
        "核心内容栏目",
        "供应链",
        "证明 Zhutova 的平台能力和执行能力",
        "工厂直供、全球贸易、本地仓储、智能履约、QC、订单管理、物流节点、供应链金融、合作伙伴生态",
        "幕后视频、流程图、H5素材改编、商务合作帖",
        "这是承接信任和 B2B 合作的内容，不一定追求最高播放量。",
    ),
]

ws.append(headers)
for row in rows:
    ws.append(row)

thin = Side(style="thin", color="D9E2EC")
for row in ws.iter_rows():
    for cell in row:
        cell.alignment = Alignment(vertical="top", wrap_text=True)
        cell.border = Border(bottom=thin)

for cell in ws[1]:
    cell.font = Font(bold=True, color="FFFFFF")
    cell.fill = PatternFill("solid", fgColor="0A1628")
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

widths = {
    "A": 18,
    "B": 24,
    "C": 34,
    "D": 72,
    "E": 42,
    "F": 52,
}
for col, width in widths.items():
    ws.column_dimensions[col].width = width

for r in range(2, ws.max_row + 1):
    ws.row_dimensions[r].height = 96
    if ws.cell(r, 1).value == "内容策略":
        fill = "E8F4FF"
    else:
        fill = "FFF4D6"
    for c in range(1, ws.max_column + 1):
        ws.cell(r, c).fill = PatternFill("solid", fgColor=fill)

wb.save(PLAN)
print(PLAN)
