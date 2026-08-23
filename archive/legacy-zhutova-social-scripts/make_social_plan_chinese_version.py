from pathlib import Path
from datetime import date, timedelta
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side


ROOT = Path(r"D:\zhutova 2清洗 正确")
SRC = ROOT / "zhutova社媒规划_新版判断品牌计划.xlsx"
OUT = ROOT / "zhutova社媒规划_新版判断品牌计划_中文版本.xlsx"

wb = openpyxl.load_workbook(SRC)

# Remove English-only duplicate sheet in the Chinese version.
if "Tomorrow Plan_May28_EN" in wb.sheetnames:
    del wb["Tomorrow Plan_May28_EN"]


def clear_sheet(ws):
    ws.delete_rows(1, ws.max_row)


def style_sheet(sheet, widths=None, freeze="A2"):
    sheet.freeze_panes = freeze
    sheet.sheet_view.showGridLines = False
    widths = widths or {}
    for col, width in widths.items():
        sheet.column_dimensions[col].width = width
    thin = Side(style="thin", color="D9E2EC")
    for row in sheet.iter_rows():
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.border = Border(bottom=thin)
    for cell in sheet[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="0A1628")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    sheet.row_dimensions[1].height = 30


def write_table(ws, headers, rows):
    clear_sheet(ws)
    for c, h in enumerate(headers, 1):
        ws.cell(1, c, h)
    for r, row in enumerate(rows, 2):
        for c, v in enumerate(row, 1):
            ws.cell(r, c, v)


# 1) Chinese June daily content calendar.
cal = wb["6月每日排期"] if "6月每日排期" in wb.sheetnames else wb.create_sheet("6月每日排期")
cal_headers = [
    "日期",
    "星期",
    "阶段",
    "内容主线",
    "短视频1（流量/情绪）",
    "短视频或轮播2（信任/分析）",
    "中文Hook方向",
    "结论角度",
    "CTA",
    "素材/制作说明",
]
weekdays = ["周一", "周二", "周三", "周四", "周五", "周六", "周日"]
themes = [
    ("值不值得进口？", "产品看起来有利润，但可能并不赚钱", "中国价格 vs 巴西售价 + 真实到岸成本", "这个产品看起来有利润，但可能不值得进口。", "值得 / 不值得 / 只有规模化才值得"),
    ("本周踩坑", "便宜供应商可能更贵", "靠谱供应商检查清单", "便宜供应商，最后可能非常贵。", "风险可能大于省下的钱"),
    ("中国 vs 巴西", "中国采购价 vs 巴西销售价", "加上运费和税费后的真实利润", "中国很便宜，巴西卖得贵，但利润真的还在吗？", "价格差不等于利润"),
    ("卖家错误", "忘记平台佣金和费用", "真实成本表拆解", "很多卖家都会忘记这个成本。", "利润往往消失在细节里"),
    ("值不值得进口？", "爆品不等于赚钱", "MOQ + 库存积压风险", "产品火不代表利润好。", "必须先验证再放大"),
    ("本周踩坑", "便宜物流背后的隐藏风险", "物流如何吃掉利润", "便宜运费可能让你付出更大代价。", "物流不是附属项，是利润项"),
    ("卖家错误", "第一单买太多", "如何用低 MOQ 测试", "第一单正确，不一定是单价最低。", "现金流优先"),
]
cal_rows = []
start = date(2026, 6, 1)
for i in range(30):
    d = start + timedelta(days=i)
    pillar, traffic, trust, hook, verdict = themes[i % len(themes)]
    stage = "上线周" if i < 7 else ("证明判断力" if i < 14 else ("信任重复" if i < 23 else "轻转化推进"))
    if i == 0:
        pillar = "判断品牌正式上线"
        traffic = "不要再盲目进口（正式上线）"
        trust = "Zhutova 开始帮助卖家分析产品、供应商、MOQ、真实成本和风险"
        hook = "从今天开始，不要再盲目进口。"
        verdict = "Zhutova 正式开始帮助巴西卖家做更好的供应链判断"
    cal_rows.append((
        d.strftime("%Y-%m-%d"),
        weekdays[d.weekday()],
        stage,
        pillar,
        traffic,
        trust,
        hook,
        verdict,
        "想让 Zhutova 分析你的产品？通过 WhatsApp 发给我们。",
        "短视频：口播 + 产品图 + 成本表。轮播：6-7页结构：Hook / 中国价格 / 巴西售价 / 真实成本 / 风险 / 结论 / CTA。",
    ))
write_table(cal, cal_headers, cal_rows)
style_sheet(cal, {"A": 13, "B": 10, "C": 16, "D": 22, "E": 38, "F": 46, "G": 40, "H": 34, "I": 42, "J": 58})
for row in range(2, cal.max_row + 1):
    if cal.cell(row, 3).value == "上线周":
        for col in range(1, cal.max_column + 1):
            cal.cell(row, col).fill = PatternFill("solid", fgColor="E8F4FF")


# 2) Chinese ready-to-post launch copy sheet.
copy = wb["预热与上线文案"] if "预热与上线文案" in wb.sheetnames else wb.create_sheet("预热与上线文案")
copy_headers = ["日期/场景", "素材主题", "发布图片/素材", "中文发布文案（可直接复制）", "标签", "素材文件夹"]
copy_rows = [
    (
        "2026-05-28",
        "中国低价不等于真实利润",
        r"D:\zhutova 2清洗 正确\社媒\instagram_warmup_2026_05_28_31_EN\2026-05-28_warmup_low-price.png",
        "中国的低采购价，不等于真实利润。\n\n在下单之前，卖家需要计算完整成本：产品价格、国际运费、进口税费、平台费用、包装、退货，以及现金流风险。\n\n最便宜的供应商，不一定是最安全的选择。\n\n一种更聪明的选品和采购方式即将开始。\n\n6月1日正式上线。",
        "#Zhutova #ChinaSourcing #ImportFromChina #ProductSourcing #EcommerceSellers #OnlineSellers #SupplierSourcing #MarketplaceSeller #CrossBorderCommerce #SmallBusiness",
        r"D:\zhutova 2清洗 正确\社媒\instagram_warmup_2026_05_28_31_EN",
    ),
    (
        "2026-05-29",
        "不要只按价格选择供应商",
        r"D:\zhutova 2清洗 正确\社媒\instagram_warmup_2026_05_28_31_EN\2026-05-29_warmup_supplier-check.png",
        "不要只按价格选择供应商。\n\n在你付款或承诺大订单之前，一个靠谱供应商应该能给出清晰、稳定、具体的回答。\n\n第一单之前，请检查：\n- 是否有真实产品图片\n- 是否可以提供样品\n- MOQ 是否清楚\n- 是否有出口经验\n- 沟通是否一致稳定\n\n更好的供应商判断，应该发生在付款之前。\n\nZhutova 将于 6 月 1 日上线。",
        "#Zhutova #ChinaSourcing #SupplierSourcing #ImportFromChina #ProductSourcing #EcommerceSellers #OnlineSellers #MarketplaceSeller #CrossBorderCommerce #SmallBusiness",
        r"D:\zhutova 2清洗 正确\社媒\instagram_warmup_2026_05_28_31_EN",
    ),
    (
        "2026-05-30",
        "MOQ 会保护现金流，也可能伤害现金流",
        r"D:\zhutova 2清洗 正确\社媒\instagram_warmup_2026_05_28_31_EN\2026-05-30_warmup_moq.png",
        "MOQ 可能保护你的现金流，也可能伤害你的现金流。\n\n更低的单价看起来很诱人，但过高的最低起订量，可能会把太多钱压在未经验证的库存里。\n\n放大采购前，卖家应该先问：\n- 我能用更小数量测试需求吗？\n- 这些库存能足够快卖出去吗？\n- 加上运费和平台费用后还划算吗？\n- 如果产品卖不动怎么办？\n\n正确的第一单，不一定是单价最低的那一单，而是你能安全验证的那一单。\n\nZhutova 将于 6 月 1 日上线。",
        "#Zhutova #ChinaSourcing #MOQ #ImportFromChina #ProductSourcing #EcommerceSellers #OnlineSellers #MarketplaceSeller #InventoryManagement #CrossBorderCommerce",
        r"D:\zhutova 2清洗 正确\社媒\instagram_warmup_2026_05_28_31_EN",
    ),
    (
        "2026-05-31",
        "Zhutova 明天上线",
        r"D:\zhutova 2清洗 正确\社媒\instagram_warmup_2026_05_28_31_EN\2026-05-31_warmup_launch-tomorrow.png",
        "Zhutova 明天上线。\n\n这是一个帮助线上卖家更清楚、更少盲猜地从中国采购的平台。\n\n通过 Zhutova，卖家可以：\n- 找到产品机会\n- 判断供应商是否可靠\n- 在下单前理解 MOQ\n- 估算真实到岸成本\n- 降低采购风险\n\n明天，一种更聪明的采购方式开始。\n\n6月1日正式上线。",
        "#Zhutova #ChinaSourcing #ImportFromChina #ProductSourcing #SupplierSourcing #EcommerceSellers #OnlineSellers #MarketplaceSeller #CrossBorderCommerce #SmallBusiness",
        r"D:\zhutova 2清洗 正确\社媒\instagram_warmup_2026_05_28_31_EN",
    ),
    (
        "2026-06-01",
        "不要再盲目进口",
        r"D:\zhutova 2清洗 正确\社媒\instagram_launch_2026_06_01_EN",
        "不要再盲目进口。\n\n从今天开始，Zhutova 帮助巴西卖家在进口前做更好的判断：产品、供应商、MOQ、运费、税费、利润和风险。\n\n买得便宜还不够。真正重要的是，完整成本算完之后，利润是否还存在。\n\n想让 Zhutova 分析你的产品？通过 WhatsApp 发给我们。",
        "#Zhutova #ChinaSourcing #ImportFromChina #ProductSourcing #SupplierSourcing #EcommerceSellers #OnlineSellers #MarketplaceSeller #CrossBorderCommerce #ChinaBrazil",
        r"D:\zhutova 2清洗 正确\社媒",
    ),
]
write_table(copy, copy_headers, copy_rows)
style_sheet(copy, {"A": 16, "B": 30, "C": 70, "D": 86, "E": 52, "F": 58})
for row in range(2, copy.max_row + 1):
    asset = copy.cell(row, 3).value
    if asset:
        copy.cell(row, 3).hyperlink = asset
        copy.cell(row, 3).style = "Hyperlink"


# 3) Chinese tomorrow plan. Keep existing if present, but rewrite in a cleaner Chinese version.
tom = wb["明日计划_5月28"] if "明日计划_5月28" in wb.sheetnames else wb.create_sheet("明日计划_5月28", 0)
clear_sheet(tom)
tom.sheet_view.showGridLines = False
tom.freeze_panes = "A4"
tom["A1"] = "Zhutova 5月28日社媒执行计划（中文）"
tom["A2"] = "周主题：Stop importing in the dark / 不要再盲目进口"
tom["A3"] = "当天主题：中国低价不等于真实利润"
for cell in ["A1", "A2", "A3"]:
    tom[cell].font = Font(bold=True, color="FFFFFF" if cell == "A1" else "0A1628", size=16 if cell == "A1" else 12)
    tom[cell].fill = PatternFill("solid", fgColor="5A0F1B" if cell == "A1" else "E8F4FF")
    tom[cell].alignment = Alignment(vertical="center", wrap_text=True)
tom.merge_cells("A1:I1")
tom.merge_cells("A2:I2")
tom.merge_cells("A3:I3")
tom_headers = ["时间", "平台", "内容形式", "发布目的", "中文Hook", "内容要点", "发布文案/CTA", "素材文件", "检查项"]
tom_rows = [
    ("09:30", "Instagram Feed", "单图 Banner", "第一天预热；建立 Zhutova 懂利润和风险的认知", "中国低价不等于真实利润。", "采购价只是第一步；真实利润要看运费、进口税费、平台费用、包装、退货和现金流风险。", copy_rows[0][3], copy_rows[0][2], "图片无错字；日期正确；文案与图片主题一致；Bio/WhatsApp 链接可用。"),
    ("12:30", "Instagram Story", "投票 + 转发早上 Banner", "引导互动，测试受众是否理解真实成本痛点", "你进口前会计算真实到岸成本吗？", "Story 1 转发 banner；Story 2 投票：会 / 还不会；Story 3 提醒运费+税费+平台费会改变利润。", "CTA：回复 Story，或把想分析的产品发给 Zhutova。", copy_rows[0][2], "投票文案要短；不要堆太多字；保留 Zhutova 标识。"),
    ("16:30", "TikTok / Reels", "15-25秒短视频", "用“低价不等于赚钱”的反常识做流量入口", "买得便宜，不代表能赚钱。", "脚本：1. 这个产品在中国看起来很便宜；2. 但你还要付运费、税费、平台费和退货成本；3. 如果下单前不算，利润会消失；4. 不要再盲目进口。", "短文案：低价不承担亏损。Zhutova 6月1日上线，帮助卖家在进口前分析产品、成本和风险。", "产品图 + 成本表录屏 + 口播；也可用 5/28 banner 做动态缩放。", "前3秒必须有 Hook；字幕要大；结尾出现 6月1日上线。"),
    ("19:30", "Facebook", "讨论帖", "引评论，不要一上来介绍品牌", "进口最容易吃掉利润的是：运费、税费，还是平台费？", "发起讨论：很多卖家只看中国采购价，但真正决定利润的是完整成本。让大家评论最容易被低估的成本。", "中国低价不等于真实利润。\n\n实际进口到巴西销售时，最容易吃掉利润的是哪一项？\n\n1. 国际运费\n2. 进口税费\n3. 平台费用\n4. 退货\n5. 库存积压\n\n评论告诉我们，哪一项最影响你的利润。", copy_rows[0][2], "问题放第一行；不要硬介绍平台；评论后可引导到 WhatsApp。"),
    ("21:00", "内部复盘", "数据记录", "判断首日痛点是否有效", "记录互动和咨询", "记录 IG 收藏/分享/评论、Story 投票、TikTok 完播、Facebook 评论、WhatsApp 咨询。", "复盘问题：低价不等于利润有没有引发评论？有没有人发产品来问？哪条内容最容易互动？", "后台数据 / 评论截图", "保留评论截图；把用户问题整理成 5/29 供应商主题素材。"),
]
for c, h in enumerate(tom_headers, 1):
    tom.cell(5, c, h)
for r, row in enumerate(tom_rows, 6):
    for c, v in enumerate(row, 1):
        tom.cell(r, c, v)
style_sheet(tom, {"A": 12, "B": 18, "C": 20, "D": 30, "E": 34, "F": 52, "G": 78, "H": 58, "I": 42}, freeze="A4")
for r in range(6, tom.max_row + 1):
    tom.row_dimensions[r].height = 126


# Put Chinese execution sheets first.
front = ["明日计划_5月28", "新版总览", "四大内容IP", "6月每日排期", "平台打法", "预热与上线文案"]
for idx, name in enumerate(front):
    if name in wb.sheetnames:
        obj = wb[name]
        wb._sheets.remove(obj)
        wb._sheets.insert(idx, obj)

for sheet in wb.worksheets:
    for row in sheet.iter_rows():
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)

wb.save(OUT)
print(OUT)
