from pathlib import Path

from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side


WORKBOOK = Path(r"D:\zhutova 2清洗 正确\社媒\Zhutova_社媒运营计划_简洁版.xlsx")

NEWS_SCRIPT = """大家好，这里是 ZHUTOVA 一分钟供应链周报。

本周有三条跨境新闻，
值得所有做外贸和跨境电商的人关注。

第一，亚马逊美国站正式更新了 FBM 自配送政策。
平台对发货时效和履约能力要求越来越高。
如果还是等订单来了再安排发货，
未来压力会更大。
提前备货、海外仓这些能力，
会越来越重要。

第二，欧盟从 7 月开始取消 150 欧元以下商品的进口免税政策。
对于做欧洲市场的卖家来说，
低价小包的成本会进一步上涨。
未来拼的不只是价格，
更是供应链和合规能力。

第三，最近海运价格持续上涨。
不少外贸企业都表示，
运费几乎每周都在调整。
如果你的报价周期比较长，
建议把运费波动提前考虑进去，
避免利润被物流成本吃掉。

以上就是本周值得关注的三条供应链资讯。
如果你想了解更多全球采购、跨境电商和供应链趋势，
记得关注 ZHUTOVA。"""


def style_sheet(ws):
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A2"
    thin = Side(style="thin", color="D9E2EC")
    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.border = Border(bottom=thin)
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="0A1628")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    widths = {
        "A": 12,
        "B": 24,
        "C": 46,
        "D": 46,
        "E": 46,
        "F": 56,
        "G": 26,
        "H": 22,
    }
    for col, width in widths.items():
        ws.column_dimensions[col].width = width
    ws.row_dimensions[1].height = 30
    for row_idx in range(2, ws.max_row + 1):
        ws.row_dimensions[row_idx].height = 110


def upsert_news_input_sheet(wb):
    name = "07_新闻热点输入"
    if name in wb.sheetnames:
        ws = wb[name]
        ws.delete_rows(1, ws.max_row)
    else:
        ws = wb.create_sheet(name)

    ws.append(["周次", "热点主题", "新闻1", "新闻2", "新闻3", "我提供的原始文案", "使用方式", "备注"])
    ws.append([
        "第5周",
        "ZHUTOVA 一分钟供应链周报",
        "亚马逊美国站更新 FBM 自配送政策，平台更重视发货时效和履约能力。",
        "欧盟 7 月开始取消 150 欧元以下商品进口免税政策，低价小包成本上涨。",
        "海运价格持续上涨，报价周期较长时需要提前考虑运费波动。",
        NEWS_SCRIPT,
        "用于每周第1条：实时行业热点新闻",
        "之后每周你把新闻热点填在这里，我再重新规划口播。",
    ])
    style_sheet(ws)


def main():
    wb = load_workbook(WORKBOOK)

    upsert_news_input_sheet(wb)

    ws04 = wb["04_未来4周主题安排"]
    for row_idx in range(2, ws04.max_row + 1):
        if ws04.cell(row_idx, 1).value == "第5周":
            ws04.cell(row_idx, 2).value = "实时行业热点新闻：ZHUTOVA 一分钟供应链周报"
            break

    ws05 = wb["05_口播文案库"]
    for row_idx in range(2, ws05.max_row + 1):
        if ws05.cell(row_idx, 1).value == "第5周" and ws05.cell(row_idx, 3).value == "实时行业热点新闻":
            ws05.cell(row_idx, 2).value = "实时行业热点新闻：ZHUTOVA 一分钟供应链周报"
            ws05.cell(row_idx, 4).value = NEWS_SCRIPT
            ws05.cell(row_idx, 5).value = "实时行业热点新闻 / 供应链周报 / 60秒"
            ws05.row_dimensions[row_idx].height = 190
            break

    wb.save(WORKBOOK)
    print(WORKBOOK)


if __name__ == "__main__":
    main()
