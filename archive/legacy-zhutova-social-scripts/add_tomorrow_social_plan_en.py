from pathlib import Path
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


ROOT = Path(r"D:\zhutova 2清洗 正确")
XLSX = ROOT / "zhutova社媒规划_新版判断品牌计划.xlsx"

wb = openpyxl.load_workbook(XLSX)
sheet_name = "Tomorrow Plan_May28_EN"
if sheet_name in wb.sheetnames:
    del wb[sheet_name]

ws = wb.create_sheet(sheet_name, 0)
ws.sheet_view.showGridLines = False
ws.freeze_panes = "A4"

navy = "0A1628"
wine = "5A0F1B"
blue = "1C97FF"
light_blue = "E8F4FF"
cream = "FFF9F2"
line = "D9E2EC"

ws["A1"] = "Zhutova May 28 Social Media Execution Plan"
ws["A2"] = "Weekly theme: Stop importing in the dark"
ws["A3"] = "Daily theme: A low China price does not mean real profit"
for cell in ["A1", "A2", "A3"]:
    ws[cell].font = Font(bold=True, color="FFFFFF" if cell == "A1" else navy, size=16 if cell == "A1" else 12)
    ws[cell].fill = PatternFill("solid", fgColor=wine if cell == "A1" else light_blue)
    ws[cell].alignment = Alignment(vertical="center", wrap_text=True)
ws.merge_cells("A1:I1")
ws.merge_cells("A2:I2")
ws.merge_cells("A3:I3")

headers = [
    "Time",
    "Platform",
    "Format",
    "Goal",
    "English Hook",
    "Content Points",
    "Caption / CTA",
    "Creative Asset",
    "Checklist",
]

rows = [
    [
        "09:30",
        "Instagram Feed",
        "Single-image banner",
        "First warm-up post; establish that Zhutova understands profit and risk, not just sourcing.",
        "A low China price is not real profit.",
        "The purchase price is only the first number. Real profit depends on freight, import taxes, marketplace fees, packaging, returns, and cash-flow risk.",
        "A low China price does not mean real profit.\n\nBefore placing an order, online sellers need to calculate the full cost: product price, international freight, import taxes, marketplace fees, packaging, returns and cash-flow risk.\n\nThe cheapest supplier is not always the safest decision.\n\nA smarter way to source is coming.\n\nLaunching June 1.",
        r"D:\zhutova 2清洗 正确\社媒\instagram_warmup_2026_05_28_31_EN\2026-05-28_warmup_low-price.png",
        "Use English creative; confirm date says MAY 28; caption is English; hashtags at the end; bio/WhatsApp link works.",
    ],
    [
        "12:30",
        "Instagram Story",
        "Poll + repost morning banner",
        "Create lightweight interaction and test whether the audience relates to real landed cost.",
        "Do you calculate real landed cost before importing?",
        "Story 1: repost banner.\nStory 2: poll Yes / Not yet.\nStory 3: remind that freight + taxes + fees can change the margin.",
        "CTA: Reply to this story or send us the product you want analyzed.",
        "Use the same May 28 English banner; add Instagram poll sticker.",
        "Keep poll text short; do not overcrowd the Story; keep Zhutova visible.",
    ],
    [
        "16:30",
        "TikTok / Reels",
        "15-25 sec short video",
        "Traffic post; use a sharp anti-cheap-price angle to stop the scroll.",
        "Importing cheap does not mean making money.",
        "Script:\n1. This product looks cheap in China.\n2. But you still need to pay freight, taxes, marketplace fees and returns.\n3. If you do not calculate before buying, profit disappears.\n4. Stop importing in the dark.",
        "Short caption: Low price does not pay for losses. Zhutova launches June 1 to help sellers analyze product, cost and risk before importing.",
        "Use product image + cost table screen recording + voiceover; or animate the May 28 banner.",
        "Hook must appear in the first 3 seconds; captions should be large; end with 'Launching June 1'.",
    ],
    [
        "19:30",
        "Facebook",
        "Discussion post",
        "Drive comments; do not start with a brand intro.",
        "What kills import margin most: freight, taxes or marketplace fees?",
        "Invite discussion: many sellers only look at the China purchase price, but total cost decides profit. Ask people to comment which cost is most underestimated.",
        "Post: A low China price does not mean real profit.\n\nIn practice, what kills margin the most when importing to sell online?\n\n1. International freight\n2. Import taxes\n3. Marketplace fees\n4. Returns\n5. Dead stock\n\nComment which cost weighs the most in your operation.",
        "Use May 28 English banner or a plain text discussion post.",
        "Put the question in the first line; avoid platform introduction; reply to comments and guide relevant users to WhatsApp.",
    ],
    [
        "21:00",
        "Internal Review",
        "Data log",
        "Check whether the first warm-up angle resonates.",
        "Record engagement and inbound questions",
        "Track IG saves/shares/comments, Story poll results, TikTok watch completion, Facebook comments, and WhatsApp inquiries.",
        "Review questions: Did 'low price does not mean profit' trigger comments? Did anyone send a product for analysis? Which content generated the strongest interaction?",
        "Record in sheet or platform dashboard.",
        "Save comment screenshots; turn user questions into May 29 supplier-check content.",
    ],
]

start_row = 5
for c, h in enumerate(headers, 1):
    cell = ws.cell(start_row, c, h)
    cell.font = Font(bold=True, color="FFFFFF")
    cell.fill = PatternFill("solid", fgColor=navy)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

for r, row in enumerate(rows, start_row + 1):
    for c, value in enumerate(row, 1):
        cell = ws.cell(r, c, value)
        cell.alignment = Alignment(vertical="top", wrap_text=True)
        cell.fill = PatternFill("solid", fgColor=cream if r % 2 == 0 else "FFFFFF")

thin = Side(style="thin", color=line)
for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=len(headers)):
    for cell in row:
        cell.border = Border(bottom=thin)

widths = [12, 18, 20, 30, 36, 54, 74, 58, 44]
for i, width in enumerate(widths, 1):
    ws.column_dimensions[get_column_letter(i)].width = width

for r in range(1, ws.max_row + 1):
    ws.row_dimensions[r].height = 24
for r in range(start_row + 1, ws.max_row + 1):
    ws.row_dimensions[r].height = 126

wb.save(XLSX)
print(XLSX)
