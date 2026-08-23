from pathlib import Path
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side


ROOT = Path(r"D:\zhutova 2清洗 正确")
PLAN = ROOT / "社媒" / "Zhutova_社媒运营计划_最终简洁版.xlsx"

wb = openpyxl.load_workbook(PLAN)
sheet_name = "06_后续规划方向"
if sheet_name in wb.sheetnames:
    del wb[sheet_name]

ws = wb.create_sheet(sheet_name)
headers = ["阶段", "周期", "核心任务", "内容方向", "判断标准"]
rows = [
    (
        "第1阶段：建立判断力",
        "第1-2周",
        "让用户先记住：Zhutova 懂进口利润和风险。",
        "重点发流量型内容：低价不等于利润、供应商风险、MOQ现金流、中国vs巴西价格差、卖家常见错误。",
        "是否有播放、收藏、评论；是否有人开始发产品来问。",
    ),
    (
        "第2阶段：补平台信任",
        "第3-4周",
        "让用户知道 Zhutova 不只是内容号，而是一站式跨境 B2B 平台。",
        "增加品牌/商务内容：Zhutova 是什么、Zhutova 服务、工厂直供到智能履约、供应链幕后、平台运营环境。",
        "是否有人理解平台能力；是否出现商务咨询或更高质量私信。",
    ),
    (
        "第3阶段：做案例沉淀",
        "第2个月",
        "用案例证明 Zhutova 的判断和执行能力。",
        "持续做模拟/真实案例：一个产品是否值得进口、一个供应商是否可靠、一个MOQ是否合理、一个品类怎么测。",
        "案例内容是否带来收藏、转发、WhatsApp产品分析请求。",
    ),
    (
        "第4阶段：商务合作承接",
        "第3个月",
        "把社媒流量转成合作线索和长期客户关系。",
        "发布合作伙伴、平台生态、供应链金融、定制采购、卖家合作流程、供应商合作机会。",
        "是否产生B端合作询盘、供应商/卖家/渠道伙伴咨询。",
    ),
    (
        "长期循环方法",
        "每月持续",
        "栏目不变，案例和品类不断换。",
        "每周仍按6个栏目循环：值不值得进口、本周踩坑、中国vs巴西、卖家错误、Zhutova服务、供应链。每月换一个主主题，例如利润月、供应商月、履约月、案例月。",
        "每月复盘哪类内容带来高质量咨询，下月加大这类内容比例。",
    ),
]

ws.append(headers)
for row in rows:
    ws.append(row)

thin = Side(style="thin", color="D9E2EC")
for row in ws.iter_rows():
    for cell in row:
        cell.alignment = Alignment(vertical="top", wrap_text=True)
        cell.border = Border(bottom=thin)

for cell in ws[1]:
    cell.font = Font(bold=True, color="FFFFFF")
    cell.fill = PatternFill("solid", fgColor="0A1628")
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

widths = {"A": 24, "B": 14, "C": 44, "D": 78, "E": 46}
for col, width in widths.items():
    ws.column_dimensions[col].width = width

for r in range(2, ws.max_row + 1):
    ws.row_dimensions[r].height = 105

wb.save(PLAN)
print(PLAN)
