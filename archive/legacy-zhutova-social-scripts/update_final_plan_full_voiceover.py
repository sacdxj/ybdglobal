from pathlib import Path
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side


ROOT = Path(r"D:\zhutova 2清洗 正确")
PLAN = ROOT / "社媒" / "Zhutova_社媒运营计划_最终简洁版.xlsx"

wb = openpyxl.load_workbook(PLAN)
sheet_name = "05_口播文案模板"
if sheet_name in wb.sheetnames:
    ws = wb[sheet_name]
    ws.delete_rows(1, ws.max_row)
else:
    ws = wb.create_sheet(sheet_name)

headers = ["栏目", "口播结构", "完整20秒口播内容", "可替换变量"]
rows = [
    (
        "值不值得进口？",
        "产品表面机会 -> 拆真实成本 -> 给出判断",
        "这个产品看起来很赚钱，但不一定值得进口。\n\n很多卖家只看中国采购价和巴西售价，中间的运费、税费、平台费、MOQ 和退货风险都没算进去。\n\n如果这些成本加完，利润还很薄，就不能直接上大单。\n\n进口前，先判断它到底值不值得做。",
        "这个产品 / 中国采购价 / 巴西售价 / 运费 / 税费 / MOQ",
    ),
    (
        "本周踩坑",
        "指出一个坑 -> 解释为什么亏 -> 给正确做法",
        "本周踩坑：便宜供应商，可能是最贵的错误。\n\n很多卖家只看报价低，却没有确认真实图片、样品、MOQ、出口经验和沟通稳定性。\n\n等到货不对版、交期延误，或者售后没人回，省下的钱早就不够赔。\n\n下单前，先判断供应商风险。",
        "便宜供应商 / 低价运费 / MOQ过高 / 没看样品 / 交期延误",
    ),
    (
        "中国 vs 巴西价格差",
        "展示价格差 -> 加入真实成本 -> 判断利润是否还在",
        "中国采购价很低，巴西售价看起来很高，但这不代表一定赚钱。\n\n真正要看的，是产品到巴西后的完整成本：国际运费、进口税费、平台佣金、包装、退货和库存风险。\n\n价格差只是入口，真实利润才是重点。\n\n不要只看差价，要看最后还剩多少利润。",
        "中国价格 / 巴西售价 / 产品类目 / 平台费用 / 真实利润",
    ),
    (
        "卖家常见错误",
        "指出错误 -> 讲后果 -> 给修正方式",
        "很多卖家进口亏钱，不是因为产品不好，而是因为只看了采购价。\n\n他们忘了平台费、运费、税费、包装、退货，还有库存卖不动的风险。\n\n结果产品看起来有利润，真正卖的时候利润却被吃掉。\n\n进口前，一定要先算完整成本。",
        "只看采购价 / 忘记平台费 / 第一单买太多 / 不算退货 / 忽略包装",
    ),
    (
        "Zhutova 服务",
        "用户问题 -> Zhutova 能做什么 -> 引导咨询",
        "你以为 Zhutova 只是帮你找供应商？不只是。\n\n真正重要的是，在进口前判断这个产品能不能赚钱，供应商靠不靠谱，MOQ 合不合理，完整成本算完还有没有利润。\n\nZhutova 帮卖家做产品、供应商、成本和风险分析。\n\n想分析产品，可以发给 Zhutova。",
        "产品分析 / 供应商判断 / MOQ建议 / 到岸成本 / 风险分析",
    ),
    (
        "供应链",
        "打破误解 -> 展示链路 -> 回到平台价值",
        "找到工厂，不代表采购结束。\n\n你还要处理交易、仓储、物流、履约和售后服务。\n\n如果只看一个报价，后面的风险很容易被忽略。\n\nZhutova 连接工厂直供、全球贸易、本地仓储和智能履约，让采购从单点找货变成全链路判断。",
        "工厂直供 / 全球贸易 / 本地仓储 / 智能履约 / 物流节点",
    ),
]

ws.append(headers)
for row in rows:
    ws.append(row)

thin = Side(style="thin", color="D9E2EC")
for row in ws.iter_rows():
    for cell in row:
        cell.alignment = Alignment(vertical="top", wrap_text=True)
        cell.border = Border(bottom=thin)

for cell in ws[1]:
    cell.font = Font(bold=True, color="FFFFFF")
    cell.fill = PatternFill("solid", fgColor="0A1628")
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

widths = {"A": 24, "B": 38, "C": 90, "D": 42}
for col, width in widths.items():
    ws.column_dimensions[col].width = width

for r in range(2, ws.max_row + 1):
    ws.row_dimensions[r].height = 145

wb.save(PLAN)
print(PLAN)
