from pathlib import Path
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side


ROOT = Path(r"D:\zhutova 2清洗 正确")
PLAN = ROOT / "社媒" / "Zhutova_社媒内容汇总_2026_06" / "00_社媒规划" / "zhutova社媒规划_中文版本.xlsx"

wb = openpyxl.load_workbook(PLAN)
sheet_name = "一周视频规划_每日1条"
if sheet_name in wb.sheetnames:
    del wb[sheet_name]

ws = wb.create_sheet(sheet_name, 0)
ws.sheet_view.showGridLines = False
ws.freeze_panes = "A2"

headers = [
    "天数",
    "内容主题",
    "视频标题/Hook",
    "核心目的",
    "30-45秒脚本结构",
    "画面/素材建议",
    "字幕关键词",
    "CTA",
    "适合平台",
]

rows = [
    (
        "Day 1",
        "品牌定位：Zhutova 是什么",
        "Zhutova is not just sourcing. It is a cross-border B2B platform.",
        "先校准认知：Zhutova 不是普通货代/采购号，而是一站式跨境 B2B 平台。",
        "0-3秒：Not every seller needs another supplier list.\n3-12秒：What sellers need is a smarter way to judge products, suppliers, cost and risk.\n12-28秒：Zhutova connects factory direct supply, global trading, local warehousing and smart fulfillment.\n28-40秒：So sellers can move from product idea to sourcing decision with more clarity.",
        "H5首页/平台图、世界地图、中国制造产品、工厂/仓储/物流图标、Zhutova logo。",
        "One-stop B2B platform / Factory supply / Warehousing / Fulfillment / Less guesswork",
        "Want to source with less risk? Follow Zhutova.",
        "Instagram Reels / TikTok / YouTube Shorts",
    ),
    (
        "Day 2",
        "判断力：低价不等于利润",
        "A low China price does not mean real profit.",
        "延续核心心智：Zhutova 懂利润和风险。",
        "0-3秒：A product can be cheap in China and still lose money.\n3-15秒：Because product cost is only one part of the calculation.\n15-30秒：You still need freight, import taxes, marketplace fees, packaging, returns and MOQ risk.\n30-45秒：Before importing, sellers need to know if profit still exists after the full cost.",
        "成本表动画：Product Cost + Freight + Taxes + Fees + Returns；产品价格对比图。",
        "China price / Real landed cost / Margin / Risk",
        "Send your product to Zhutova for analysis.",
        "Instagram Reels / TikTok",
    ),
    (
        "Day 3",
        "平台能力：从工厂直供到智能履约",
        "From factory direct supply to smart fulfillment.",
        "把 H5 里的平台链路讲清：工厂直供、全球贸易、本地仓储、智能履约。",
        "0-3秒：Sourcing is not finished when you find a factory.\n3-12秒：You still need trading support, warehousing, logistics and fulfillment.\n12-32秒：Zhutova connects four steps: factory direct supply, global trading, local warehousing and smart fulfillment.\n32-45秒：That is why sourcing should be managed as a full chain, not a single quote.",
        "四步流程图：Factory Direct Supply -> Global Trading -> Local Warehousing -> Smart Fulfillment；配图用 H5 图标风格。",
        "Factory / Trading / Warehousing / Fulfillment / Full chain",
        "Want a more stable sourcing process? Talk to Zhutova.",
        "Instagram Reels / LinkedIn / TikTok",
    ),
    (
        "Day 4",
        "供应商判断：不要只看价格",
        "Do not choose a supplier by price alone.",
        "建立 Zhutova 的供应商验证专业感。",
        "0-3秒：The cheapest supplier can become the most expensive mistake.\n3-15秒：Before payment, check real product photos, samples, MOQ, export experience and communication quality.\n15-32秒：A reliable supplier gives clear answers before you commit to a large order.\n32-45秒：Zhutova helps sellers check risk before buying.",
        "供应商聊天截图样式、checklist 动画、红旗/绿旗对比。",
        "Real photos / Sample / MOQ / Export experience / Supplier risk",
        "Want Zhutova to check a supplier? Send it on WhatsApp.",
        "TikTok / Instagram Reels / Facebook",
    ),
    (
        "Day 5",
        "运营环境：幕后流程",
        "Before a product reaches the seller, these checks should happen.",
        "展示运营环境和幕后流程，增强信任。",
        "0-3秒：Good sourcing is not luck. It is a process.\n3-15秒：Before ordering, you need product specs, supplier communication, sample checks and landed cost estimation.\n15-32秒：Then come order placement, logistics tracking, warehousing and service support.\n32-45秒：Zhutova turns sourcing into a more structured process.",
        "幕后素材：表格、报价单、物流节点、仓储、QC流程；没有真实素材可用 PPT 动画。",
        "Process / Supplier check / Sample / Cost / Logistics / Service",
        "Follow Zhutova for smarter sourcing decisions.",
        "Instagram Reels / YouTube Shorts",
    ),
    (
        "Day 6",
        "客户案例/模拟案例",
        "A seller wanted to import this product. Here is what we checked first.",
        "用案例证明平台不是空谈，而是能帮助判断。",
        "0-3秒：A seller asked if this product was worth importing.\n3-15秒：We did not start with the supplier price. We checked demand, MOQ, freight, taxes and selling channel fees.\n15-32秒：Then we looked at supplier reliability and fulfillment risk.\n32-45秒：The answer was not simply yes or no. It depends on volume, cost and risk.",
        "模拟产品图 + 判断表：Demand / MOQ / Freight / Fees / Risk / Verdict。",
        "Case study / Product analysis / Not yes or no / Depends on margin",
        "Want your product analyzed? Send it to Zhutova.",
        "Instagram Reels / TikTok / Facebook",
    ),
    (
        "Day 7",
        "生态与合作：平台背书",
        "Zhutova connects more than products. It connects the sourcing ecosystem.",
        "补充合作伙伴、全球市场、商务合作认知。",
        "0-3秒：A strong sourcing platform is not only about finding products.\n3-15秒：It also connects sales channels, supply chain services, warehousing, payments and fulfillment.\n15-32秒：Zhutova works within a broader ecosystem including ecommerce platforms, service partners and global markets.\n32-45秒：For sellers, that means more support from sourcing decision to business execution.",
        "H5合作伙伴 logo 页、地图页、平台生态图；注意不要夸大合作关系，用“ecosystem / partners / channels”表达。",
        "Ecosystem / Partners / Sales channels / Global markets / Business support",
        "For business cooperation, contact Zhutova.",
        "LinkedIn / Instagram Reels / Facebook",
    ),
]

for c, h in enumerate(headers, 1):
    ws.cell(1, c, h)
for r, row in enumerate(rows, 2):
    for c, value in enumerate(row, 1):
        ws.cell(r, c, value)

thin = Side(style="thin", color="D9E2EC")
for row in ws.iter_rows():
    for cell in row:
        cell.alignment = Alignment(vertical="top", wrap_text=True)
        cell.border = Border(bottom=thin)
for cell in ws[1]:
    cell.font = Font(bold=True, color="FFFFFF")
    cell.fill = PatternFill("solid", fgColor="0A1628")
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

widths = {
    "A": 10,
    "B": 24,
    "C": 42,
    "D": 34,
    "E": 70,
    "F": 52,
    "G": 36,
    "H": 42,
    "I": 28,
}
for col, width in widths.items():
    ws.column_dimensions[col].width = width
for r in range(2, ws.max_row + 1):
    ws.row_dimensions[r].height = 138

wb.save(PLAN)
print(PLAN)
