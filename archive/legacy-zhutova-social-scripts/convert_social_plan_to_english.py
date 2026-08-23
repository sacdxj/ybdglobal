from pathlib import Path
from datetime import date, timedelta
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side


ROOT = Path(r"D:\zhutova 2清洗 正确")
XLSX = ROOT / "zhutova社媒规划_新版判断品牌计划.xlsx"

wb = openpyxl.load_workbook(XLSX)

# Remove Portuguese duplicate tomorrow sheet; the English version already exists.
if "明日计划_5月28" in wb.sheetnames:
    del wb["明日计划_5月28"]

# Rename new planning sheets to English where possible.
rename_map = {
    "新版总览": "Overview_EN",
    "四大内容IP": "Content IP_EN",
    "6月每日排期": "June Daily Plan_EN",
    "平台打法": "Platform Playbook_EN",
    "预热与上线文案": "Warmup Launch Copy_EN",
}
for old, new in rename_map.items():
    if old in wb.sheetnames:
        if new in wb.sheetnames:
            del wb[new]
        wb[old].title = new


def clear_sheet(ws):
    ws.delete_rows(1, ws.max_row)


def style_sheet(sheet, widths=None, freeze="A2"):
    sheet.freeze_panes = freeze
    sheet.sheet_view.showGridLines = False
    widths = widths or {}
    for col, width in widths.items():
        sheet.column_dimensions[col].width = width
    thin = Side(style="thin", color="D9E2EC")
    for row in sheet.iter_rows():
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.border = Border(bottom=thin)
    for cell in sheet[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="0A1628")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    sheet.row_dimensions[1].height = 30


def write_table(sheet, headers, rows, start_row=1):
    for c, h in enumerate(headers, 1):
        sheet.cell(start_row, c, h)
    for r, row in enumerate(rows, start_row + 1):
        for c, value in enumerate(row, 1):
            sheet.cell(r, c, value)


def replace_sheet(name, headers, rows, widths):
    if name not in wb.sheetnames:
        ws = wb.create_sheet(name)
    else:
        ws = wb[name]
        clear_sheet(ws)
    write_table(ws, headers, rows)
    style_sheet(ws, widths)
    return ws


overview_rows = [
    ("Core Positioning", "Zhutova is not simply launching a website. It is launching as a supply-chain decision brand for Brazilian sellers. Core perception: Zhutova understands import profit and risk."),
    ("June Message", "Stop importing in the dark. All content should revolve around profit, risk, and better sourcing decisions."),
    ("June Objective", "Build the first layer of brand awareness: Zhutova helps sellers judge whether a product is worth importing, whether a supplier is reliable, whether MOQ makes sense, and whether margin survives the full cost calculation."),
    ("Do Not Lead With", "Do not focus on website launch, welcome messages, feature introductions, or 'who we are' content. In the first 30 days, talk less about registration and features; show decision-making ability."),
    ("Lead With", "Use examples and breakdowns to prove that Zhutova understands margin, risk, cash flow, and Brazilian ecommerce seller realities better than a generic freight forwarder or supplier directory."),
    ("Core CTA", "Want Zhutova to analyze your product? Send it to us on WhatsApp."),
    ("Primary Metrics", "Prioritize saves, comments, WhatsApp product inquiries, shares, and video completion. Follower count and likes are secondary."),
    ("Tone", "English-facing planning, but local execution may be adapted by channel. Be direct, judgment-led, risk-aware, and seller-focused. Avoid abstract brand language."),
    ("Posting Rhythm", "TikTok/Reels: 2 posts per day, one emotional traffic post and one analytical trust post. Instagram: 1 Reel + 1 Carousel per day. Facebook: discussion posts. YouTube: search-driven content."),
]
overview = replace_sheet("Overview_EN", ["Section", "English Plan / Execution Requirement"], overview_rows, {"A": 24, "B": 110})
overview["A1"].fill = PatternFill("solid", fgColor="5A0F1B")
overview["B1"].fill = PatternFill("solid", fgColor="5A0F1B")


pillars_rows = [
    (
        "Is It Worth Importing?",
        "The most important core IP. It shifts Zhutova from 'teaching import' to 'helping sellers decide'.",
        "Hook -> China purchase cost -> Brazil selling price -> freight/tax/fees/MOQ -> risk -> Worth it? Not worth it? Only at scale?",
        "This product looks profitable... but maybe it is not.\nImporting cheap does not mean profit.",
        "Worth it? Not worth it? Tight margin? Only at scale?\nCTA: Want Zhutova to analyze your product? Send it on WhatsApp.",
        "TikTok, Reels, IG Carousel, YouTube Shorts",
    ),
    (
        "Trap of the Week",
        "Creates emotion and shareability by exposing sourcing mistakes and hidden risks.",
        "Trap -> why sellers fall for it -> where the loss happens -> how to avoid it -> CTA",
        "Cheap freight can become expensive.\nThis MOQ destroyed the seller's cash flow.\nA viral product does not mean profit.",
        "CTA: Before buying, ask Zhutova for an analysis.",
        "TikTok, Reels, Facebook",
    ),
    (
        "China vs Brazil",
        "The easiest format to save and share. Use the price gap as the entry point, then educate through real cost.",
        "Left: China cost -> Right: Brazil price -> add freight/tax/fees -> margin -> risk -> Is it worth importing?",
        "It costs X in China and sells for Y in Brazil. But the real calculation is different.",
        "CTA: Want to analyze a category? Send it to Zhutova.",
        "Instagram Carousel, Reels, Facebook",
    ),
    (
        "Seller Mistake",
        "Triggers self-recognition, comments, saves, and shares.",
        "Mistake -> consequence -> hidden cost -> rule of thumb -> CTA",
        "The mistake that makes sellers lose money importing.\nThe biggest mistake is not always import tax.\nMany sellers forget this cost.",
        "CTA: Stop importing in the dark. Talk to Zhutova.",
        "Facebook, Instagram, YouTube Shorts",
    ),
]
replace_sheet(
    "Content IP_EN",
    ["Content Pillar", "Role", "Fixed Structure", "Example Hooks", "Verdict / CTA", "Best Platforms"],
    pillars_rows,
    {"A": 24, "B": 40, "C": 50, "D": 44, "E": 44, "F": 30},
)


calendar_headers = [
    "Date",
    "Day",
    "Stage",
    "Daily Pillar",
    "Reel/TikTok 1 (Traffic)",
    "Reel/Carousel 2 (Trust)",
    "English Hook",
    "Verdict Angle",
    "CTA",
    "Notes / Creative Direction",
]
weekdays = ["Mon", "Tue", "Wed", "Thu", "Fri", "Sat", "Sun"]
themes = [
    ("Is It Worth Importing?", "Product looks profitable, but maybe it is not", "China vs Brazil + real landed cost", "This product looks profitable... but maybe it is not.", "Worth it / Not worth it / Only at scale"),
    ("Trap of the Week", "Cheap supplier can become expensive", "Reliable supplier checklist", "A cheap supplier can become very expensive.", "Risk can be bigger than savings"),
    ("China vs Brazil", "China price vs Brazil selling price", "Margin after freight and taxes", "It costs little in China and sells higher in Brazil. But what about margin?", "Price gap does not equal profit"),
    ("Seller Mistake", "Forgetting marketplace fees", "Real cost table", "Many sellers forget this cost.", "Margin disappears in the details"),
    ("Is It Worth Importing?", "Viral product does not mean profit", "MOQ + dead stock", "A viral product does not pay the bills if margin is weak.", "Only worth it with validation"),
    ("Trap of the Week", "Cheap freight with bad timing or risk", "When freight kills margin", "Cheap freight can become expensive.", "Logistics is a profit factor"),
    ("Seller Mistake", "Buying too much in the first order", "How to test with low MOQ", "The right first order is not always the cheapest one.", "Cash flow first"),
]
rows = []
start = date(2026, 6, 1)
for i in range(30):
    d = start + timedelta(days=i)
    theme, reel1, reel2, hook, verdict = themes[i % len(themes)]
    stage = "Launch Week" if i < 7 else ("Proof Building" if i < 14 else ("Trust + Repetition" if i < 23 else "Soft Conversion Push"))
    if i == 0:
        theme = "Decision Brand Launch"
        reel1 = "Stop importing in the dark (official launch)"
        reel2 = "Zhutova starts analyzing product, supplier, and risk"
        hook = "Starting today, stop importing in the dark."
        verdict = "Zhutova officially starts helping Brazilian sellers make better sourcing decisions"
    rows.append((
        d.strftime("%Y-%m-%d"),
        weekdays[d.weekday()],
        stage,
        theme,
        reel1,
        reel2,
        hook,
        verdict,
        "Want Zhutova to analyze your product? Send it to us on WhatsApp.",
        "Short video: voiceover + PPT/cost table recording. Carousel: 6-7 pages: Hook / China price / Brazil price / real cost / risk / verdict / CTA.",
    ))
june = replace_sheet("June Daily Plan_EN", calendar_headers, rows, {"A": 13, "B": 10, "C": 18, "D": 24, "E": 38, "F": 42, "G": 46, "H": 34, "I": 46, "J": 58})
for row in range(2, june.max_row + 1):
    if june.cell(row, 3).value == "Launch Week":
        for col in range(1, june.max_column + 1):
            june.cell(row, col).fill = PatternFill("solid", fgColor="E8F4FF")


platform_rows = [
    ("TikTok", "Use emotional risk-led content to attract sellers and make them realize importing is not only about price.", "2 posts/day", "Trap of the Week / Seller Mistake / Is It Worth Importing?", "First 3 seconds must challenge an assumption: Importing cheap does not mean profit. 15-35 seconds, voiceover + product/cost table.", "Completion rate, comments, DMs"),
    ("Instagram Reels", "Build decision authority and direct interested users to WhatsApp.", "1 post/day", "Short product, supplier, MOQ, or cost-risk breakdown", "Clean visual style, sharp title, clear CTA at the end.", "Saves, shares, profile visits, WhatsApp clicks"),
    ("Instagram Carousel", "Create saveable decision cards.", "Daily or at least 5/week", "China vs Brazil / Is It Worth Importing? / Checklist", "7 pages: Hook / China price / Brazil price / real cost / risk / verdict / CTA.", "Save rate, shares, comments"),
    ("Facebook", "Start discussion instead of posting hard brand ads.", "3-5 posts/week", "Discussion posts, polls, longer experience posts", "Question-led titles: Is this worth importing from China in 2026? Does this MOQ make sense?", "Comments, group interaction, DMs"),
    ("YouTube Shorts", "Reuse short videos and build recommendation/search surface.", "3-5/week", "Mistake / Is It Worth Importing? / China vs Brazil", "Searchable titles: How to calculate real profit importing from China.", "Watch time, subscriptions, traffic"),
    ("YouTube Long", "Build search-driven trust assets.", "1/week", "Complete tutorials", "Do not make company intro videos. Make: How to calculate real profit / How to find reliable suppliers / How much importing costs.", "Search traffic, watch time"),
]
replace_sheet("Platform Playbook_EN", ["Platform", "June Job", "Frequency", "Content Types", "Writing / Production Notes", "Key Metrics"], platform_rows, {"A": 20, "B": 42, "C": 16, "D": 36, "E": 60, "F": 28})


copy_rows = [
    (
        "2026-05-28",
        "A low China price is not real profit",
        "A low China price does not mean real profit.\n\nBefore placing an order, online sellers need to calculate the full cost: product price, international freight, import taxes, marketplace fees, packaging, returns and cash-flow risk.\n\nThe cheapest supplier is not always the safest decision.\n\nA smarter way to source is coming.\n\nLaunching June 1.",
        "#Zhutova #ChinaSourcing #ImportFromChina #ProductSourcing #EcommerceSellers #OnlineSellers #SupplierSourcing #MarketplaceSeller #CrossBorderCommerce #SmallBusiness",
        r"D:\zhutova 2清洗 正确\社媒\instagram_warmup_2026_05_28_31_EN",
    ),
    (
        "2026-05-29",
        "Do not choose a supplier by price alone",
        "Do not choose a supplier by price alone.\n\nA reliable supplier should give clear answers before you send money or commit to a large order.\n\nBefore placing your first order, check:\n- Real product photos\n- Sample availability\n- Clear MOQ\n- Export experience\n- Consistent communication\n\nBetter supplier decisions start before the payment.\n\nZhutova launches June 1.",
        "#Zhutova #ChinaSourcing #SupplierSourcing #ImportFromChina #ProductSourcing #EcommerceSellers #OnlineSellers #MarketplaceSeller #CrossBorderCommerce #SmallBusiness",
        r"D:\zhutova 2清洗 正确\社媒\instagram_warmup_2026_05_28_31_EN",
    ),
    (
        "2026-05-30",
        "MOQ can protect or hurt your cash flow",
        "MOQ can protect or hurt your cash flow.\n\nA lower unit price can look attractive, but a large minimum order can lock too much money into untested inventory.\n\nBefore scaling, sellers should ask:\n- Can I test demand with a smaller quantity?\n- Can I sell this stock fast enough?\n- Will freight and fees still make sense?\n- What happens if the product does not move?\n\nThe right first order is not always the cheapest one. It is the one you can validate safely.\n\nZhutova launches June 1.",
        "#Zhutova #ChinaSourcing #MOQ #ImportFromChina #ProductSourcing #EcommerceSellers #OnlineSellers #MarketplaceSeller #InventoryManagement #CrossBorderCommerce",
        r"D:\zhutova 2清洗 正确\社媒\instagram_warmup_2026_05_28_31_EN",
    ),
    (
        "2026-05-31",
        "Zhutova launches tomorrow",
        "Zhutova launches tomorrow.\n\nA sourcing platform for online sellers who want to import from China with more clarity, less guesswork, and better decisions before buying.\n\nWith Zhutova, sellers can:\n- Find product opportunities\n- Check supplier reliability\n- Understand MOQ before ordering\n- Estimate real landed cost\n- Source with less risk\n\nTomorrow, a smarter way to source begins.\n\nLaunching June 1.",
        "#Zhutova #ChinaSourcing #ImportFromChina #ProductSourcing #SupplierSourcing #EcommerceSellers #OnlineSellers #MarketplaceSeller #CrossBorderCommerce #SmallBusiness",
        r"D:\zhutova 2清洗 正确\社媒\instagram_warmup_2026_05_28_31_EN",
    ),
    (
        "2026-06-01",
        "Stop importing in the dark",
        "Stop importing in the dark.\n\nStarting today, Zhutova helps Brazilian sellers make better decisions before importing: product, supplier, MOQ, freight, taxes, margin, and risk.\n\nBuying cheap is not enough. What matters is whether profit still exists after the full cost calculation.\n\nWant Zhutova to analyze your product? Send it to us on WhatsApp.",
        "#Zhutova #ChinaSourcing #ImportFromChina #ProductSourcing #SupplierSourcing #EcommerceSellers #OnlineSellers #MarketplaceSeller #CrossBorderCommerce #ChinaBrazil",
        r"D:\zhutova 2清洗 正确\社媒",
    ),
]
replace_sheet("Warmup Launch Copy_EN", ["Date / Scenario", "Creative Theme", "English Caption (Ready to Post)", "Hashtags", "Asset Folder"], copy_rows, {"A": 16, "B": 34, "C": 82, "D": 52, "E": 58})


# Reorder English planning sheets first.
front = ["Tomorrow Plan_May28_EN", "Overview_EN", "Content IP_EN", "June Daily Plan_EN", "Platform Playbook_EN", "Warmup Launch Copy_EN"]
for idx, name in enumerate(front):
    if name in wb.sheetnames:
        obj = wb[name]
        wb._sheets.remove(obj)
        wb._sheets.insert(idx, obj)

for ws in wb.worksheets:
    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)

try:
    wb.save(XLSX)
    print(XLSX)
except PermissionError:
    fallback = ROOT / "zhutova社媒规划_新版判断品牌计划_EN.xlsx"
    wb.save(fallback)
    print(fallback)
