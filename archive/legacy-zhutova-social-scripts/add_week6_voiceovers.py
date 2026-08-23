from pathlib import Path

from openpyxl import load_workbook


WORKBOOK = Path(r"D:\zhutova 2清洗 正确\社媒\Zhutova_社媒运营计划_简洁版.xlsx")
WEEK = "第6周"

NEWS_RAW = """大家好，这里是 ZHUTOVA 一分钟跨境供应链周报。

第一条，TikTok Shop 美国站把直播拍卖正式加入大促玩法。
以后拼的不只是价格，而是互动和供应链速度。
如果你的产品能快速出样、小批量供货，就更容易抓住流量机会。

第二条，eBay 从 7 月开始调整卖家佣金。
不同品类费率重新划分，二手、翻新产品成本更有优势。
做跨境的卖家，建议重新算一遍利润，而不是继续按老费率运营。

第三条，SHEIN 美国站全面限制自发货。
越来越多平台开始要求商家使用平台合作物流和海外仓。
未来竞争的不只是产品，而是谁的履约能力更稳定。

这就是本周值得关注的三件事。

关注 ZHUTOVA，每周一分钟，带你了解最新跨境供应链动态。"""

TOPICS = [
    "实时行业热点新闻：ZHUTOVA 一分钟跨境供应链周报",
    "产品精选：适合直播大促的小件产品怎么判断？",
    "供应链与履约：为什么小批量快反会越来越重要？",
    "跨境贸易风险与踩坑：平台佣金调整后，为什么要重算利润？",
    "案例与商务合作 vlog：海外仓和平台物流，到底要提前看什么？",
]

VOICEOVERS = [
    (
        WEEK,
        TOPICS[0],
        "实时行业热点新闻",
        """本周这三件事，
都在提醒跨境卖家：
平台越来越看重供应链速度和履约能力。

这里是 ZHUTOVA 一分钟跨境供应链周报。

第一，TikTok Shop 美国站把直播拍卖，
正式加入大促玩法。

以后拼的不只是价格，
还有互动能力和供应链速度。

如果你的产品能快速出样，
还能小批量供货，
就更容易抓住直播流量机会。

第二，eBay 从 7 月开始调整卖家佣金。
不同品类费率重新划分，
二手、翻新产品的成本优势会更明显。

做跨境的卖家，
建议重新算一遍利润，
不要继续按老费率运营。

第三，SHEIN 美国站全面限制自发货。
越来越多平台开始要求商家使用平台合作物流和海外仓。

未来竞争的不只是产品，
而是谁的履约能力更稳定。

关注 ZHUTOVA，
每周一分钟，
带你了解最新跨境供应链动态。""",
        "实时行业热点新闻 / 供应链周报 / 60秒",
    ),
    (
        WEEK,
        TOPICS[1],
        "产品精选",
        """直播大促里的产品，
不是越便宜越好卖。

真正适合直播场景的产品，
要先看三个点。

第一，展示效果要直接。
观众几秒钟内，
就要看懂它有什么用。

第二，体积和重量要可控。
否则订单起来以后，
物流成本会吃掉利润。

第三，供应要能跟上。
直播爆单不是最难的，
爆单以后还能不能补货，
才是真正考验供应链的地方。

所以 Zhutova 做产品精选，
不会只看价格和热度。

还会看这个产品，
能不能适应内容场景、物流成本和后续供货。""",
        "产品精选 / 直播大促选品 / 30-45秒",
    ),
    (
        WEEK,
        TOPICS[2],
        "供应链与履约",
        """现在跨境电商越来越需要小批量快反。

原因很简单。

平台流量变化快，
爆品周期也越来越短。

如果一开始就压很大库存，
风险会很高。

但如果供应链只能大批量生产，
又很难快速跟上市场变化。

所以更理想的方式，
是先小批量测试，
看到数据以后再快速补货。

这背后需要的，
不是单纯找一个便宜工厂。

而是供应商配合度、生产节奏、仓储和物流，
都能跟得上。

这也是 Zhutova 一直强调供应链协同的原因。""",
        "供应链与履约 / 小批量快反 / 30-45秒",
    ),
    (
        WEEK,
        TOPICS[3],
        "跨境贸易风险与踩坑",
        """平台佣金一调整，
很多产品的利润就要重新算。

千万不要继续用老成本表做判断。

因为跨境利润，
不是只看采购价和销售价。

平台佣金、物流费、仓储费、退货成本，
任何一个变化，
都会影响最后利润。

特别是二手、翻新、低客单价产品，
费率变化以后，
可能从能做变成不值得做，
也可能出现新的机会。

所以卖家每次看到平台规则变化，
第一件事不是马上跟风选品。

而是重新算一遍完整成本。""",
        "跨境贸易风险与踩坑 / 成本重算 / 30-45秒",
    ),
    (
        WEEK,
        TOPICS[4],
        "案例与商务合作",
        """现在越来越多平台，
开始限制自发货，
或者要求使用合作物流和海外仓。

这对卖家来说，
不是简单换一个发货方式。

你要提前看清楚几个问题。

第一，产品适不适合提前备货。
第二，海外仓成本能不能接受。
第三，补货周期会不会太长。
第四，退货和售后怎么处理。

如果这些没提前算清楚，
订单越多，
压力反而越大。

所以 Zhutova 看海外仓和平台物流，
不是只看能不能发货，
而是看它能不能支撑长期稳定经营。""",
        "案例与商务合作 vlog / 海外仓与平台物流 / 30-45秒",
    ),
]


def upsert_news_input(wb):
    ws = wb["07_新闻热点输入"]
    target_row = None
    for row_idx in range(2, ws.max_row + 1):
        if ws.cell(row_idx, 1).value == WEEK:
            target_row = row_idx
            break
    if target_row is None:
        target_row = ws.max_row + 1

    values = [
        WEEK,
        "ZHUTOVA 一分钟跨境供应链周报",
        "TikTok Shop 美国站把直播拍卖正式加入大促玩法，供应链速度更重要。",
        "eBay 7 月开始调整卖家佣金，卖家需要重新计算利润。",
        "SHEIN 美国站全面限制自发货，平台物流和海外仓能力更重要。",
        NEWS_RAW,
        "用于每周第1条：实时行业热点新闻",
        "本周新闻由用户提供，口播已做短视频化整理。",
    ]
    for col_idx, value in enumerate(values, start=1):
        ws.cell(target_row, col_idx).value = value
    ws.row_dimensions[target_row].height = 120


def upsert_week_plan(wb):
    ws = wb["04_未来4周主题安排"]
    target_row = None
    for row_idx in range(2, ws.max_row + 1):
        if ws.cell(row_idx, 1).value == WEEK:
            target_row = row_idx
            break
    if target_row is None:
        target_row = ws.max_row + 1

    for col_idx, value in enumerate([WEEK, *TOPICS], start=1):
        ws.cell(target_row, col_idx).value = value
    ws.row_dimensions[target_row].height = 98


def upsert_voiceovers(wb):
    ws = wb["05_口播文案库"]
    existing = []
    for row_idx in range(2, ws.max_row + 1):
        if ws.cell(row_idx, 1).value == WEEK:
            existing.append(row_idx)

    if existing:
        start = existing[0]
        for i, row in enumerate(VOICEOVERS):
            target = start + i
            for col_idx, value in enumerate(row, start=1):
                ws.cell(target, col_idx).value = value
            ws.row_dimensions[target].height = 170 if i == 0 else 145
    else:
        for i, row in enumerate(VOICEOVERS):
            ws.append(row)
            ws.row_dimensions[ws.max_row].height = 170 if i == 0 else 145


def main():
    wb = load_workbook(WORKBOOK)
    upsert_news_input(wb)
    upsert_week_plan(wb)
    upsert_voiceovers(wb)
    wb.save(WORKBOOK)
    print(WORKBOOK)


if __name__ == "__main__":
    main()
