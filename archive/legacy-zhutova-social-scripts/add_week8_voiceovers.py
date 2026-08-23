from pathlib import Path

from openpyxl import load_workbook


WORKBOOK = Path(r"D:\zhutova 2清洗 正确\社媒\Zhutova_社媒运营计划_简洁版.xlsx")
FALLBACK_WORKBOOK = Path(r"D:\zhutova 2清洗 正确\社媒\Zhutova_社媒运营计划_简洁版_第8周口播.xlsx")
WEEK = "第8周"

NEWS_RAW = """大家好，这里是 ZHUTOVA 一分钟跨境供应链周报。

第一条，亚马逊美国站进一步收紧电子产品合规要求。
带电产品的安全申报正在进入强执行阶段。
如果你做 3C、智能硬件或消费电子，现在不仅要有产品，更要提前准备完整的认证和合规资料。

第二条，欧洲继续加强对非欧盟卖家的监管。
未来平台、清关和产品合规都会越来越严格。
想长期做好欧洲市场，除了价格，更要提前做好标签、包装和产品合规。

本周最大的变化只有一个：
跨境竞争，正在从价格竞争，转向供应链和合规能力的竞争。

关注 ZHUTOVA，每周一分钟，带你了解最新跨境供应链动态。"""

TOPICS = [
    "实时行业热点新闻：ZHUTOVA 一分钟跨境供应链周报",
    "产品精选：做3C和智能硬件，选品前要先看什么？",
    "供应链与履约：为什么合规资料要提前准备？",
    "跨境贸易风险与踩坑：只看低价做欧洲市场，为什么越来越危险？",
    "案例与商务合作 vlog：产品标签和包装，出货前要检查什么？",
]

VOICEOVERS = [
    (
        WEEK,
        TOPICS[0],
        "实时行业热点新闻",
        """这周跨境卖家最该关注的，
不是哪一个产品又火了。

而是平台正在把合规要求，
推到更强执行阶段。

这里是 ZHUTOVA 一分钟跨境供应链周报。

第一，亚马逊美国站进一步收紧电子产品合规要求。

带电产品的安全申报，
正在进入强执行阶段。

如果你做 3C、智能硬件，或者消费电子，
现在不仅要有产品，
更要提前准备完整的认证和合规资料。

第二，欧洲继续加强对非欧盟卖家的监管。

未来平台、清关和产品合规，
都会越来越严格。

想长期做好欧洲市场，
除了价格，
更要提前做好标签、包装和产品合规。

本周最大的变化只有一个：
跨境竞争，
正在从价格竞争，
转向供应链和合规能力的竞争。

关注 ZHUTOVA，
每周一分钟，
带你了解最新跨境供应链动态。""",
        "实时行业热点新闻 / 合规周报 / 60秒",
    ),
    (
        WEEK,
        TOPICS[1],
        "产品精选",
        """3C 和智能硬件，
看起来很适合跨境。

客单价不低，
需求也很稳定。

但这类产品，
不是看到供应价合适就能做。

选品前，
至少要先看四件事。

第一，是否带电。
第二，是否需要安全认证。
第三，说明书、标签和包装，
能不能匹配目标市场。
第四，售后和退货风险高不高。

如果这些没提前看清楚，
产品再有需求，
后面也可能卡在合规和平台审核上。

所以 Zhutova 做产品精选，
不是只看价格和外观。

还要看这个产品，
能不能合规、稳定地进入市场。""",
        "产品精选 / 3C智能硬件选品 / 30-45秒",
    ),
    (
        WEEK,
        TOPICS[2],
        "供应链与履约",
        """合规资料，
真的不能等到出单以后再补。

尤其是带电产品、电子产品、智能硬件，
平台和清关都会越来越重视资料完整性。

如果认证、检测报告、标签、说明书，
一开始就没准备好，
后面可能不是发货慢的问题。

而是产品无法上架，
订单无法正常履约，
甚至货已经备好了，
却卡在审核和清关环节。

所以 Zhutova 看供应链，
不只是看工厂能不能生产。

还要看资料、包装、标签和履约要求，
能不能提前配合好。

这才是长期做跨境，
真正需要的稳定能力。""",
        "供应链与履约 / 合规资料准备 / 30-45秒",
    ),
    (
        WEEK,
        TOPICS[3],
        "跨境贸易风险与踩坑",
        """现在做欧洲市场，
只看低价会越来越危险。

因为平台、清关、产品合规，
都在变得更严格。

有些产品在中国看起来很便宜，
但到了欧洲市场，
可能要补标签，
要改包装，
要做认证，
还要满足平台审核。

这些成本如果一开始没算进去，
最后利润很容易被吃掉。

更麻烦的是，
如果资料不合规，
产品可能根本上不了架，
或者货到了也不好卖。

所以做欧洲市场，
不能只问价格低不低。

还要问，
这个产品能不能合规、稳定、长期地卖。""",
        "跨境贸易风险与踩坑 / 欧洲合规风险 / 30-45秒",
    ),
    (
        WEEK,
        TOPICS[4],
        "案例与商务合作",
        """出货前检查产品，
不要只看货做没做好。

标签和包装，
同样要提前看。

比如产品标签有没有目标市场需要的信息，
包装上有没有正确的警示说明，
说明书语言是否匹配，
条码、型号、认证信息是否一致。

这些细节看起来很小，
但一旦出错，
可能会影响上架、清关、售后，
甚至影响整批货的交付。

所以 Zhutova 做供应链对接，
不是只看产品本身。

也会把标签、包装、资料和交付要求，
一起放进出货前检查里。

这样客户后面进入市场时，
才会更稳。""",
        "案例与商务合作 vlog / 标签包装检查 / 30-45秒",
    ),
]


def upsert_news_input(wb):
    ws = wb["07_新闻热点输入"]
    target_row = None
    for row_idx in range(2, ws.max_row + 1):
        if ws.cell(row_idx, 1).value == WEEK:
            target_row = row_idx
            break
    if target_row is None:
        target_row = ws.max_row + 1

    values = [
        WEEK,
        "ZHUTOVA 一分钟跨境供应链周报",
        "亚马逊美国站进一步收紧电子产品合规要求，带电产品安全申报进入强执行阶段。",
        "欧洲继续加强对非欧盟卖家的监管，平台、清关和产品合规要求更严格。",
        "",
        NEWS_RAW,
        "用于每周第1条：实时行业热点新闻",
        "本周新闻由用户提供，口播已围绕合规和供应链能力延展。",
    ]
    for col_idx, value in enumerate(values, start=1):
        ws.cell(target_row, col_idx).value = value
    ws.row_dimensions[target_row].height = 120


def upsert_week_plan(wb):
    ws = wb["04_未来4周主题安排"]
    target_row = None
    for row_idx in range(2, ws.max_row + 1):
        if ws.cell(row_idx, 1).value == WEEK:
            target_row = row_idx
            break
    if target_row is None:
        target_row = ws.max_row + 1
    for col_idx, value in enumerate([WEEK, *TOPICS], start=1):
        ws.cell(target_row, col_idx).value = value
    ws.row_dimensions[target_row].height = 98


def upsert_voiceovers(wb):
    ws = wb["05_口播文案库"]
    existing = []
    for row_idx in range(2, ws.max_row + 1):
        if ws.cell(row_idx, 1).value == WEEK:
            existing.append(row_idx)

    if existing:
        start = existing[0]
        for i, row in enumerate(VOICEOVERS):
            target = start + i
            for col_idx, value in enumerate(row, start=1):
                ws.cell(target, col_idx).value = value
            ws.row_dimensions[target].height = 170 if i == 0 else 145
    else:
        for i, row in enumerate(VOICEOVERS):
            ws.append(row)
            ws.row_dimensions[ws.max_row].height = 170 if i == 0 else 145


def main():
    wb = load_workbook(WORKBOOK)
    upsert_news_input(wb)
    upsert_week_plan(wb)
    upsert_voiceovers(wb)
    try:
        wb.save(WORKBOOK)
        print(WORKBOOK)
    except PermissionError:
        wb.save(FALLBACK_WORKBOOK)
        print(FALLBACK_WORKBOOK)


if __name__ == "__main__":
    main()
