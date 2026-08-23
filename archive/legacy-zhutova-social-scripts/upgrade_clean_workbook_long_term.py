from pathlib import Path
from datetime import date, timedelta
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side


ROOT = Path(r"D:\zhutova 2清洗 正确")
OUT = ROOT / "社媒" / "Zhutova_社媒运营计划_简洁版.xlsx"

wb = openpyxl.load_workbook(OUT)


def replace_sheet(name, headers, rows, widths, index=None):
    if name in wb.sheetnames:
        ws = wb[name]
        ws.delete_rows(1, ws.max_row)
    else:
        ws = wb.create_sheet(name, index if index is not None else None)

    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A2"
    ws.append(headers)
    for row in rows:
        ws.append(row)

    thin = Side(style="thin", color="D9E2EC")
    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.border = Border(bottom=thin)

    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="0A1628")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for col, width in widths.items():
        ws.column_dimensions[col].width = width
    ws.row_dimensions[1].height = 30
    for r in range(2, ws.max_row + 1):
        ws.row_dimensions[r].height = 92
    return ws


# 1. Long-term roadmap
roadmap_rows = [
    ("第1阶段", "第1-2周", "建立认知", "让用户记住：Zhutova 懂进口利润和风险", "低价不等于利润、供应商风险、MOQ、真实成本", "播放量、完播率、评论、收藏"),
    ("第2阶段", "第3-4周", "建立信任", "让用户知道 Zhutova 是一站式跨境 B2B 平台", "平台如何工作、工厂直供、本地仓储、智能履约、运营幕后", "收藏、主页访问、私信、WhatsApp点击"),
    ("第3阶段", "第2个月", "建立专业资产", "把 Zhutova 从内容号变成卖家采购前会想到的判断品牌", "产品案例、成本模板、供应商清单、品类分析、FAQ", "产品分析请求、私信质量、WhatsApp咨询"),
    ("第4阶段", "第3个月", "转化与商务合作", "用案例和平台能力承接 B2B 合作、供应链合作和卖家咨询", "客户案例、合作伙伴、平台生态、定制采购、供应链金融", "商务询盘、合作线索、复购咨询"),
]
replace_sheet(
    "02_长期运营路线",
    ["阶段", "周期", "核心目标", "要建立的用户认知", "主要内容", "核心指标"],
    roadmap_rows,
    {"A": 14, "B": 14, "C": 22, "D": 52, "E": 58, "F": 32},
    index=1,
)


# 2. Monthly themes
monthly_rows = [
    ("第1月", "不要再盲目进口", "建立 Zhutova 的判断力：产品是否值得进口、供应商是否可靠、MOQ是否合理、真实成本是否还有利润。", "爆点流量 60%；品牌背书 40%", "低价不等于利润 / 供应商风险 / MOQ / China vs Brazil / Zhutova 是什么"),
    ("第2月", "从判断到执行", "让用户看到 Zhutova 不只会判断，还能支持工厂直供、全球贸易、本地仓储、智能履约。", "平台能力 50%；案例 30%；爆点 20%", "How Zhutova Works / Behind the Supply Chain / 订单和物流 / 本地仓 / 服务支持"),
    ("第3月", "案例与合作", "用案例、合作伙伴和生态能力建立商务信任，开始承接更高质量咨询。", "案例 40%；商务背书 40%；爆点 20%", "客户案例 / 模拟案例 / 供应链金融 / 商务合作 / 平台生态"),
]
replace_sheet(
    "03_月度主题",
    ["月份", "月度主题", "运营重点", "内容比例", "代表选题"],
    monthly_rows,
    {"A": 12, "B": 24, "C": 70, "D": 28, "E": 62},
    index=2,
)


# 3. Weekly repeatable structure
weekly_rows = [
    ("周一", "爆点流量", "产品/利润误区", "低价不等于利润、爆品不等于赚钱", "TikTok / Reels / Shorts", "拿播放量"),
    ("周二", "品牌背书", "平台介绍/能力", "Zhutova 是什么、平台如何工作", "Reels / Facebook / LinkedIn", "建立信任"),
    ("周三", "爆点流量", "供应商风险", "便宜供应商、样品、MOQ、沟通风险", "TikTok / Reels / Facebook", "引发代入"),
    ("周四", "品牌背书", "供应链幕后", "工厂直供、仓储、物流、履约、QC", "Reels / LinkedIn / Facebook", "证明能力"),
    ("周五", "爆点流量", "现金流/MOQ/成本", "MOQ压库存、运费税费、平台费吃利润", "TikTok / Reels / Shorts", "提高收藏"),
    ("周六", "案例内容", "客户案例/模拟案例", "一个产品如何判断是否值得进口", "Reels / TikTok / Facebook", "建立实用价值"),
    ("周日", "商务背书", "生态/合作/总结", "合作伙伴、平台生态、商务合作、本周总结", "LinkedIn / Facebook / Reels", "承接B2B询盘"),
]
replace_sheet(
    "04_每周循环模板",
    ["星期", "内容风格", "内容方向", "可选主题", "建议平台", "作用"],
    weekly_rows,
    {"A": 10, "B": 16, "C": 24, "D": 60, "E": 30, "F": 24},
    index=3,
)


# 4. 30-day content calendar
start = date(2026, 6, 10)
weekday_map = {
    0: ("爆点流量", "产品/利润误区", "很多卖家进口亏钱，不是因为产品不好。"),
    1: ("品牌背书", "平台介绍/能力", "Zhutova 不只是 sourcing。"),
    2: ("爆点流量", "供应商风险", "便宜供应商，可能是最贵的错误。"),
    3: ("品牌背书", "供应链幕后", "找到工厂，不代表采购结束。"),
    4: ("爆点流量", "MOQ/现金流", "第一单买太多，可能直接压死现金流。"),
    5: ("案例内容", "客户案例/模拟案例", "这个产品不是不能做，是不能这样做。"),
    6: ("商务背书", "生态/合作/总结", "Zhutova 连接的不只是产品。"),
}
calendar_rows = []
topics = [
    "中国低价不等于真实利润",
    "Zhutova 是一站式跨境 B2B 平台",
    "供应商低价背后的隐藏风险",
    "工厂直供到智能履约的链路",
    "MOQ 如何影响现金流",
    "模拟案例：手机配件是否值得进口",
    "Zhutova 平台生态与商务合作",
    "爆品为什么不一定赚钱",
    "Zhutova 如何帮助卖家做采购判断",
    "样品、图片和规格为什么重要",
    "仓储和履约如何影响利润",
    "运费税费平台费如何吃掉利润",
    "模拟案例：小批量测试 vs 大单采购",
    "本周进口风险总结",
    "价格差不是利润",
    "Zhutova 的全品类供应能力",
    "如何判断供应商沟通是否可靠",
    "订单和物流节点如何管理",
    "第一单应该追求低价还是验证需求",
    "模拟案例：产品值得做但不适合大批量",
    "商务合作：谁适合找 Zhutova",
    "真实到岸成本怎么算",
    "从产品机会到采购执行",
    "供应商红旗清单",
    "本地仓储的价值",
    "库存积压如何毁掉利润",
    "模拟案例：卖家下单前我们检查什么",
    "合作伙伴和全球市场连接",
    "进口前最该问的5个问题",
    "月度总结：不要再盲目进口",
]
for i in range(30):
    d = start + timedelta(days=i)
    style, direction, hook = weekday_map[d.weekday()]
    calendar_rows.append((
        d.strftime("%Y-%m-%d"),
        ["周一", "周二", "周三", "周四", "周五", "周六", "周日"][d.weekday()],
        style,
        direction,
        topics[i],
        hook,
        "短视频 20-30 秒；先讲痛点/误区，再带出 Zhutova 的判断或平台能力。",
        "想让 Zhutova 分析你的产品？通过 WhatsApp 发给我们。",
    ))
replace_sheet(
    "05_未来30天内容排期",
    ["日期", "星期", "风格", "内容方向", "当天选题", "开头Hook", "制作说明", "CTA"],
    calendar_rows,
    {"A": 14, "B": 10, "C": 16, "D": 22, "E": 42, "F": 48, "G": 58, "H": 42},
    index=4,
)


# 5. Topic bank
topic_rows = [
    ("产品/利润判断", "中国低价不等于利润", "爆点流量", "采购价只是第一步，真实利润要看完整成本。"),
    ("产品/利润判断", "爆品不等于赚钱", "爆点流量", "产品火不代表利润好，关键是成本和履约。"),
    ("产品/利润判断", "价格差不是利润", "爆点流量", "China vs Brazil 价格差要加运费税费平台费。"),
    ("供应商判断", "不要只按价格选供应商", "爆点流量", "低价供应商可能带来质量、交期、售后风险。"),
    ("供应商判断", "供应商红旗清单", "实用收藏", "没有真实图、不寄样、MOQ模糊、沟通不稳定。"),
    ("供应商判断", "样品为什么重要", "实用收藏", "样品能提前暴露质量、包装、规格问题。"),
    ("MOQ/现金流", "第一单不要买太多", "爆点流量", "测试需求比压低单价更重要。"),
    ("MOQ/现金流", "MOQ 如何谈", "实用收藏", "用测试单、组合单、分批采购降低风险。"),
    ("平台能力", "Zhutova 是什么", "品牌背书", "一站式跨境 B2B 平台。"),
    ("平台能力", "How Zhutova Works", "品牌背书", "工厂直供、全球贸易、本地仓储、智能履约。"),
    ("平台能力", "全品类供应不等于都值得做", "品牌+判断", "平台有供应能力，但判断更重要。"),
    ("运营幕后", "下单前检查流程", "信任建设", "产品规格、供应商、样品、成本、物流。"),
    ("运营幕后", "物流节点如何影响利润", "信任建设", "交期、仓储、履约都会影响实际利润。"),
    ("案例", "产品不是不能做，是不能这样做", "流量+信任", "用模拟案例展示判断逻辑。"),
    ("案例", "小批量测试 vs 大单采购", "流量+信任", "不同阶段的采购策略不同。"),
    ("商务合作", "谁适合和 Zhutova 合作", "商务背书", "卖家、品牌方、供应商、渠道伙伴。"),
    ("商务合作", "Zhutova 生态连接", "商务背书", "销售渠道、供应链服务、仓储、支付、履约。"),
]
replace_sheet(
    "06_长期选题库",
    ["内容类别", "选题", "建议风格", "核心观点"],
    topic_rows,
    {"A": 18, "B": 38, "C": 18, "D": 72},
    index=5,
)


# Reorder core sheets
front = [
    "01_运营总览",
    "02_长期运营路线",
    "03_月度主题",
    "04_每周循环模板",
    "05_未来30天内容排期",
    "06_长期选题库",
    "02_一周视频计划",
    "03_视频口播文案",
    "04_素材清单",
    "05_发布复盘",
]
for idx, name in enumerate(front):
    if name in wb.sheetnames:
        obj = wb[name]
        wb._sheets.remove(obj)
        wb._sheets.insert(idx, obj)

try:
    wb.save(OUT)
    print(OUT)
except PermissionError:
    fallback = ROOT / "社媒" / "Zhutova_社媒运营计划_长期版.xlsx"
    wb.save(fallback)
    print(fallback)
