from pathlib import Path

from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side


WORKBOOK = Path(r"D:\zhutova 2清洗 正确\社媒\Zhutova_社媒运营计划_简洁版.xlsx")


def style_header_and_column(ws):
    thin = Side(style="thin", color="D9E2EC")
    ws.cell(1, 6).font = Font(bold=True, color="FFFFFF")
    ws.cell(1, 6).fill = PatternFill("solid", fgColor="0A1628")
    ws.cell(1, 6).alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.column_dimensions["F"].width = 20
    for row_idx in range(2, ws.max_row + 1):
        cell = ws.cell(row_idx, 6)
        cell.alignment = Alignment(vertical="top", wrap_text=True)
        cell.border = Border(bottom=thin)


def main():
    wb = load_workbook(WORKBOOK)

    ws01 = wb["01_核心思路"]
    ws02 = wb["02_9大栏目主题池"]

    ws01.insert_rows(6)
    ws01["A6"] = "内容属性"
    ws01["B6"] = "“Zhutova 是什么 / 平台如何运作”这类介绍内容，属于一次性基础资产内容，适合集中拍摄后反复穿插使用；热点新闻、产品精选、供应链与履约、跨境贸易风险与踩坑、案例与商务合作 vlog 属于长期周更内容。"

    ws02.insert_cols(6)
    ws02["F1"] = "内容属性"

    mapping = {
        "实时行业热点新闻": "长期周更内容",
        "Zhutova 是什么": "一次性基础资产",
        "平台如何运作": "一次性基础资产",
        "产品与供应机会": "长期周更内容",
        "供应链与履约": "长期周更内容",
        "案例与商务合作": "长期周更内容",
        "利润判断": "长期周更内容",
        "风险与踩坑": "长期周更内容",
        "中国采购现场": "长期周更内容",
    }

    for row_idx in range(2, ws02.max_row + 1):
        name = ws02.cell(row_idx, 1).value
        ws02.cell(row_idx, 6).value = mapping.get(name, "")

    style_header_and_column(ws02)
    wb.save(WORKBOOK)
    print(WORKBOOK)


if __name__ == "__main__":
    main()
