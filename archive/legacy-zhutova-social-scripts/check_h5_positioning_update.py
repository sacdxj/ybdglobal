from pathlib import Path
import openpyxl

paths = [
    Path(r"D:\zhutova 2清洗 正确\社媒\zhutova社媒规划_新版判断品牌计划_中文版本.xlsx"),
    Path(r"D:\zhutova 2清洗 正确\社媒\zhutova社媒规划_新版判断品牌计划_中文版本_H5定位校准版.xlsx"),
    Path(r"D:\zhutova 2清洗 正确\社媒\Zhutova_社媒内容汇总_2026_06\00_社媒规划\zhutova社媒规划_中文版本.xlsx"),
]

for path in paths:
    if not path.exists():
        continue
    wb = openpyxl.load_workbook(path, data_only=False)
    print("FILE", path.name)
    print(wb.sheetnames[:10])
    print("overview", wb["新版总览"]["B2"].value[:90])
    print("calendar headers", [wb["6月每日排期"].cell(1, c).value for c in range(1, 5)])
    print("calendar row2", [wb["6月每日排期"].cell(2, c).value for c in range(1, 5)])
