from pathlib import Path
import openpyxl


ROOT = Path(r"D:\zhutova 2清洗 正确")
BUNDLE = ROOT / "社媒" / "Zhutova_社媒内容汇总_2026_06"
PLAN_DIR = BUNDLE / "00_社媒规划"

asset_map = {
    "2026-05-28": BUNDLE / "01_预热图_EN" / "2026-05-28_warmup_low-price.png",
    "2026-05-29": BUNDLE / "01_预热图_EN" / "2026-05-29_warmup_supplier-check.png",
    "2026-05-30": BUNDLE / "01_预热图_EN" / "2026-05-30_warmup_moq.png",
    "2026-05-31": BUNDLE / "01_预热图_EN" / "2026-05-31_warmup_launch-tomorrow.png",
    "2026-06-01": BUNDLE / "03_上线轮播_EN",
}


def update_workbook(path: Path):
    wb = openpyxl.load_workbook(path)
    if "预热与上线文案" in wb.sheetnames:
        ws = wb["预热与上线文案"]
        headers = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
        image_col = None
        folder_col = None
        for idx, h in enumerate(headers, 1):
            if h in ("Post Image / Asset", "发布图片/素材"):
                image_col = idx
            if h in ("Asset Folder", "素材文件夹"):
                folder_col = idx
        for row in range(2, ws.max_row + 1):
            date_key = str(ws.cell(row, 1).value)
            asset = asset_map.get(date_key)
            if asset and image_col:
                cell = ws.cell(row, image_col)
                cell.value = str(asset)
                cell.hyperlink = str(asset)
                cell.style = "Hyperlink"
            if asset and folder_col:
                folder = asset if asset.is_dir() else asset.parent
                ws.cell(row, folder_col).value = str(folder)

    # Tomorrow execution sheets may include asset columns.
    for sheet_name in ("明日计划_5月28", "Tomorrow Plan_May28_EN"):
        if sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            for row in range(1, ws.max_row + 1):
                for col in range(1, ws.max_column + 1):
                    value = ws.cell(row, col).value
                    if isinstance(value, str):
                        for old, new in {
                            r"D:\zhutova 2清洗 正确\社媒\instagram_warmup_2026_05_28_31_EN\2026-05-28_warmup_low-price.png": str(asset_map["2026-05-28"]),
                            r"D:\zhutova 2清洗 正确\社媒\instagram_warmup_2026_05_28_31_PTBR\2026-05-28_aquecimento_preco-baixo.png": str(BUNDLE / "02_预热图_PTBR" / "2026-05-28_aquecimento_preco-baixo.png"),
                        }.items():
                            if old in value:
                                ws.cell(row, col).value = value.replace(old, new)

    wb.save(path)


for workbook in (PLAN_DIR / "zhutova社媒规划_中文版本.xlsx", PLAN_DIR / "zhutova社媒规划_执行版.xlsx"):
    if workbook.exists():
        update_workbook(workbook)

readme = f"""# Zhutova 社媒内容汇总 - 2026年6月

这个文件夹把本次社媒规划和发布素材集中到一起，方便团队管理和发布。

## 目录说明

- `00_社媒规划`：社媒规划表、中文版本、执行版本、原始社媒笔记
- `01_预热图_EN`：5月28日-5月31日英文预热图 + 英文 captions
- `02_预热图_PTBR`：5月28日-5月31日葡语预热图 + 葡语 captions
- `03_上线轮播_EN`：6月1日英文上线 Instagram 轮播图 + caption
- `04_上线轮播_PTBR`：6月1日葡语上线 Instagram 轮播图 + caption
- `05_历史图文素材`：之前制作的 IG 图文、轮播、视频文件
- `06_视频素材`：社媒视频素材

## 推荐使用

1. 内部看中文规划：`00_社媒规划/zhutova社媒规划_中文版本.xlsx`
2. 内容制作/发布看执行版：`00_社媒规划/zhutova社媒规划_执行版.xlsx`
3. 每天发布时看 `预热与上线文案` sheet，对应日期有图片路径、发布文案、hashtags。

核心传播方向：Zhutova 不是单纯网站上线，而是开始帮助巴西卖家判断产品、供应商、利润和风险。
"""

(BUNDLE / "README_先看这个.txt").write_text(readme, encoding="utf-8")
print(BUNDLE)
