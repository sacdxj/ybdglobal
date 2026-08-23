from pathlib import Path
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


ROOT = Path(r"D:\zhutova 2清洗 正确")
FILES = [
    ROOT / "社媒" / "Zhutova_社媒内容汇总_2026_06" / "00_社媒规划" / "zhutova社媒规划_中文版本.xlsx",
    ROOT / "社媒" / "zhutova社媒规划_新版判断品牌计划_中文版本_H5定位校准版.xlsx",
    ROOT / "社媒" / "zhutova社媒规划_新版判断品牌计划_中文版本.xlsx",
]

SHEET = "一周视频规划_每日1条"
COL_NAME = "20秒中文口播文案"

scripts = {
    "Day 1": "Zhutova 不只是一个采购服务。\n\n它是一站式跨境 B2B 平台，连接工厂直供、全球贸易、本地仓储和智能履约。\n\n对线上卖家来说，这不只是找产品，而是在进口前做更好的判断：产品、供应商、MOQ、到岸成本和风险。\n\n不要再盲目进口，用 Zhutova 更聪明地采购。",
    "Day 2": "中国的低采购价，不等于真实利润。\n\n进口前，卖家要算完整成本：产品价格、国际运费、进口税费、平台费用、包装、退货，还有现金流风险。\n\n最便宜的供应商，不一定是最安全的选择。\n\n想让 Zhutova 分析你的产品？发给我们。",
    "Day 3": "找到工厂，不代表采购就结束了。\n\n你还需要贸易支持、仓储、物流和履约能力。\n\nZhutova 把工厂直供、全球贸易、本地仓储和智能履约连接到一个跨境 B2B 平台里。\n\n真正聪明的采购，是全链路判断。",
    "Day 4": "不要只按价格选择供应商。\n\n付款前，要看真实产品图片、是否能寄样品、MOQ 是否清楚、有没有出口经验，以及沟通是否稳定。\n\n便宜供应商，最后可能变成最贵的错误。\n\nZhutova 帮卖家在下单前判断供应商风险。",
    "Day 5": "好的采购，不靠运气，而靠流程。\n\n下单前，卖家应该检查产品规格、供应商沟通、样品情况和真实到岸成本。\n\n之后还要管理订单、物流节点、仓储和服务支持。\n\nZhutova 把采购变成更清晰的流程。",
    "Day 6": "一个卖家问：这个产品值不值得进口？\n\n我们没有先看供应商报价。\n\n我们先看需求、MOQ、运费、税费、平台费用、供应商可靠性和履约风险。\n\n答案不一定是简单的值得或不值得，而是取决于利润、数量和风险。",
    "Day 7": "Zhutova 连接的不只是产品。\n\n它连接选品采购、销售渠道、供应链服务、仓储、支付和履约支持。\n\n对卖家来说，这意味着从产品判断到业务执行，都有更清晰的路径。\n\n如果你想谈业务合作，可以联系 Zhutova。",
}


def style(ws):
    thin = Side(style="thin", color="D9E2EC")
    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.border = Border(bottom=thin)
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="0A1628")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for r in range(2, ws.max_row + 1):
        ws.row_dimensions[r].height = 178


def ensure_file(path: Path):
    if not path.exists():
        return None
    wb = openpyxl.load_workbook(path)
    if SHEET not in wb.sheetnames:
        return "missing sheet"
    ws = wb[SHEET]

    headers = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]

    # Delete existing script column(s) so the final version is placed visibly at D.
    for idx in range(len(headers), 0, -1):
        if headers[idx - 1] == COL_NAME or headers[idx - 1] == "20秒英文口播文案":
            ws.delete_cols(idx)

    # Insert after 视频标题/Hook, which is column C in this sheet.
    insert_col = 4
    ws.insert_cols(insert_col)
    ws.cell(1, insert_col, COL_NAME)

    for row in range(2, ws.max_row + 1):
        day = ws.cell(row, 1).value
        if day in scripts:
            ws.cell(row, insert_col, scripts[day])

    ws.column_dimensions[get_column_letter(insert_col)].width = 72
    style(ws)
    wb.save(path)
    return "updated"


for file in FILES:
    try:
        result = ensure_file(file)
        print(file, result)
    except PermissionError:
        fallback = file.with_name(file.stem + "_补入口播文案版.xlsx")
        wb = openpyxl.load_workbook(file)
        if SHEET in wb.sheetnames:
            ws = wb[SHEET]
            headers = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
            for idx in range(len(headers), 0, -1):
                if headers[idx - 1] == COL_NAME or headers[idx - 1] == "20秒英文口播文案":
                    ws.delete_cols(idx)
            ws.insert_cols(4)
            ws.cell(1, 4, COL_NAME)
            for row in range(2, ws.max_row + 1):
                day = ws.cell(row, 1).value
                if day in scripts:
                    ws.cell(row, 4, scripts[day])
            ws.column_dimensions["D"].width = 72
            style(ws)
        wb.save(fallback)
        print(fallback, "fallback")
