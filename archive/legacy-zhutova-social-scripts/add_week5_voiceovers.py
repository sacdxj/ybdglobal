from pathlib import Path

from openpyxl import load_workbook


WORKBOOK = Path(r"D:\zhutova 2清洗 正确\社媒\Zhutova_社媒运营计划_简洁版.xlsx")


WEEK = "第5周"

TOPICS = [
    "实时行业热点新闻：平台流量红利变少，卖家更要看供应链",
    "产品精选：适合巴西卖家的小家电产品怎么判断？",
    "供应链与履约：为什么发货快不等于履约稳定？",
    "跨境贸易风险与踩坑：只看爆品榜选品，为什么容易踩坑？",
    "案例与商务合作 vlog：一次工厂看样，应该重点看什么？",
]

VOICEOVERS = [
    (
        WEEK,
        TOPICS[0],
        "实时行业热点新闻",
        "最近很多跨境卖家都会发现，\n平台流量越来越不是白送的。\n\n以前一个产品靠短视频爆了，\n可能很快就能出单。\n\n但现在，\n平台竞争更激烈，\n广告、内容、达人、履约，\n都会影响最后的转化。\n\n所以卖家不能只问，\n这个产品能不能火。\n\n更要问，\n如果真的出单了，\n供应能不能跟上，\n物流能不能稳，\n利润还能不能保住。\n\n这也是 Zhutova 一直强调的，\n跨境生意不是只看流量，\n还要看后面的供应链和履约能力。",
        "实时行业热点新闻 / 流量型 / 30-45秒",
    ),
    (
        WEEK,
        TOPICS[1],
        "产品精选",
        "小家电看起来很适合跨境卖家，\n因为客单价不低，\n需求也比较稳定。\n\n但不是所有小家电都适合直接做。\n\n你要先看几个点。\n\n第一，体积和重量会不会让运费太高。\n第二，电压、插头、认证要求能不能匹配当地市场。\n第三，售后风险高不高。\n第四，供应是否稳定，能不能持续补货。\n\n所以产品精选，\n不是看到价格便宜就选。\n\nZhutova 更关注的是，\n这个产品从供应到销售，\n能不能真正跑得起来。",
        "产品精选 / 品类判断 / 30-45秒",
    ),
    (
        WEEK,
        TOPICS[2],
        "供应链与履约",
        "很多人会觉得，\n发货快就代表履约好。\n\n但其实不是。\n\n真正稳定的履约，\n不只是今天能不能发出去，\n还要看库存准不准，\n包装稳不稳，\n物流节点能不能追踪，\n售后问题能不能处理。\n\n如果只追求快，\n但货不对版、包装破损、补货跟不上，\n最后还是会影响客户体验。\n\n所以 Zhutova 看履约，\n看的不是单个动作，\n而是从仓储、物流到交付，\n整条链路能不能稳定配合。",
        "供应链与履约 / 平台执行能力 / 30-45秒",
    ),
    (
        WEEK,
        TOPICS[3],
        "跨境贸易风险与踩坑",
        "很多卖家选品，\n第一反应就是看爆品榜。\n\n这个方法没错，\n但如果只看爆品榜，\n就很容易踩坑。\n\n因为爆品榜只能说明它现在有热度，\n不能说明你拿得到稳定供应，\n也不能说明你的成本结构还能赚钱。\n\n有些产品看起来卖得很好，\n但 MOQ 高、运费贵、退货率高、竞争已经很卷。\n\n所以真正要判断的，\n不是这个产品火不火，\n而是你进入之后，\n还有没有利润和交付空间。",
        "跨境贸易风险与踩坑 / 流量型 / 30-45秒",
    ),
    (
        WEEK,
        TOPICS[4],
        "案例与商务合作",
        "一次工厂看样，\n不要只看样品漂不漂亮。\n\n更重要的是看细节。\n\n比如材料是不是稳定，\n做工有没有差异，\n包装能不能适合跨境运输，\n工厂有没有持续生产能力，\n交期和 MOQ 是否清楚。\n\n很多问题，\n只看图片是看不出来的。\n\n所以 Zhutova 做工厂和供应链对接，\n不是只帮客户看一个产品，\n而是把样品、生产、交付和后续合作，\n一起放进判断里。",
        "案例与商务合作 vlog / 工厂看样素材 / 30-45秒",
    ),
]


def replace_or_append_week_row(ws):
    target_row = None
    for row_idx in range(2, ws.max_row + 1):
        if ws.cell(row_idx, 1).value == WEEK:
            target_row = row_idx
            break
    if target_row is None:
        target_row = ws.max_row + 1
    for col_idx, value in enumerate([WEEK, *TOPICS], start=1):
        ws.cell(target_row, col_idx).value = value
    ws.row_dimensions[target_row].height = 98


def replace_or_append_voiceovers(ws):
    existing = []
    for row_idx in range(2, ws.max_row + 1):
        if ws.cell(row_idx, 1).value == WEEK:
            existing.append(row_idx)

    if existing:
        start = existing[0]
        for i, row in enumerate(VOICEOVERS):
            target_row = start + i
            for col_idx, value in enumerate(row, start=1):
                ws.cell(target_row, col_idx).value = value
            ws.row_dimensions[target_row].height = 142
    else:
        for row in VOICEOVERS:
            ws.append(row)
            ws.row_dimensions[ws.max_row].height = 142


def main():
    wb = load_workbook(WORKBOOK)
    replace_or_append_week_row(wb["04_未来4周主题安排"])
    replace_or_append_voiceovers(wb["05_口播文案库"])
    wb.save(WORKBOOK)
    print(WORKBOOK)


if __name__ == "__main__":
    main()
