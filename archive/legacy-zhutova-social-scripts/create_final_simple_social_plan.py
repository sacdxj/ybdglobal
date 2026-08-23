from pathlib import Path
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side


ROOT = Path(r"D:\zhutova 2清洗 正确")
OUT = ROOT / "社媒" / "Zhutova_社媒运营计划_最终简洁版.xlsx"


def style(ws, widths):
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A2"
    thin = Side(style="thin", color="D9E2EC")
    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.border = Border(bottom=thin)
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="0A1628")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for col, width in widths.items():
        ws.column_dimensions[col].width = width
    ws.row_dimensions[1].height = 30
    for r in range(2, ws.max_row + 1):
        ws.row_dimensions[r].height = 88


def add_sheet(wb, name, headers, rows, widths):
    ws = wb.create_sheet(name)
    ws.append(headers)
    for row in rows:
        ws.append(row)
    style(ws, widths)
    return ws


wb = Workbook()
wb.remove(wb.active)


add_sheet(
    wb,
    "01_核心思路",
    ["项目", "内容"],
    [
        ("一句话定位", "Zhutova 是以供应链判断力切入的一站式跨境 B2B 平台。"),
        ("社媒主张", "不要再盲目进口。"),
        ("社媒要建立的认知", "进口前，先让 Zhutova 判断产品、供应商、成本、MOQ 和风险。"),
        ("内容打法", "用流量型内容吸引卖家，用品牌/商务背书内容建立平台信任。"),
        ("流量型内容作用", "获得播放量、评论、收藏、分享，让用户产生“这和我有关”的感觉。"),
        ("品牌/商务内容作用", "解释 Zhutova 是什么、能做什么、为什么可信，承接商务合作和高质量咨询。"),
        ("长期目标", "让 Zhutova 成为巴西卖家进口前进行供应链判断的首选品牌。"),
    ],
    {"A": 24, "B": 110},
)


add_sheet(
    wb,
    "02_内容栏目",
    ["栏目", "类型", "目的", "主要内容", "示例Hook"],
    [
        ("值不值得进口？", "流量型", "建立产品判断心智", "拆中国采购价、巴西售价、运费、税费、平台费、MOQ 和风险，最后判断是否值得进口。", "这个产品看起来有利润，但可能不值得进口。"),
        ("本周踩坑", "流量型", "制造风险代入和传播", "供应商风险、物流风险、MOQ风险、税费误判、爆品误判。", "便宜供应商，可能是最贵的错误。"),
        ("中国 vs 巴西价格差", "流量型", "用价格差吸引注意，再教育真实成本", "中国采购价 vs 巴西销售价，再加入真实成本和利润判断。", "中国便宜、巴西卖贵，但利润真的还在吗？"),
        ("卖家常见错误", "流量型", "触发卖家自我代入", "只看采购价、忽略 MOQ、忘记平台费、不看样品、不算退货。", "很多卖家亏钱，是因为忘了这个成本。"),
        ("Zhutova 服务", "品牌/商务", "让用户理解 Zhutova 能具体帮什么", "产品分析、供应商判断、MOQ建议、到岸成本估算、采购和履约支持。", "Zhutova 不只是帮你找供应商。"),
        ("供应链", "品牌/商务", "证明平台能力和执行能力", "工厂直供、全球贸易、本地仓储、智能履约、QC、订单和物流管理。", "找到工厂，不代表采购结束。"),
    ],
    {"A": 24, "B": 16, "C": 32, "D": 68, "E": 46},
)


add_sheet(
    wb,
    "03_每周发布节奏",
    ["星期", "视频风格", "主题方向", "当天目标", "视频Hook"],
    [
        ("周一", "流量型", "低价不等于利润", "开周抢注意力，强化利润判断", "很多卖家进口亏钱，不是因为产品不好。"),
        ("周二", "品牌/商务", "Zhutova 是什么", "解释平台身份，建立基础信任", "Zhutova 不只是 sourcing。"),
        ("周三", "流量型", "供应商风险", "用踩坑内容引发代入和评论", "便宜供应商，可能是最贵的错误。"),
        ("周四", "品牌/商务", "平台能力/供应链幕后", "展示工厂直供、本地仓储、智能履约", "找到工厂，不代表采购结束。"),
        ("周五", "流量型", "MOQ 与现金流", "打中卖家现金流痛点，提高收藏", "第一单买太多，可能直接压死现金流。"),
        ("周六", "流量+信任", "客户案例/模拟案例", "用案例展示 Zhutova 的判断过程", "这个产品不是不能做，是不能这样做。"),
        ("周日", "品牌/商务", "生态与合作", "承接商务合作和平台背书", "Zhutova 连接的不只是产品。"),
    ],
    {"A": 10, "B": 18, "C": 28, "D": 44, "E": 48},
)


add_sheet(
    wb,
    "04_口播文案",
    ["主题", "风格", "20秒口播文案", "CTA"],
    [
        ("低价不等于利润", "流量型", "很多卖家进口亏钱，不是因为产品不好。\n\n而是因为他们只看到了采购价。\n\n真正决定利润的，是供应商、MOQ、运费、税费、仓储和履约。\n\nZhutova 做的不是简单找货，而是帮卖家在进口前判断：这个产品到底值不值得做。", "想让 Zhutova 分析你的产品？发给我们。"),
        ("Zhutova 是什么", "品牌/商务", "Zhutova 是一站式跨境 B2B 平台。\n\n它连接中国制造、全球贸易、本地仓储和智能履约，帮助卖家从产品机会走向更稳定的采购和交付。\n\n如果你需要的不只是供应商名单，而是完整的跨境供应链支持，Zhutova 就是为此而来。", "关注 Zhutova，了解更聪明的跨境采购方式。"),
        ("供应商风险", "流量型", "便宜供应商，可能是最贵的错误。\n\n很多卖家只比较单价，却没有看真实图片、样品、MOQ、出口经验和沟通稳定性。\n\n等到货不对版、交期延误、售后没人回，省下的钱早就不够赔。", "想让 Zhutova 帮你看供应商？发给我们。"),
        ("平台能力链路", "品牌/商务", "找到工厂，不代表采购结束。\n\n你还要处理交易、仓储、物流、履约和服务。\n\nZhutova 把工厂直供、全球贸易、本地仓储和智能履约连接起来，让采购从单点找货变成全链路判断。", "如果你需要更稳定的跨境采购流程，可以联系 Zhutova。"),
        ("MOQ 与现金流", "流量型", "第一单买太多，可能直接压死现金流。\n\n很多供应商会用更低单价吸引你加大 MOQ。\n\n但如果产品还没验证，库存卖不动，现金就被锁住了。\n\n正确的第一单，是你能安全测试需求的数量。", "想知道你的 MOQ 合不合理？把产品发给 Zhutova。"),
        ("客户案例", "流量+信任", "这个产品不是不能做，是不能这样做。\n\n一个卖家想进口一个看起来很有利润的产品。\n\n我们没有先看供应商报价，而是先看需求、MOQ、运费、税费、平台费和履约风险。\n\n最后结论是：小批量测试可以，大单直接上风险太高。", "想让 Zhutova 分析你的产品？发给我们。"),
        ("生态与合作", "品牌/商务", "Zhutova 连接的不只是产品。\n\n它连接的是选品采购、销售渠道、供应链服务、仓储、支付和履约支持。\n\n对卖家来说，这意味着从产品判断到业务执行，都有更清晰的路径。", "商务合作请联系 Zhutova。"),
    ],
    {"A": 24, "B": 16, "C": 88, "D": 38},
)


add_sheet(
    wb,
    "05_复盘表",
    ["日期", "平台", "主题", "风格", "播放量", "完播率", "收藏/分享", "评论/私信", "WhatsApp咨询", "复盘结论"],
    [("", "", "", "", "", "", "", "", "", "") for _ in range(14)],
    {"A": 14, "B": 16, "C": 24, "D": 16, "E": 14, "F": 14, "G": 18, "H": 18, "I": 18, "J": 54},
)


for ws in wb.worksheets:
    if ws.title == "04_口播文案":
        for r in range(2, ws.max_row + 1):
            ws.row_dimensions[r].height = 145

wb.save(OUT)
print(OUT)
