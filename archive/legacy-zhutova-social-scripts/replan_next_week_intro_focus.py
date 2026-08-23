from pathlib import Path

from openpyxl import load_workbook


WORKBOOK = Path(r"D:\zhutova 2清洗 正确\社媒\Zhutova_社媒运营计划_简洁版.xlsx")


WEEK4_TOPICS = [
    "基础介绍：Zhutova 到底是什么？",
    "基础介绍：Zhutova 平台怎么运作？",
    "基础介绍：从工厂直供到智能履约是什么链路？",
    "基础介绍：为什么 Zhutova 不是普通采购服务？",
    "基础介绍：Zhutova 可以提供哪些合作支持？",
]

WEEK4_VOICEOVERS = [
    (
        "第4周",
        "基础介绍：Zhutova 到底是什么？",
        "Zhutova 是什么",
        "很多人第一次听到 Zhutova，\n会觉得，\n这是不是就是一个找货的平台。\n\n其实不只是这样。\n\nZhutova 更准确来说，\n是一个连接中国制造和全球市场的\n一站式跨境 B2B 平台。\n\n它做的，\n不是只帮你找一个产品，\n而是把后面的工厂对接、跨境交易、仓储、物流、履约，\n这些环节一起串起来。\n\n所以对海外卖家、品牌方和合作伙伴来说，\nZhutova 的价值，\n不是只让你看到中国供应链便宜，\n而是帮你把从选品到交付这整件事，\n做得更完整，也更顺。",
        "基本资料 / 品牌认知 / 40-60秒介绍",
    ),
    (
        "第4周",
        "基础介绍：Zhutova 平台怎么运作？",
        "平台如何运作",
        "很多人会以为，\n跨境采购最重要的，\n就是先找到工厂。\n\n但实际上，\n找到工厂只是第一步。\n\n后面还有产品确认、交易流程、订单处理、仓储安排、物流管理、履约协同，\n甚至售后服务。\n\nZhutova 平台做的，\n就是把这些原本很散的环节接起来。\n\n让客户不是只拿到一个报价，\n而是进入一套更完整、更清楚的采购和交付流程。\n\n所以它不是单点找货，\n而是把整条链路跑顺。",
        "基本资料 / 平台流程 / 40-60秒介绍",
    ),
    (
        "第4周",
        "基础介绍：从工厂直供到智能履约是什么链路？",
        "供应链与履约",
        "从工厂直供到智能履约，\n中间其实有很长一条链路。\n\n不是简单发个货，\n就结束了。\n\n这里面会经过供应确认、样品判断、订单处理、仓储安排、物流节点管理，\n最后才到真正的交付协同。\n\n只要有一个环节不稳，\n影响的就不只是时效，\n最后还会影响客户体验，\n甚至影响整盘生意。\n\n所以 Zhutova 一直强调工厂直供、本地仓储和智能履约，\n不是为了把话说大，\n而是希望客户拿到的，\n不只是产品，\n而是一套更稳、更可控的交付能力。",
        "基本资料 / 供应链能力 / 40-60秒介绍",
    ),
    (
        "第4周",
        "基础介绍：为什么 Zhutova 不是普通采购服务？",
        "Zhutova 是什么",
        "普通采购服务，\n一般更关注前面的动作。\n\n比如帮你找供应商、问价格、拿样品、确认订单。\n\n这些当然重要，\n但对跨境 B2B 来说，\n真正容易出问题的，\n往往是在后面。\n\n比如供应能不能持续，\n订单流程清不清楚，\n仓储和物流能不能配合，\n履约和售后能不能跟上。\n\nZhutova 和普通采购服务的区别就在这里。\n\n它不是只帮你完成采购动作，\n而是希望把供应、交易、仓储、物流和履约放在一起考虑。\n\n让客户不是只完成一单采购，\n而是更有机会把跨境合作长期跑稳。",
        "基本资料 / 品牌认知 / 40-60秒介绍",
    ),
    (
        "第4周",
        "基础介绍：Zhutova 可以提供哪些合作支持？",
        "案例与商务合作",
        "Zhutova 能提供的合作支持，\n不只是产品信息，\n也不只是一份供应商名单。\n\n如果你是卖家，\n它可以帮助你看产品机会、供应稳定性和履约成本。\n\n如果你是品牌方，\n它可以帮你对接中国制造资源，\n完善从供应到交付的链路。\n\n如果你是渠道伙伴，\n它可以一起探索产品、仓储、物流和本地市场合作。\n\n如果你是供应方，\n也可以通过 Zhutova，\n连接更多海外市场机会。\n\n所以 Zhutova 提供的，\n不是某一个单点服务，\n而是围绕跨境 B2B 合作，\n把资源、供应、履约和商务机会连接起来。",
        "基本资料 / 商务合作 / 40-60秒介绍",
    ),
]


def main():
    wb = load_workbook(WORKBOOK)

    ws04 = wb["04_未来4周主题安排"]
    for col_idx, value in enumerate(["第4周", *WEEK4_TOPICS], start=1):
        ws04.cell(5, col_idx).value = value

    ws05 = wb["05_口播文案库"]
    week4_rows = []
    for row_idx in range(2, ws05.max_row + 1):
        if ws05.cell(row_idx, 1).value == "第4周":
            week4_rows.append(row_idx)

    if week4_rows:
        start_row = week4_rows[0]
        for i, row in enumerate(WEEK4_VOICEOVERS):
            target_row = start_row + i
            for col_idx, value in enumerate(row, start=1):
                ws05.cell(target_row, col_idx).value = value
            ws05.row_dimensions[target_row].height = 142
    else:
        for row in WEEK4_VOICEOVERS:
            ws05.append(row)
            ws05.row_dimensions[ws05.max_row].height = 142

    wb.save(WORKBOOK)
    print(WORKBOOK)


if __name__ == "__main__":
    main()
