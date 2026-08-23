from pathlib import Path
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side


ROOT = Path(r"D:\zhutova 2清洗 正确")
PLAN = ROOT / "社媒" / "Zhutova_社媒内容汇总_2026_06" / "00_社媒规划" / "zhutova社媒规划_中文版本.xlsx"
OUTER = ROOT / "社媒" / "zhutova社媒规划_新版判断品牌计划_中文版本_H5定位校准版.xlsx"

scripts = {
    "Day 1": (
        "Zhutova is not just a sourcing service.\n\n"
        "It is a one-stop cross-border B2B platform connecting factory direct supply, global trading, local warehousing and smart fulfillment.\n\n"
        "For online sellers, this means better decisions before importing: product, supplier, MOQ, landed cost and risk.\n\n"
        "Stop importing in the dark. Source smarter with Zhutova."
    ),
    "Day 2": (
        "A low China price does not mean real profit.\n\n"
        "Before importing, sellers need to calculate the full cost: product price, freight, import taxes, marketplace fees, packaging, returns and cash-flow risk.\n\n"
        "The cheapest supplier is not always the safest decision.\n\n"
        "Want Zhutova to analyze your product? Send it to us."
    ),
    "Day 3": (
        "Sourcing is not finished when you find a factory.\n\n"
        "You still need trading support, warehousing, logistics and fulfillment.\n\n"
        "Zhutova connects factory direct supply, global trading, local warehousing and smart fulfillment in one cross-border B2B platform.\n\n"
        "Because smart sourcing is a full-chain decision."
    ),
    "Day 4": (
        "Do not choose a supplier by price alone.\n\n"
        "Before payment, check real product photos, sample availability, MOQ, export experience and communication quality.\n\n"
        "A cheap supplier can become an expensive mistake.\n\n"
        "Zhutova helps sellers check supplier risk before buying."
    ),
    "Day 5": (
        "Good sourcing is not luck. It is a process.\n\n"
        "Before ordering, sellers should check product specs, supplier communication, samples and real landed cost.\n\n"
        "Then comes order placement, logistics tracking, warehousing and service support.\n\n"
        "Zhutova turns sourcing into a structured process."
    ),
    "Day 6": (
        "A seller asked if this product was worth importing.\n\n"
        "We did not start with the supplier price.\n\n"
        "We checked demand, MOQ, freight, taxes, marketplace fees, supplier reliability and fulfillment risk.\n\n"
        "The answer is not always yes or no. It depends on margin, volume and risk."
    ),
    "Day 7": (
        "Zhutova connects more than products.\n\n"
        "It connects sourcing, sales channels, supply-chain services, warehousing, payments and fulfillment support.\n\n"
        "For sellers, that means more clarity from product decision to business execution.\n\n"
        "For business cooperation, contact Zhutova."
    ),
}


def update(path: Path):
    wb = openpyxl.load_workbook(path)
    ws = wb["一周视频规划_每日1条"]
    headers = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
    if "20秒英文口播文案" in headers:
        col = headers.index("20秒英文口播文案") + 1
    else:
        col = ws.max_column + 1
        ws.cell(1, col, "20秒英文口播文案")

    for row in range(2, ws.max_row + 1):
        day = ws.cell(row, 1).value
        if day in scripts:
            ws.cell(row, col, scripts[day])

    thin = Side(style="thin", color="D9E2EC")
    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.border = Border(bottom=thin)
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="0A1628")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 68
    for r in range(2, ws.max_row + 1):
        ws.row_dimensions[r].height = 162
    wb.save(path)


update(PLAN)
if OUTER.exists():
    update(OUTER)

print(PLAN)
