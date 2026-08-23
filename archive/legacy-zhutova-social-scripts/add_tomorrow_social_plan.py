from pathlib import Path
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side


ROOT = Path(r"D:\zhutova 2清洗 正确")
XLSX = ROOT / "zhutova社媒规划_新版判断品牌计划.xlsx"

wb = openpyxl.load_workbook(XLSX)
sheet_name = "明日计划_5月28"
if sheet_name in wb.sheetnames:
    del wb[sheet_name]

ws = wb.create_sheet(sheet_name, 0)
ws.sheet_view.showGridLines = False
ws.freeze_panes = "A4"

navy = "0A1628"
wine = "5A0F1B"
blue = "1C97FF"
light_blue = "E8F4FF"
cream = "FFF9F2"
line = "D9E2EC"

ws["A1"] = "Zhutova 5月28日社媒执行计划"
ws["A2"] = "周主题：Pare de importar no escuro / 不要再盲目进口"
ws["A3"] = "当天主题：Preço baixo na China não significa lucro real / 中国低价不等于真实利润"
for cell in ["A1", "A2", "A3"]:
    ws[cell].font = Font(bold=True, color="FFFFFF" if cell == "A1" else navy, size=16 if cell == "A1" else 12)
    ws[cell].fill = PatternFill("solid", fgColor=wine if cell == "A1" else light_blue)
    ws[cell].alignment = Alignment(vertical="center", wrap_text=True)
ws.merge_cells("A1:I1")
ws.merge_cells("A2:I2")
ws.merge_cells("A3:I3")

headers = [
    "时间",
    "平台",
    "内容形式",
    "发布目的",
    "葡语标题 / Hook",
    "内容要点",
    "Caption / CTA",
    "素材文件",
    "检查项",
]

rows = [
    [
        "09:30",
        "Instagram Feed",
        "单图 Banner",
        "第一天预热；建立“Zhutova懂利润和风险”的认知",
        "Preço baixo na China não é lucro real.",
        "产品采购价只是第一步；真实利润要算 frete、impostos、taxas do marketplace、embalagem、devoluções、risco de caixa。",
        "Preço baixo na China não significa lucro real.\n\nAntes de fechar um pedido, o vendedor precisa calcular o custo completo: preço do produto, frete internacional, impostos, taxas do marketplace, embalagem, devoluções e risco de caixa.\n\nO fornecedor mais barato nem sempre é a decisão mais segura.\n\nUma forma mais inteligente de importar vem aí.\n\nLançamento em 1 de junho.",
        r"D:\zhutova 2清洗 正确\社媒\instagram_warmup_2026_05_28_31_PTBR\2026-05-28_aquecimento_preco-baixo.png",
        "图片无错字；日期为28 MAI；caption 用葡语；hashtags 放最后；Bio/WhatsApp链接可用。",
    ],
    [
        "12:30",
        "Instagram Story",
        "投票 + 转发早上Banner",
        "引导互动；测试受众是否理解“真实成本”痛点",
        "Você calcula o custo real antes de importar?",
        "Story 1：转发 banner。\nStory 2：投票 Sim / Ainda não。\nStory 3：提示 frete + imposto + taxa podem mudar a margem。",
        "CTA: Responda nos stories ou envie seu produto para análise.",
        "使用同一张 5/28 葡语 banner，Story 可加投票贴纸。",
        "投票按钮简洁；不要写太多字；保留 Zhutova 标识。",
    ],
    [
        "16:30",
        "TikTok / Reels",
        "15-25秒短视频",
        "情绪型流量；用“打脸低价”吸引卖家停留",
        "Importar barato não significa ganhar dinheiro.",
        "脚本：\n1. Esse produto custa barato na China.\n2. Mas você ainda precisa pagar frete, imposto, taxa do marketplace e devolução.\n3. Se não calcular antes, o lucro some.\n4. Pare de importar no escuro.",
        "Legenda curta: Preço baixo não paga prejuízo. A Zhutova lança em 1 de junho para ajudar vendedores a analisar produto, custo e risco antes de importar.",
        "可用产品图 + 成本表录屏 + 口播；也可用早上 banner 做动态缩放。",
        "前3秒必须有Hook；字幕大；结尾出现“Lançamento em 1 de junho”。",
    ],
    [
        "19:30",
        "Facebook",
        "讨论帖",
        "引评论；不要硬广，让卖家说出自己的成本痛点",
        "O que mais mata a margem na importação: frete, imposto ou taxa do marketplace?",
        "发起讨论：很多卖家只看中国采购价，但真正决定利润的是总成本。请大家评论最容易被低估的成本。",
        "Post: Preço baixo na China não significa lucro real.\n\nNa prática, o que mais mata a margem quando você importa para vender no Brasil?\n\n1. Frete internacional\n2. Impostos\n3. Taxas do marketplace\n4. Devoluções\n5. Estoque parado\n\nComenta aqui qual desses custos mais pesa na sua operação.",
        "可配 5/28 banner 或纯文字帖。",
        "问题要放第一行；不要一开始介绍平台；评论后可私信引导 WhatsApp。",
    ],
    [
        "21:00",
        "内部复盘",
        "数据记录",
        "判断首日痛点是否有效",
        "记录互动和私信",
        "记录 IG 收藏/分享/评论、Story 投票、TikTok完播、Facebook评论、WhatsApp咨询。",
        "复盘问题：低价≠利润这个角度是否引发评论？有没有人发产品来问？哪条内容最容易互动？",
        "在表格或后台记录。",
        "保留评论截图；把用户问题整理成 5/29 供应商主题素材。",
    ],
]

start_row = 5
for c, h in enumerate(headers, 1):
    cell = ws.cell(start_row, c, h)
    cell.font = Font(bold=True, color="FFFFFF")
    cell.fill = PatternFill("solid", fgColor=navy)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

for r, row in enumerate(rows, start_row + 1):
    for c, value in enumerate(row, 1):
        cell = ws.cell(r, c, value)
        cell.alignment = Alignment(vertical="top", wrap_text=True)
        cell.fill = PatternFill("solid", fgColor=cream if r % 2 == 0 else "FFFFFF")

thin = Side(style="thin", color=line)
for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=len(headers)):
    for cell in row:
        cell.border = Border(bottom=thin)

widths = [12, 18, 20, 28, 36, 52, 72, 56, 42]
for i, width in enumerate(widths, 1):
    ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = width

for r in range(1, ws.max_row + 1):
    ws.row_dimensions[r].height = 24
for r in range(start_row + 1, ws.max_row + 1):
    ws.row_dimensions[r].height = 126

wb.save(XLSX)
print(XLSX)
