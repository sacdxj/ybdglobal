from pathlib import Path
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side


ROOT = Path(r"D:\zhutova 2清洗 正确")
PLAN = ROOT / "社媒" / "Zhutova_社媒内容汇总_2026_06" / "00_社媒规划" / "zhutova社媒规划_中文版本.xlsx"

wb = openpyxl.load_workbook(PLAN)
sheet_name = "一周视频规划_每日1条"
if sheet_name in wb.sheetnames:
    ws = wb[sheet_name]
    ws.delete_rows(1, ws.max_row)
else:
    ws = wb.create_sheet(sheet_name, 0)

ws.sheet_view.showGridLines = False
ws.freeze_panes = "A2"

headers = [
    "天数",
    "建议风格/目标",
    "内容主题",
    "视频标题/Hook",
    "本条采用文案",
    "20秒爆点流量版口播",
    "20秒品牌/商务背书版口播",
    "核心目的",
    "画面/素材建议",
    "字幕关键词",
    "CTA",
    "适合平台",
]

rows = [
    (
        "Day 1",
        "爆点流量版",
        "低价不等于利润",
        "很多卖家进口亏钱，不是因为产品不好。",
        "爆点流量版",
        "很多卖家进口亏钱，不是因为产品不好。\n\n而是因为他们只看到了采购价。\n\n真正决定利润的，是供应商、MOQ、运费、税费、仓储和履约。\n\nZhutova 做的不是简单找货，而是帮卖家在进口前判断：这个产品到底值不值得做。\n\n不要再盲目进口。",
        "Zhutova 是一站式跨境 B2B 平台，连接工厂直供、全球贸易、本地仓储和智能履约。\n\n它帮助卖家在进口前判断产品、供应商、MOQ、到岸成本和风险。\n\n对卖家来说，Zhutova 不只是采购服务，而是从产品机会到履约执行的跨境供应链支持。",
        "用强痛点建立第一印象：Zhutova 懂利润和风险。",
        "成本表、产品价格对比、红色亏损提示、最后出现 Zhutova logo。",
        "采购价 / 真实利润 / MOQ / 运费税费 / 不要盲目进口",
        "想让 Zhutova 分析你的产品？发给我们。",
        "TikTok / Instagram Reels / YouTube Shorts",
    ),
    (
        "Day 2",
        "品牌/商务背书版",
        "Zhutova 是什么",
        "Zhutova 不只是 sourcing。",
        "品牌/商务背书版",
        "你以为 Zhutova 只是帮你找供应商？\n\n不只是。\n\n真正难的不是找到一个报价，而是判断这个产品能不能赚钱、供应商靠不靠谱、履约能不能稳定。\n\nZhutova 把选品、供应链、仓储和履约连接起来，让卖家少一点盲猜，多一点判断。",
        "Zhutova 是一站式跨境 B2B 平台。\n\n它连接中国制造、全球贸易、本地仓储和智能履约，帮助卖家从产品机会走向更稳定的采购和交付。\n\n如果你需要的不只是一个供应商名单，而是更完整的跨境供应链支持，Zhutova 就是为此而来。",
        "补充平台身份，适合做置顶/品牌背书。",
        "H5首页、世界地图、工厂直供/全球贸易/本地仓储/智能履约四步图。",
        "一站式跨境 B2B 平台 / 中国制造 / 全球贸易 / 智能履约",
        "关注 Zhutova，了解更聪明的跨境采购方式。",
        "Instagram Reels / LinkedIn / Facebook",
    ),
    (
        "Day 3",
        "爆点流量版",
        "供应商风险",
        "便宜供应商，可能是最贵的错误。",
        "爆点流量版",
        "便宜供应商，可能是最贵的错误。\n\n很多卖家只比较单价，却没有看真实图片、样品、MOQ、出口经验和沟通稳定性。\n\n等到货不对版、交期延误、售后没人回，省下的钱早就不够赔。\n\n下单前，先判断供应商风险。",
        "Zhutova 帮助卖家在采购前评估供应商风险。\n\n我们关注的不只是报价，还包括产品信息、样品可能性、MOQ、出口经验、沟通质量和后续履约稳定性。\n\n更好的供应商判断，应该发生在付款之前。",
        "用供应商踩坑制造代入和评论。",
        "供应商聊天界面、红旗/绿旗 checklist、价格低但风险高的对比图。",
        "供应商风险 / 样品 / MOQ / 出口经验 / 付款前判断",
        "想让 Zhutova 帮你看供应商？通过 WhatsApp 发给我们。",
        "TikTok / Instagram Reels / Facebook",
    ),
    (
        "Day 4",
        "品牌/商务背书版",
        "平台能力链路",
        "找到工厂，不代表采购结束。",
        "品牌/商务背书版",
        "找到工厂，不代表采购结束。\n\n你还要处理交易、仓储、物流、履约和服务。\n\n如果只看一个报价，后面的风险很容易被忽略。\n\nZhutova 把工厂直供、全球贸易、本地仓储和智能履约连接起来，让采购从单点找货变成全链路判断。",
        "Zhutova 的平台能力覆盖从工厂直供到智能履约。\n\nFactory Direct Supply、Global Trading、Local Warehousing、Smart Fulfillment，是卖家从产品到交付之间需要被管理的关键链路。\n\n这就是一站式跨境 B2B 平台的价值。",
        "解释 H5 里的四段平台能力，建立专业度。",
        "四步流程图：Factory Direct Supply -> Global Trading -> Local Warehousing -> Smart Fulfillment。",
        "工厂直供 / 全球贸易 / 本地仓储 / 智能履约 / 全链路",
        "如果你需要更稳定的跨境采购流程，可以联系 Zhutova。",
        "LinkedIn / Instagram Reels / Facebook",
    ),
    (
        "Day 5",
        "爆点流量版",
        "MOQ 与现金流",
        "第一单买太多，可能直接压死现金流。",
        "爆点流量版",
        "第一单买太多，可能直接压死现金流。\n\n很多供应商会用更低单价吸引你加大 MOQ。\n\n但如果产品还没验证，库存卖不动，现金就被锁住了。\n\n正确的第一单，不是单价最低，而是你能安全测试需求的数量。",
        "Zhutova 帮助卖家在下单前评估 MOQ 是否合理。\n\n我们会结合产品需求、销售渠道、库存风险、运费和平台费用，判断第一单应该测试还是放大。\n\nMOQ 不是越大越好，适合你的阶段才重要。",
        "用现金流风险做爆点，适合高播放。",
        "库存堆积、现金流锁住、MOQ数字对比、测试订单 vs 大订单。",
        "MOQ / 现金流 / 库存积压 / 测试需求 / 第一单",
        "想知道你的 MOQ 合不合理？把产品发给 Zhutova。",
        "TikTok / Instagram Reels / YouTube Shorts",
    ),
    (
        "Day 6",
        "爆点流量版",
        "客户案例/模拟案例",
        "这个产品不是不能做，是不能这样做。",
        "爆点流量版",
        "这个产品不是不能做，是不能这样做。\n\n一个卖家想进口一个看起来很有利润的产品。\n\n我们没有先看供应商报价，而是先看需求、MOQ、运费、税费、平台费和履约风险。\n\n最后结论是：小批量测试可以，大单直接上风险太高。",
        "一个卖家想判断某个产品是否值得进口。\n\nZhutova 会先看需求、供应商可靠性、MOQ、真实到岸成本、销售渠道费用和履约风险。\n\n我们的目标不是简单说可以或不可以，而是帮助卖家找到更稳的执行方式。",
        "用案例证明 Zhutova 的判断力，兼具流量和信任。",
        "模拟产品卡片、判断表、Verdict：Test first / Not at scale yet。",
        "案例 / 产品分析 / 小批量测试 / 大单风险 / 判断力",
        "想让 Zhutova 分析你的产品？发给我们。",
        "TikTok / Instagram Reels / Facebook",
    ),
    (
        "Day 7",
        "品牌/商务背书版",
        "生态与商务合作",
        "Zhutova 连接的不只是产品。",
        "品牌/商务背书版",
        "Zhutova 连接的不只是产品。\n\n它连接的是选品采购、销售渠道、供应链服务、仓储、支付和履约支持。\n\n对卖家来说，这意味着从产品判断到业务执行，都有更清晰的路径。\n\n如果你要谈供应链合作，可以联系 Zhutova。",
        "Zhutova 是连接中国制造与全球市场的一站式跨境 B2B 平台。\n\n平台能力覆盖全品类供应、全球销售网络、可靠供应链、交易体验和供应链服务。\n\n我们欢迎卖家、品牌方、供应商和渠道伙伴进行商务合作。",
        "周末做品牌背书/商务合作，承接潜在B端咨询。",
        "H5合作伙伴页、地图、平台生态图、商务合作关键词。",
        "生态 / 合作伙伴 / 全球市场 / 商务合作 / B2B平台",
        "商务合作请联系 Zhutova。",
        "LinkedIn / Facebook / Instagram Reels",
    ),
]

for c, h in enumerate(headers, 1):
    ws.cell(1, c, h)
for r, row in enumerate(rows, 2):
    for c, value in enumerate(row, 1):
        ws.cell(r, c, value)

thin = Side(style="thin", color="D9E2EC")
for row in ws.iter_rows():
    for cell in row:
        cell.alignment = Alignment(vertical="top", wrap_text=True)
        cell.border = Border(bottom=thin)

for cell in ws[1]:
    cell.font = Font(bold=True, color="FFFFFF")
    cell.fill = PatternFill("solid", fgColor="0A1628")
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

widths = {
    "A": 9,
    "B": 18,
    "C": 24,
    "D": 38,
    "E": 16,
    "F": 62,
    "G": 62,
    "H": 34,
    "I": 46,
    "J": 34,
    "K": 36,
    "L": 28,
}
for col, width in widths.items():
    ws.column_dimensions[col].width = width
for r in range(2, ws.max_row + 1):
    ws.row_dimensions[r].height = 180

# Color-code the two styles.
for r in range(2, ws.max_row + 1):
    style_value = ws.cell(r, 2).value or ""
    color = "E8F4FF" if "爆点" in style_value else "FFF4D6"
    for c in range(1, ws.max_column + 1):
        ws.cell(r, c).fill = PatternFill("solid", fgColor=color)

wb.save(PLAN)
print(PLAN)
