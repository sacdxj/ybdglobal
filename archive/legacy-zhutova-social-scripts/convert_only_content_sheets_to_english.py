from pathlib import Path
from datetime import date, timedelta
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side


ROOT = Path(r"D:\zhutova 2清洗 正确")
XLSX = ROOT / "zhutova社媒规划_新版判断品牌计划.xlsx"

wb = openpyxl.load_workbook(XLSX)


def clear_sheet(ws):
    ws.delete_rows(1, ws.max_row)


def style_sheet(sheet, widths=None, freeze="A2"):
    sheet.freeze_panes = freeze
    sheet.sheet_view.showGridLines = False
    widths = widths or {}
    for col, width in widths.items():
        sheet.column_dimensions[col].width = width
    thin = Side(style="thin", color="D9E2EC")
    for row in sheet.iter_rows():
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.border = Border(bottom=thin)
    for cell in sheet[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="0A1628")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    sheet.row_dimensions[1].height = 30


def write_table(sheet, headers, rows):
    clear_sheet(sheet)
    for c, h in enumerate(headers, 1):
        sheet.cell(1, c, h)
    for r, row in enumerate(rows, 2):
        for c, value in enumerate(row, 1):
            sheet.cell(r, c, value)


# Sheet 1: 6月每日排期 -> content-production-facing fields in English.
if "6月每日排期" not in wb.sheetnames:
    calendar = wb.create_sheet("6月每日排期")
else:
    calendar = wb["6月每日排期"]

calendar_headers = [
    "Date",
    "Day",
    "Stage",
    "Content Pillar",
    "Reel/TikTok 1 (Traffic Post)",
    "Reel/Carousel 2 (Trust Post)",
    "English Hook",
    "Verdict Angle",
    "CTA",
    "Notes / Creative Direction",
]
weekdays = ["Mon", "Tue", "Wed", "Thu", "Fri", "Sat", "Sun"]
themes = [
    ("Is It Worth Importing?", "Product looks profitable, but maybe it is not", "China vs Brazil + real landed cost", "This product looks profitable... but maybe it is not.", "Worth it / Not worth it / Only at scale"),
    ("Trap of the Week", "Cheap supplier can become expensive", "Reliable supplier checklist", "A cheap supplier can become very expensive.", "Risk can be bigger than savings"),
    ("China vs Brazil", "China price vs Brazil selling price", "Margin after freight and taxes", "It costs little in China and sells higher in Brazil. But what about margin?", "Price gap does not equal profit"),
    ("Seller Mistake", "Forgetting marketplace fees", "Real cost table", "Many sellers forget this cost.", "Margin disappears in the details"),
    ("Is It Worth Importing?", "Viral product does not mean profit", "MOQ + dead stock", "A viral product does not pay the bills if margin is weak.", "Only worth it with validation"),
    ("Trap of the Week", "Cheap freight with hidden risk", "When freight kills margin", "Cheap freight can become expensive.", "Logistics is a profit factor"),
    ("Seller Mistake", "Buying too much in the first order", "How to test with low MOQ", "The right first order is not always the cheapest one.", "Cash flow first"),
]

calendar_rows = []
start = date(2026, 6, 1)
for i in range(30):
    d = start + timedelta(days=i)
    pillar, traffic, trust, hook, verdict = themes[i % len(themes)]
    stage = "Launch Week" if i < 7 else ("Proof Building" if i < 14 else ("Trust + Repetition" if i < 23 else "Soft Conversion Push"))
    if i == 0:
        pillar = "Decision Brand Launch"
        traffic = "Stop importing in the dark (official launch)"
        trust = "Zhutova starts analyzing product, supplier, MOQ, real cost, and risk"
        hook = "Starting today, stop importing in the dark."
        verdict = "Zhutova officially starts helping Brazilian sellers make better sourcing decisions"
    calendar_rows.append((
        d.strftime("%Y-%m-%d"),
        weekdays[d.weekday()],
        stage,
        pillar,
        traffic,
        trust,
        hook,
        verdict,
        "Want Zhutova to analyze your product? Send it to us on WhatsApp.",
        "Short video: voiceover + product image + cost table. Carousel: 6-7 pages: Hook / China price / Brazil price / real cost / risk / verdict / CTA.",
    ))

write_table(calendar, calendar_headers, calendar_rows)
style_sheet(calendar, {"A": 13, "B": 10, "C": 18, "D": 24, "E": 40, "F": 46, "G": 46, "H": 36, "I": 48, "J": 62})
for row in range(2, calendar.max_row + 1):
    if calendar.cell(row, 3).value == "Launch Week":
        for col in range(1, calendar.max_column + 1):
            calendar.cell(row, col).fill = PatternFill("solid", fgColor="E8F4FF")


# Sheet 2: 预热与上线文案 -> ready-to-post English captions.
if "预热与上线文案" not in wb.sheetnames:
    copy = wb.create_sheet("预热与上线文案")
else:
    copy = wb["预热与上线文案"]

copy_headers = ["Date / Scenario", "Creative Theme", "English Caption (Ready to Post)", "Hashtags", "Asset Folder"]
copy_rows = [
    (
        "2026-05-28",
        "A low China price is not real profit",
        "A low China price does not mean real profit.\n\nBefore placing an order, online sellers need to calculate the full cost: product price, international freight, import taxes, marketplace fees, packaging, returns and cash-flow risk.\n\nThe cheapest supplier is not always the safest decision.\n\nA smarter way to source is coming.\n\nLaunching June 1.",
        "#Zhutova #ChinaSourcing #ImportFromChina #ProductSourcing #EcommerceSellers #OnlineSellers #SupplierSourcing #MarketplaceSeller #CrossBorderCommerce #SmallBusiness",
        r"D:\zhutova 2清洗 正确\社媒\instagram_warmup_2026_05_28_31_EN",
    ),
    (
        "2026-05-29",
        "Do not choose a supplier by price alone",
        "Do not choose a supplier by price alone.\n\nA reliable supplier should give clear answers before you send money or commit to a large order.\n\nBefore placing your first order, check:\n- Real product photos\n- Sample availability\n- Clear MOQ\n- Export experience\n- Consistent communication\n\nBetter supplier decisions start before the payment.\n\nZhutova launches June 1.",
        "#Zhutova #ChinaSourcing #SupplierSourcing #ImportFromChina #ProductSourcing #EcommerceSellers #OnlineSellers #MarketplaceSeller #CrossBorderCommerce #SmallBusiness",
        r"D:\zhutova 2清洗 正确\社媒\instagram_warmup_2026_05_28_31_EN",
    ),
    (
        "2026-05-30",
        "MOQ can protect or hurt your cash flow",
        "MOQ can protect or hurt your cash flow.\n\nA lower unit price can look attractive, but a large minimum order can lock too much money into untested inventory.\n\nBefore scaling, sellers should ask:\n- Can I test demand with a smaller quantity?\n- Can I sell this stock fast enough?\n- Will freight and fees still make sense?\n- What happens if the product does not move?\n\nThe right first order is not always the cheapest one. It is the one you can validate safely.\n\nZhutova launches June 1.",
        "#Zhutova #ChinaSourcing #MOQ #ImportFromChina #ProductSourcing #EcommerceSellers #OnlineSellers #MarketplaceSeller #InventoryManagement #CrossBorderCommerce",
        r"D:\zhutova 2清洗 正确\社媒\instagram_warmup_2026_05_28_31_EN",
    ),
    (
        "2026-05-31",
        "Zhutova launches tomorrow",
        "Zhutova launches tomorrow.\n\nA sourcing platform for online sellers who want to import from China with more clarity, less guesswork, and better decisions before buying.\n\nWith Zhutova, sellers can:\n- Find product opportunities\n- Check supplier reliability\n- Understand MOQ before ordering\n- Estimate real landed cost\n- Source with less risk\n\nTomorrow, a smarter way to source begins.\n\nLaunching June 1.",
        "#Zhutova #ChinaSourcing #ImportFromChina #ProductSourcing #SupplierSourcing #EcommerceSellers #OnlineSellers #MarketplaceSeller #CrossBorderCommerce #SmallBusiness",
        r"D:\zhutova 2清洗 正确\社媒\instagram_warmup_2026_05_28_31_EN",
    ),
    (
        "2026-06-01",
        "Stop importing in the dark",
        "Stop importing in the dark.\n\nStarting today, Zhutova helps Brazilian sellers make better decisions before importing: product, supplier, MOQ, freight, taxes, margin, and risk.\n\nBuying cheap is not enough. What matters is whether profit still exists after the full cost calculation.\n\nWant Zhutova to analyze your product? Send it to us on WhatsApp.",
        "#Zhutova #ChinaSourcing #ImportFromChina #ProductSourcing #SupplierSourcing #EcommerceSellers #OnlineSellers #MarketplaceSeller #CrossBorderCommerce #ChinaBrazil",
        r"D:\zhutova 2清洗 正确\社媒",
    ),
]

write_table(copy, copy_headers, copy_rows)
style_sheet(copy, {"A": 18, "B": 36, "C": 86, "D": 54, "E": 60})

try:
    wb.save(XLSX)
    print(XLSX)
except PermissionError:
    fallback = ROOT / "zhutova社媒规划_新版判断品牌计划_仅内容英文版.xlsx"
    wb.save(fallback)
    print(fallback)
