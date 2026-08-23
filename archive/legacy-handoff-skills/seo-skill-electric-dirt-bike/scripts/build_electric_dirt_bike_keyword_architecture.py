from __future__ import annotations

import argparse
import re
import unicodedata
from collections import Counter
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


DEFAULT_INPUT = "electric_dirt_bike.xlsx"
DEFAULT_AGENT = "agent.md"
DEFAULT_OUTPUT = Path("outputs") / "electric_dirt_bike_keyword_architecture.xlsx"


def strip_accents(value: str) -> str:
    return "".join(ch for ch in unicodedata.normalize("NFKD", value) if not unicodedata.combining(ch))


def norm(value) -> str:
    text = "" if pd.isna(value) else str(value)
    text = strip_accents(text).lower()
    text = re.sub(r"(?<=\d),(?=\d)", "", text)
    text = re.sub(r"[^a-z0-9\s\-\/\+]", " ", text)
    return re.sub(r"\s+", " ", text).strip()


def has_any(text: str, patterns: list[str]) -> bool:
    return any(re.search(pattern, text) for pattern in patterns)


CORE_PRODUCT = [
    r"\belectric dirt bike(s)?\b",
    r"\belectric dirtbike(s)?\b",
    r"\be dirt bike(s)?\b",
    r"\be dirtbike(s)?\b",
    r"\beletric dirt bike(s)?\b",
    r"\beletric dirtbike(s)?\b",
    r"\bdirt bike(s)?\b.*\belectric\b",
    r"\bdirt bike(s)? that are electric\b",
    r"\bdirt rocket\b",
    r"\belectric motocross\b",
    r"\be[- ]?moto\b",
    r"\belectric pit bike(s)?\b",
    r"\belectric trail bike(s)?\b",
    r"\belectric mini bike(s)?\b",
]
ADULT = [r"\badult(s)?\b", r"\bfor adults\b", r"\bfull size\b", r"\bhigh performance\b"]
KIDS = [
    r"\bkid(s)?\b", r"\byouth\b", r"\bteen(s)?\b", r"\bchildren\b",
    r"\bfor boys\b", r"\bfor girls\b", r"\bage(s)?\b",
    r"\b(?:8|9|10|11|12|13)\s*(?:year|yr)s?[- ]?old\b",
]
STREET_LEGAL = [r"\bstreet legal\b", r"\broad legal\b", r"\bregister(ed|able)?\b", r"\bregistration\b", r"\blicense\b", r"\bdual sport\b"]
SPEED = [r"\bspeed\b", r"\bfast(est)?\b", r"\bmph\b", r"\b40\+?\s*mph\b", r"\b50\+?\s*mph\b", r"\btop speed\b"]
POWER = [r"\b72v\b", r"\b60v\b", r"\b48v\b", r"\b36v\b", r"\b24v\b", r"\bvolt(age)?\b", r"\bwatt(s)?\b", r"\bkw\b", r"\bmotor\b"]
BATTERY = [r"\bbattery\b", r"\bcharger\b", r"\bcharging\b", r"\brange\b", r"\blithium\b", r"\bruntime\b"]
BUYING = [r"\bbest\b", r"\bfor sale\b", r"\bbuy\b", r"\bprice\b", r"\bcost\b", r"\bcheap\b", r"\bbudget\b", r"\baffordable\b", r"\bfinanc(e|ing)\b", r"\bpay later\b"]
COMPARE = [r"\bvs\b", r"\bversus\b", r"\balternative(s)?\b", r"\bcompare\b", r"\bcomparison\b", r"\bsurron\b", r"\bsur ron\b", r"\brazor\b", r"\btalaria\b", r"\bktm\b", r"\bstark\b"]
PARTS = [r"\bpart(s)?\b", r"\baccessor(y|ies)\b", r"\btire(s)?\b", r"\bbrake(s)?\b", r"\bsuspension\b", r"\bcontroller\b", r"\bchain\b", r"\bsprocket\b", r"\bhelmet\b"]
SUPPORT = [r"\bmaintenance\b", r"\brepair\b", r"\btroubleshoot\b", r"\bwarranty\b", r"\bmanual\b", r"\bservice\b"]
SAFETY = [r"\bsafety\b", r"\bsafety gear\b", r"\bprotective gear\b", r"\blaw(s)?\b", r"\brule(s)?\b", r"\blegal\b", r"\bhelmet\b"]
COMMERCIAL = [r"\bdealer(s)?\b", r"\bdistributor(s)?\b", r"\bwholesale\b", r"\bbulk\b", r"\bquote\b"]
NEWS_RISK = [
    r"\bkilled\b", r"\bdied\b", r"\bdeath\b", r"\bfatal\b", r"\bcrash(ed)?\b",
    r"\baccident\b", r"\binjur(ed|y|ies)\b", r"\bpolice\b",
]
OUT_OF_SCOPE = [
    r"\bgame(s)?\b", r"\btoy(s)? only\b", r"\bcoupon\b", r"\bpromo code\b", r"\blogin\b",
    r"\btracking\b", r"\bapp\b", r"\bmovie\b", r"\bcelebrity\b", r"\bmedical\b",
    r"\bporn\b", r"\badult content\b", r"\bmobility scooter\b", r"\batv\b", r"\bgas dirt bike\b",
]

BRANDS = ["surron", "sur ron", "razor", "talaria", "ktm", "stark", "segway", "cake", "kuberg", "yamaha", "honda"]
DIRT_BIKE_FOCUSED_BRANDS = {"surron", "sur ron", "talaria", "stark", "cake", "kuberg"}
KNOWN_DIRT_BIKE_MODELS = [
    r"\brazor\s+(?:mx125|mx350|mx400|mx500|mx650|sx500)\b",
    r"\bsegway\s+x(?:160|260)\b",
]


def volume_value(row: pd.Series) -> int:
    for col in ["Search Volume", "Volume", "Avg. monthly searches"]:
        if col in row.index:
            try:
                return int(float(row.get(col) or 0))
            except (TypeError, ValueError):
                return 0
    return 0


def classify(row: pd.Series) -> dict:
    keyword = row["Keyword"]
    k = norm(keyword)
    volume = volume_value(row)
    intents = str(row.get("Keyword Intents", "") or "")

    product = has_any(k, CORE_PRODUCT)
    adult = has_any(k, ADULT)
    kids = has_any(k, KIDS)
    street = has_any(k, STREET_LEGAL)
    speed = has_any(k, SPEED)
    power = has_any(k, POWER)
    voltage_72 = bool(re.search(r"\b72\s*(?:v|volt)\b", k))
    voltage_60 = bool(re.search(r"\b60\s*(?:v|volt)\b", k))
    battery = has_any(k, BATTERY)
    buying = has_any(k, BUYING)
    compare = has_any(k, COMPARE)
    parts = has_any(k, PARTS)
    support = has_any(k, SUPPORT)
    safety = has_any(k, SAFETY)
    commercial = has_any(k, COMMERCIAL)
    news_risk = has_any(k, NEWS_RISK)
    out_scope = has_any(k, OUT_OF_SCOPE)
    brand_hits = [b for b in BRANDS if re.search(rf"\b{re.escape(b)}\b", k)]
    dirt_bike_explicit = has_any(k, [r"\bdirt bike(s)?\b", r"\bdirtbike(s)?\b"])
    scooter_only = has_any(k, [r"\be[- ]?scooter(s)?\b", r"\belectric scooter(s)?\b", r"\bscooter(s)?\b"]) and not dirt_bike_explicit
    ebike_only = has_any(k, [r"\be[- ]?bike(s)?\b", r"\belectric bike(s)?\b", r"\belectric bicycle(s)?\b", r"\bbicycle(s)?\b", r"\bcycle(s)?\b"]) and not dirt_bike_explicit
    motorcycle_only = has_any(k, [r"\belectric motorcycle(s)?\b", r"\bmotorcycle(s)?\b"]) and not dirt_bike_explicit
    generic_accessory_only = (battery or parts or support) and not product and not dirt_bike_explicit

    business_signals = sum(bool(x) for x in [product, adult, kids, street, speed, power, battery, buying, compare, parts, support, safety, commercial])
    focused_brand = any(brand in DIRT_BIKE_FOCUSED_BRANDS for brand in brand_hits)
    known_dirt_bike_model = has_any(k, KNOWN_DIRT_BIKE_MODELS)
    has_business_signal = product or focused_brand or known_dirt_bike_model or bool(brand_hits and dirt_bike_explicit)

    risk_flags = []
    if out_scope:
        risk_flags.append("out-of-scope")
    if scooter_only:
        risk_flags.append("scooter-only")
    if ebike_only:
        risk_flags.append("ebike-only")
    if motorcycle_only:
        risk_flags.append("motorcycle-only")
    if generic_accessory_only:
        risk_flags.append("generic-accessory-only")
    if brand_hits and not compare:
        risk_flags.append("competitor/navigation-review")
    if "gas dirt bike" in k and "electric" not in k:
        risk_flags.append("gas-only")
    if len(k.split()) <= 2 and volume >= 10000 and not product:
        risk_flags.append("generic-head-term")
    if street:
        risk_flags.append("legal-compliance-review")
    if news_risk:
        risk_flags.append("accident-news-intent")
    if ("battery" in k or "charger" in k) and any(x in k for x in ["modify", "bypass", "overvolt"]):
        risk_flags.append("unsafe-battery-modification")

    if out_scope or news_risk or scooter_only or ebike_only or motorcycle_only or generic_accessory_only or not has_business_signal:
        business_scope = "Out of Scope"
    elif product or buying or commercial or compare or street:
        business_scope = "Core"
    else:
        business_scope = "Adjacent"

    budget_value = has_any(k, [r"\bcheap\b", r"\bbudget\b", r"\baffordable\b", r"\bunder\s*\$?\d+\b", r"\bfinanc(e|ing)\b", r"\bpay later\b"])

    if compare or brand_hits or budget_value:
        hub, pillar, topic, functional, section, role, cluster = (
            "Compare & Save",
            "Competitor and Budget Comparisons",
            "Brand Alternatives" if brand_hits else "Budget-Friendly Buying Guides",
            "Comparison",
            "Compare",
            "comparison page",
            "Competitor alternatives and budget decisions",
        )
    elif street:
        hub, pillar, topic, functional, section, role, cluster = (
            "Street Legal",
            "Registry and Compliance",
            "Street legal electric dirt bikes",
            "Street Legal",
            "Street Legal",
            "guide page",
            "Street legal, dual-sport, registration and compliance",
        )
    elif kids:
        hub, pillar, topic, functional, section, role, cluster = (
            "Kids & Youth",
            "Kids and Youth Electric Dirt Bikes",
            "Age and safety fit",
            "Kids / Youth",
            "Kids & Youth",
            "collection or guide",
            "Ages, youth sizing, low-voltage safety and parent buying",
        )
    elif adult or speed or (power and (voltage_72 or voltage_60)):
        voltage = "72V Powerhouses" if voltage_72 else "60V Performance" if voltage_60 else "High-Speed Adult Bikes" if speed else "Adult Performance"
        hub, pillar, topic, functional, section, role, cluster = (
            "Adult Dirt Bikes",
            "Adult Performance Electric Dirt Bikes",
            voltage,
            "Adult / Performance",
            "Adult Dirt Bikes",
            "collection or guide",
            "Adult bikes, high speed, voltage, power and terrain fit",
        )
    elif battery or power:
        hub, pillar, topic, functional, section, role, cluster = (
            "Buying Guides",
            "Battery, Range and Power",
            "Battery and voltage education",
            "Battery / Power",
            "Buying Guides",
            "SEO article",
            "Battery, range, charging, motor and voltage decisions",
        )
    elif parts or support:
        hub, pillar, topic, functional, section, role, cluster = (
            "Parts And Support",
            "Parts, Accessories and Maintenance",
            "Parts and after-sales support",
            "Parts / Support",
            "Parts And Support",
            "support page",
            "Parts, accessories, maintenance and troubleshooting",
        )
    elif safety:
        hub, pillar, topic, functional, section, role, cluster = (
            "Safety And Rules",
            "Safety Gear and Riding Rules",
            "Safety and riding rules",
            "Safety / Rules",
            "Safety And Rules",
            "SEO article",
            "Safety gear, youth riding, trail rules and risk boundaries",
        )
    elif buying or product:
        hub, pillar, topic, functional, section, role, cluster = (
            "Shop All",
            "Electric Dirt Bike Catalog",
            "Shop and buying intent",
            "Shop / Buyer",
            "Shop All",
            "collection page",
            "Shop all electric dirt bikes by buyer intent",
        )
    else:
        hub, pillar, topic, functional, section, role, cluster = (
            "Out of Scope / Review",
            "Unqualified Traffic",
            "Review unqualified keyword intent",
            "Review",
            "Review",
            "review",
            "Unqualified or unrelated intent",
        )

    audience = "Parent Buyer" if kids else "Adult Rider" if adult else "Dealer / Wholesale" if commercial else "Comparison Buyer" if compare else "General Buyer"
    geo = "en-US"
    page_candidate = make_page_candidate(role, topic, keyword, brand_hits)

    qa_score = 0
    qa_score += 35 if business_scope == "Core" else 15 if business_scope == "Adjacent" else 0
    qa_score += min(20, business_signals * 4)
    qa_score += 15 if buying or commercial or compare else 12 if product else 0
    qa_score += 10 if volume >= 100 else 5 if volume > 0 else 0
    qa_score -= 35 if out_scope else 0
    qa_score -= 12 if "competitor/navigation-review" in risk_flags and not compare else 0
    qa_score -= 10 if "generic-head-term" in risk_flags else 0
    qa_score = max(0, min(100, qa_score))

    serious_flags = any(flag in risk_flags for flag in [
        "out-of-scope", "gas-only", "unsafe-battery-modification", "accident-news-intent",
        "scooter-only", "ebike-only", "motorcycle-only", "generic-accessory-only",
    ])
    if business_scope == "Core" and qa_score >= 60 and not serious_flags:
        decision = "Build now" if volume >= 100 else "Build later"
    elif business_scope in {"Core", "Adjacent"} and qa_score >= 40:
        decision = "Manual review before build" if risk_flags else "Build later"
    elif business_scope == "Adjacent":
        decision = "Merge into parent/FAQ"
    else:
        decision = "Reject/deprioritize"

    if decision == "Build now" and functional in {"Shop / Buyer", "Adult / Performance", "Kids / Youth", "Street Legal", "Comparison"}:
        batch = "Batch 1 - Commercial/Core"
    elif decision == "Build now":
        batch = "Batch 1 - SEO Articles"
    elif decision == "Build later":
        batch = "Batch 2 - SEO Expansion"
    elif decision == "Merge into parent/FAQ":
        batch = "Parent Support"
    elif decision == "Manual review before build":
        batch = "Review Queue"
    else:
        batch = "Rejected"

    return {
        "Keyword": keyword,
        "Normalized Keyword": k,
        "Search Volume": volume,
        "Keyword Difficulty": row.get("Keyword Difficulty", ""),
        "Keyword Intents": intents,
        "Business Scope": business_scope,
        "Hub": hub,
        "Macro Pillar": pillar,
        "Topic": topic,
        "Cluster": cluster,
        "Page / Article Candidate": page_candidate,
        "Functional Filter": functional,
        "Platform Filter": ", ".join(brand_hits),
        "Seller Profile": audience,
        "Search Intent Class": intent_class(k, buying, compare, street, parts, support, safety),
        "Site Section": section,
        "Page Role": role,
        "Risk Flags": ", ".join(risk_flags),
        "Geo / Language Intent": geo,
        "Production Decision": decision,
        "Production Batch": batch,
        "QA Score": qa_score,
        "QA Reason": decision_reason(business_scope, hub, functional, risk_flags, has_business_signal, volume, intents),
    }


def intent_class(k: str, buying: bool, compare: bool, street: bool, parts: bool, support: bool, safety: bool) -> str:
    if compare:
        return "comparison"
    if buying or any(term in k for term in ["for sale", "buy", "price", "cheap", "budget"]):
        return "commercial"
    if street:
        return "legal/compliance"
    if parts or support:
        return "support"
    if safety:
        return "safety/informational"
    return "informational"


def make_page_candidate(role: str, topic: str, keyword: str, brand_hits: list[str]) -> str:
    if role == "review":
        return "Review unqualified keyword intent"
    if brand_hits:
        brand = brand_hits[0].replace("sur ron", "surron").title()
        return f"Brand vs {brand}: electric dirt bike alternative guide"
    return f"{topic}: {keyword}"


STOPWORDS = {
    "electric", "dirt", "bike", "bikes", "for", "the", "and", "with", "near", "me",
    "best", "buy", "sale", "price", "adult", "adults", "kids", "youth", "street", "legal",
}


def keyword_category(k: str) -> str:
    tokens = [t for t in k.split() if len(t) > 2 and t not in STOPWORDS]
    return " ".join(tokens[:4])[:80] if tokens else ""


def budget_amount(k: str) -> str:
    match = re.search(r"\bunder\s+(?:usd\s*)?(\d{2,5})\b", k)
    return match.group(1) if match else ""


def keyword_brand(k: str) -> str:
    for brand in sorted(BRANDS, key=len, reverse=True):
        if re.search(rf"\b{re.escape(brand)}\b", k):
            return "Surron" if brand in {"surron", "sur ron"} else brand.title()
    return ""


def page_variant(row: pd.Series) -> tuple[str, str, str, str]:
    k = row["Normalized Keyword"]
    functional = row["Functional Filter"]
    topic = row["Topic"]
    role = row["Page Role"]
    category = keyword_category(k)

    if role == "review":
        return ("Review", "Review", "Unqualified / manual SERP validation", "Review unqualified keyword intent")
    if functional == "Shop / Buyer":
        if "for sale" in k or "buy" in k:
            return ("Core Page", "Shop", "Shop All For Sale", "Electric Dirt Bikes for Sale")
        if "price" in k or "cost" in k:
            return ("SEO Article", "Shop", "Pricing Guide", "Electric Dirt Bike Price Guide")
        if "best" in k:
            return ("SEO Article", "Shop", "Best Electric Dirt Bikes", "Best Electric Dirt Bikes")
        return ("Core Page", "Shop", "Shop All Electric Dirt Bikes", "Shop All Electric Dirt Bikes")
    if functional == "Adult / Performance":
        if re.search(r"\b72\s*(?:v|volt)\b", k):
            return ("Core Page", "Adult", "72V Powerhouses", "72V Electric Dirt Bikes")
        if re.search(r"\b60\s*(?:v|volt)\b", k):
            return ("Core Page", "Adult", "60V Performance", "60V Electric Dirt Bikes")
        if "speed" in k or "fast" in k or "mph" in k:
            return ("SEO Article", "Adult", "High-Speed Electric Dirt Bikes", "Fast Electric Dirt Bikes 40+ MPH")
        return ("Core Page", "Adult", "Adult Dirt Bikes", "Electric Dirt Bikes for Adults")
    if functional == "Kids / Youth":
        if any(age in k.split() for age in ["8", "9", "10", "11", "12"]):
            return ("Core Page", "Kids", "Ages 8-12", "Electric Dirt Bikes for Ages 8-12")
        if "teen" in k or "youth" in k or "13" in k:
            return ("Core Page", "Kids", "Teens & Youth", "Electric Dirt Bikes for Teens and Youth")
        if "24v" in k or "36v" in k or "safety" in k:
            return ("SEO Article", "Kids", "Safety First Low Voltage", "Low Voltage Electric Dirt Bikes for Kids")
        return ("Core Page", "Kids", "Kids & Youth", "Electric Dirt Bikes for Kids")
    if functional == "Street Legal":
        if "dual" in k:
            return ("Core Page", "Street Legal", "Dual-Sport Bikes", "Street Legal Dual-Sport Electric Dirt Bikes")
        if "register" in k or "license" in k:
            return ("SEO Article", "Street Legal", "Registry & Compliance Guide", "How to Register a Street Legal Electric Dirt Bike")
        return ("Core Page", "Street Legal", "Street Legal Bikes", "Street Legal Electric Dirt Bikes")
    if functional == "Comparison":
        if "surron" in k or "sur ron" in k:
            return ("SEO Article", "Compare", "Brand vs Surron", "Electric Dirt Bike vs Surron")
        if "razor" in k:
            return ("SEO Article", "Compare", "Brand vs Razor", "Electric Dirt Bike vs Razor")
        amount = budget_amount(k)
        if amount:
            return ("SEO Article", "Compare", "Budget by Price", f"Electric Dirt Bikes Under ${int(amount):,}")
        if "budget" in k or "cheap" in k or "affordable" in k:
            return ("SEO Article", "Compare", "Budget-Friendly Guides", "Best Budget Electric Dirt Bikes")
        if "financ" in k or "pay later" in k:
            return ("SEO Article", "Compare", "Financing", "Electric Dirt Bike Financing Guide")
        brand = keyword_brand(k)
        if brand:
            return ("SEO Article", "Compare", "Brand Alternatives", f"{brand} Electric Dirt Bike Alternatives")
        return ("SEO Article", "Compare", "Comparison Guide", "Electric Dirt Bike Comparison Guide")
    if functional == "Battery / Power":
        if "72v" in k and "60v" in k:
            return ("SEO Article", "Buying Guide", "72V vs 60V", "72V vs 60V Electric Dirt Bikes")
        if "battery" in k:
            return ("SEO Article", "Buying Guide", "Battery Life", "How Long Does an Electric Dirt Bike Battery Last")
        voltage = re.search(r"\b(24|36|48|60|72)\s*(?:v|volt)\b", k)
        if voltage:
            volts = voltage.group(1)
            return ("SEO Article", "Buying Guide", f"{volts}V Guide", f"{volts}V Electric Dirt Bike Guide")
        if "range" in k:
            return ("SEO Article", "Buying Guide", "Range", "Electric Dirt Bike Range Guide")
        if "charger" in k or "charging" in k:
            return ("SEO Article", "Buying Guide", "Charging", "Electric Dirt Bike Charging Guide")
        if "motor" in k or "watt" in k or re.search(r"\bkw\b", k):
            return ("SEO Article", "Buying Guide", "Motor and Power", "Electric Dirt Bike Motor and Power Guide")
        return ("SEO Article", "Buying Guide", "Power and Voltage", "Electric Dirt Bike Power and Voltage Guide")
    if functional == "Parts / Support":
        return ("SEO Article", "Support", "Parts and Maintenance", f"{category.title() if category else 'Electric Dirt Bike Parts'} Guide")
    if functional == "Safety / Rules":
        if "legal" in k or "law" in k or "rule" in k:
            return ("SEO Article", "Safety", "Laws and Rules", "Electric Dirt Bike Laws and Riding Rules")
        return ("SEO Article", "Safety", "Safety and Rules", "Electric Dirt Bike Safety Guide")
    return ("SEO Article", row["Hub"], topic, row["Page / Article Candidate"])


def expanded_page_candidate(row: pd.Series) -> str:
    return page_variant(row)[3]


PREFERRED_PRIMARY_KEYWORDS = {
    "72V Electric Dirt Bikes": "72v electric dirt bikes",
    "60V Electric Dirt Bikes": "60v electric dirt bike",
    "Street Legal Electric Dirt Bikes": "electric street legal dirt bike",
    "Electric Dirt Bikes for Sale": "electric dirt bikes for sale",
    "Electric Dirt Bike Price Guide": "electric dirt bike price",
    "Best Electric Dirt Bikes": "best electric dirt bikes",
    "Best Budget Electric Dirt Bikes": "budget friendly electric dirt bike",
}


def preferred_primary_keyword(candidate: str, top_keyword: str) -> str:
    return PREFERRED_PRIMARY_KEYWORDS.get(str(candidate), top_keyword)


def decision_reason(scope: str, hub: str, functional: str, flags: list[str], business: bool, volume: int, intents: str) -> str:
    bits = [f"{scope} fit in {hub}", f"functional={functional}"]
    if not business:
        bits.append("no clear electric dirt bike buyer/support signal")
    if flags:
        bits.append("flags=" + ", ".join(flags))
    if volume >= 10000:
        bits.append("high-volume term needs intent discipline")
    if intents:
        bits.append("source intents=" + intents)
    return "; ".join(bits)


def compact_join(series: pd.Series) -> str:
    values = []
    for raw in series.dropna().astype(str):
        for item in raw.split(","):
            item = item.strip()
            if item and item not in values:
                values.append(item)
    return ", ".join(values)


def aggregate_articles(master: pd.DataFrame) -> pd.DataFrame:
    keep = master[master["Production Decision"] != "Reject/deprioritize"].copy()
    if keep.empty:
        return pd.DataFrame()
    variants = keep.apply(page_variant, axis=1, result_type="expand")
    variants.columns = ["Page Tier", "Opportunity Type", "Opportunity Topic", "Expanded Page Candidate"]
    keep = pd.concat([keep.reset_index(drop=True), variants.reset_index(drop=True)], axis=1)
    group_cols = ["Hub", "Macro Pillar", "Topic", "Cluster", "Page / Article Candidate", "Page Role", "Site Section"]
    rows = []
    for key, grp in keep.groupby(group_cols, dropna=False):
        top = grp.sort_values(["Search Volume", "QA Score"], ascending=False).iloc[0]
        rows.append(dict(zip(group_cols, key), **article_row(grp, top, article_decision(grp), article_batch(grp))))
    return pd.DataFrame(rows).sort_values(["Production Batch", "Total Search Volume"], ascending=[True, False])


def aggregate_expanded_pages(master: pd.DataFrame) -> pd.DataFrame:
    keep = master[master["Production Decision"] != "Reject/deprioritize"].copy()
    if keep.empty:
        return pd.DataFrame()
    variants = keep.apply(page_variant, axis=1, result_type="expand")
    variants.columns = ["Page Tier", "Opportunity Type", "Opportunity Topic", "Expanded Page Candidate"]
    keep = pd.concat([keep.reset_index(drop=True), variants.reset_index(drop=True)], axis=1)
    group_cols = [
        "Hub", "Macro Pillar", "Topic", "Cluster", "Page Tier", "Opportunity Type",
        "Opportunity Topic", "Expanded Page Candidate", "Page Role", "Site Section",
    ]
    rows = []
    for key, grp in keep.groupby(group_cols, dropna=False):
        top = grp.sort_values(["Search Volume", "QA Score"], ascending=False).iloc[0]
        decision = expanded_decision(grp)
        rows.append(dict(zip(group_cols, key), **article_row(grp, top, decision, expanded_batch(decision, str(key[4])))))
    out = pd.DataFrame(rows)
    return out.sort_values(["Production Batch", "Page Tier", "Total Search Volume"], ascending=[True, True, False]) if not out.empty else out


def article_row(grp: pd.DataFrame, top: pd.Series, decision: str, batch: str) -> dict:
    candidate = str(grp.get("Expanded Page Candidate", pd.Series([""])).iloc[0]) if "Expanded Page Candidate" in grp else ""
    return {
        "Primary Keyword": preferred_primary_keyword(candidate, top["Keyword"]),
        "Keyword Count": len(grp),
        "Total Search Volume": int(grp["Search Volume"].sum()),
        "Average KD": round(pd.to_numeric(grp["Keyword Difficulty"], errors="coerce").mean(), 1),
        "Business Scope Mix": ", ".join(sorted(set(grp["Business Scope"]))),
        "Risk Flags": compact_join(grp["Risk Flags"]),
        "Production Decision": decision,
        "Production Batch": batch,
        "QA Score": int(round(grp["QA Score"].mean())),
        "Supporting Keywords": "; ".join(grp.sort_values("Search Volume", ascending=False)["Keyword"].astype(str).head(15)),
    }


def article_decision(grp: pd.DataFrame) -> str:
    counts = Counter(grp["Production Decision"])
    flags = compact_join(grp["Risk Flags"])
    if any(flag in flags for flag in ["out-of-scope", "gas-only", "unsafe-battery-modification"]):
        return "Reject/deprioritize"
    if "legal-compliance-review" in flags and counts["Build now"]:
        return "Manual review before build"
    if counts["Build now"]:
        return "Build now"
    if counts["Build later"]:
        return "Build later"
    if counts["Manual review before build"]:
        return "Manual review before build"
    if counts["Merge into parent/FAQ"]:
        return "Merge into parent/FAQ"
    return "Reject/deprioritize"


def expanded_decision(grp: pd.DataFrame) -> str:
    flags = compact_join(grp["Risk Flags"])
    tier = str(grp["Page Tier"].iloc[0])
    opportunity_type = str(grp["Opportunity Type"].iloc[0])
    counts = Counter(grp["Production Decision"])
    total_volume = int(grp["Search Volume"].sum())
    avg_score = float(grp["QA Score"].mean())
    if tier == "Review":
        return "Manual review before build"
    if any(flag in flags for flag in ["out-of-scope", "gas-only", "unsafe-battery-modification"]):
        return "Reject/deprioritize"
    if "legal-compliance-review" in flags and opportunity_type == "Street Legal":
        return "Manual review before build"
    if tier == "Core Page" and counts["Build now"] and avg_score >= 55:
        return "Build now"
    if opportunity_type == "Compare" and str(grp["Opportunity Topic"].iloc[0]) == "Budget by Price":
        if total_volume >= 2000 and avg_score >= 55:
            return "Build now"
        if total_volume >= 500 and avg_score >= 55:
            return "Build later"
        return "Merge into parent/FAQ"
    if opportunity_type in {"Shop", "Adult", "Kids", "Street Legal", "Compare"} and total_volume >= 500 and avg_score >= 55:
        return "Build now"
    if opportunity_type in {"Buying Guide", "Support", "Safety"} and total_volume >= 100:
        return "Build later"
    if counts["Manual review before build"]:
        return "Manual review before build"
    return "Merge into parent/FAQ"


def expanded_batch(decision: str, tier: str) -> str:
    if decision == "Build now" and tier == "Core Page":
        return "Batch 1 - Commercial/Core"
    if decision == "Build now":
        return "Batch 1 - SEO Articles"
    if decision == "Build later":
        return "Batch 2 - SEO Expansion"
    if decision == "Manual review before build":
        return "Review Queue"
    if decision == "Merge into parent/FAQ":
        return "Parent Support"
    return "Rejected"


def article_batch(grp: pd.DataFrame) -> str:
    decision = article_decision(grp)
    if decision == "Build now":
        batches = [b for b in grp["Production Batch"] if str(b).startswith("Batch 1")]
        return Counter(batches).most_common(1)[0][0] if batches else "Batch 1 - SEO Articles"
    if decision == "Build later":
        return "Batch 2 - SEO Expansion"
    if decision == "Manual review before build":
        return "Review Queue"
    if decision == "Merge into parent/FAQ":
        return "Parent Support"
    return "Rejected"


def summarize(master: pd.DataFrame, cols: list[str]) -> pd.DataFrame:
    if master.empty:
        return pd.DataFrame()
    rows = []
    for key, grp in master.groupby(cols, dropna=False):
        if not isinstance(key, tuple):
            key = (key,)
        rows.append(dict(zip(cols, key), **{
            "Keyword Count": len(grp),
            "Total Search Volume": int(grp["Search Volume"].sum()),
            "Build Now Keywords": int((grp["Production Decision"] == "Build now").sum()),
            "Review Keywords": int((grp["Production Decision"] == "Manual review before build").sum()),
            "Rejected Keywords": int((grp["Production Decision"] == "Reject/deprioritize").sum()),
            "Average QA Score": round(grp["QA Score"].mean(), 1),
            "Top Keyword": grp.sort_values("Search Volume", ascending=False).iloc[0]["Keyword"],
        }))
    return pd.DataFrame(rows).sort_values("Total Search Volume", ascending=False)


def execution_roadmap(pages: pd.DataFrame) -> pd.DataFrame:
    if pages.empty:
        return pd.DataFrame()
    rows = []
    for key, grp in pages.groupby(["Production Batch", "Hub", "Macro Pillar", "Cluster"], dropna=False):
        rows.append(dict(zip(["Production Batch", "Hub", "Macro Pillar", "Cluster"], key), **{
            "Page Count": len(grp),
            "Total Search Volume": int(grp["Total Search Volume"].sum()),
            "Build Now": int((grp["Production Decision"] == "Build now").sum()),
            "Build Later": int((grp["Production Decision"] == "Build later").sum()),
            "Review": int((grp["Production Decision"] == "Manual review before build").sum()),
            "Parent Support": int((grp["Production Decision"] == "Merge into parent/FAQ").sum()),
            "Top Page Candidate": grp.sort_values("Total Search Volume", ascending=False).iloc[0]["Expanded Page Candidate"],
        }))
    return pd.DataFrame(rows).sort_values(["Production Batch", "Total Search Volume"], ascending=[True, False])


def site_foundation() -> pd.DataFrame:
    rows = [
        ["Shop All", "All own-brand electric dirt bike products and product filters.", "Build now"],
        ["Adult Dirt Bikes", "Own-brand adult performance collection; supports high-speed, 72V and 60V paths.", "Build now"],
        ["Kids & Youth", "Own-brand youth collection split by age, teen fit and low-voltage safety.", "Build now"],
        ["Street Legal", "Own-brand street legal and dual-sport collection plus compliance content.", "Build after compliance review"],
        ["Compare & Save", "Blog-only competitor, alternative and budget acquisition content that links back to the own brand.", "SEO Blog only"],
        ["Buying Guides", "Footer SEO pool for battery, range, voltage, speed, safety and ownership questions.", "Build after core pages"],
        ["Parts And Support", "Batteries, chargers, controllers, brakes, tires, suspension, maintenance and warranty.", "Build after core pages"],
        ["Customer Support", "Contact, shipping, delivery and financing pages for conversion support.", "Confirm"],
    ]
    return pd.DataFrame(rows, columns=["Site Area", "Purpose", "Recommended Action"])


def audit(source_rows: int, unique_rows: int, master: pd.DataFrame, agent_path: Path) -> pd.DataFrame:
    counts = Counter(master["Production Decision"])
    rows = [
        ["Source rows loaded", "Enabled", source_rows, "Raw rows before duplicate cleanup"],
        ["Unique keywords retained", "Enabled", unique_rows, "One row per normalized keyword"],
        ["Duplicate rows removed", "Enabled", source_rows - unique_rows, "Duplicate volume is not summed"],
        ["Project agent rules", "Enabled" if agent_path.exists() else "Missing", str(agent_path), "Rules source for electric dirt bike business scope"],
        ["Rule-based classification", "Enabled", "regex + deterministic scoring", "No AI or embedding is claimed"],
        ["Navigation positioning", "Enabled", "Shop All / Adult / Kids / Street Legal / Buying Guides", "Competitor brands are Blog-only and excluded from commercial pillars"],
        ["Semantic embedding QA", "Not enabled", "", "Do not claim semantic recognition for this output"],
        ["Build now keyword count", "Enabled", counts.get("Build now", 0), "Qualified keyword-level decisions"],
        ["Manual review keyword count", "Enabled", counts.get("Manual review before build", 0), "Mixed or risky opportunities"],
        ["Rejected keyword count", "Enabled", counts.get("Reject/deprioritize", 0), "Out-of-scope or weak-fit traffic"],
    ]
    return pd.DataFrame(rows, columns=["Feature", "Status", "Evidence", "Notes"])


def write_workbook(output: Path, sheets: dict[str, pd.DataFrame]) -> None:
    output.parent.mkdir(parents=True, exist_ok=True)
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        for name, df in sheets.items():
            df.to_excel(writer, index=False, sheet_name=name[:31])
    wb = load_workbook(output)
    fill = PatternFill("solid", fgColor="1F4E78")
    font = Font(color="FFFFFF", bold=True)
    for ws in wb.worksheets:
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions
        for cell in ws[1]:
            cell.fill = fill
            cell.font = font
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        for col_idx, col_cells in enumerate(ws.iter_cols(min_row=1, max_row=min(ws.max_row, 200)), start=1):
            max_len = 10
            for cell in col_cells:
                if cell.value is not None:
                    max_len = max(max_len, min(64, len(str(cell.value)) + 2))
            ws.column_dimensions[get_column_letter(col_idx)].width = max_len
    wb.save(output)


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("--input", default=DEFAULT_INPUT)
    parser.add_argument("--agent", default=DEFAULT_AGENT)
    parser.add_argument("--output", default=str(DEFAULT_OUTPUT))
    args = parser.parse_args()

    input_path = Path(args.input)
    agent_path = Path(args.agent)
    output_path = Path(args.output)
    if not input_path.exists():
        raise FileNotFoundError(f"Input workbook not found: {input_path}")

    df = pd.read_excel(input_path, sheet_name=0)
    source_rows = len(df)
    if "Search Volume" not in df.columns and "Volume" in df.columns:
        df = df.rename(columns={"Volume": "Search Volume"})
    if "Keyword" not in df.columns or "Search Volume" not in df.columns:
        raise ValueError("Missing required source columns: Keyword and Search Volume/Volume")
    for optional in ["Keyword Difficulty", "Keyword Intents"]:
        if optional not in df.columns:
            df[optional] = ""
    df["_Normalized Dedup Key"] = df["Keyword"].map(norm)
    df["_Volume Sort"] = pd.to_numeric(df["Search Volume"], errors="coerce").fillna(0)
    df = (
        df.sort_values("_Volume Sort", ascending=False)
        .drop_duplicates("_Normalized Dedup Key", keep="first")
        .drop(columns=["_Normalized Dedup Key", "_Volume Sort"])
        .reset_index(drop=True)
    )

    classified = pd.DataFrame([classify(row) for _, row in df.iterrows()])
    articles = aggregate_articles(classified)
    expanded_pages = aggregate_expanded_pages(classified)
    topic_map = summarize(classified, ["Hub", "Macro Pillar"])
    topic_mid = summarize(classified[classified["Business Scope"] != "Out of Scope"], ["Hub", "Macro Pillar", "Topic"])
    clusters = summarize(classified[classified["Business Scope"] != "Out of Scope"], ["Hub", "Macro Pillar", "Topic", "Cluster"])
    qa_topics = topic_mid.copy()
    if not qa_topics.empty:
        qa_topics["QA Decision"] = qa_topics["Average QA Score"].apply(lambda x: "Pass" if x >= 50 else "Review")
    qa_clusters = clusters.copy()
    if not qa_clusters.empty:
        qa_clusters["QA Decision"] = qa_clusters["Average QA Score"].apply(lambda x: "Pass" if x >= 50 else "Review")
    qa_articles = expanded_pages.copy()
    if not qa_articles.empty:
        qa_articles["QA Decision"] = qa_articles["QA Score"].apply(lambda x: "Pass" if x >= 55 else "Review")

    sheets = {
        "Keyword_Master": classified,
        "Topic_Map": topic_map,
        "Topic_Mid": topic_mid,
        "Mid_Clusters": clusters,
        "Article_Plan_Mid": articles,
        "Page_Opportunity_Expansion": expanded_pages,
        "QA_Scored_Topics": qa_topics,
        "QA_Scored_Clusters": qa_clusters,
        "QA_Scored_Articles": qa_articles,
        "Production_Batches": expanded_pages,
        "Execution_Roadmap": execution_roadmap(expanded_pages),
        "Site_Foundation": site_foundation(),
        "Feature_Consistency_Audit": audit(source_rows, len(df), classified, agent_path),
    }
    write_workbook(output_path, sheets)
    print(f"wrote {output_path.resolve()}")
    print(f"source rows: {source_rows}")
    print(f"unique keywords after deduplication: {len(df)}")
    print(f"production candidates: {len(expanded_pages)}")
    print(f"decisions: {dict(Counter(expanded_pages['Production Decision'])) if not expanded_pages.empty else {}}")


if __name__ == "__main__":
    main()
