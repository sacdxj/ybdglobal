from __future__ import annotations

import argparse
import json
import re
import shutil
import time
import unicodedata
import urllib.error
import urllib.request
from collections import Counter, defaultdict
from pathlib import Path

import numpy as np
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


DEFAULT_SOURCE = Path("outputs") / "electric_dirt_bike_keyword_architecture.xlsx"
DEFAULT_OUTPUT = Path("outputs") / "electric_dirt_bike_keyword_architecture_semantic.xlsx"
DEFAULT_CACHE = Path("outputs") / "electric_dirt_bike_semantic_embedding_cache.jsonl"
DEFAULT_ENV = Path(r"D:\obsidian-erp\erp.env")

MODEL = "gemini-embedding-2"
BATCH_SIZE = 50
API_RETRIES = 4
SEMANTIC_PASS = 0.86
SEMANTIC_REVIEW = 0.80
GROUP_MIN_SIZE = 3


def norm(text) -> str:
    text = str(text or "").strip().lower()
    text = unicodedata.normalize("NFKD", text)
    text = "".join(ch for ch in text if not unicodedata.combining(ch))
    text = re.sub(r"[^a-z0-9\s\-\/]", " ", text)
    text = re.sub(r"\s+", " ", text).strip()
    return text


def load_key(env_file: Path) -> str:
    text = env_file.read_text(encoding="utf-8-sig", errors="ignore")
    for line in text.splitlines():
        s = line.strip()
        if not s or s.startswith("#"):
            continue
        m = re.match(r"(?:\$env:)?(GOOGLE_API_KEY|GEMINI_API_KEY|GOOGLE_GENERATIVE_AI_API_KEY)\s*=\s*(.*)", s)
        if m:
            return m.group(2).strip().strip('"').strip("'")
    raise RuntimeError(f"No Google/Gemini API key found in {env_file}")


def load_cache(cache_file: Path) -> dict[str, list[float]]:
    cache = {}
    if not cache_file.exists():
        return cache
    with cache_file.open("r", encoding="utf-8") as f:
        for line in f:
            if not line.strip():
                continue
            obj = json.loads(line)
            cache[obj["text"]] = obj["embedding"]
    return cache


def append_cache(cache_file: Path, items) -> None:
    cache_file.parent.mkdir(parents=True, exist_ok=True)
    with cache_file.open("a", encoding="utf-8") as f:
        for text, emb in items:
            f.write(json.dumps({"text": text, "embedding": emb}, ensure_ascii=False) + "\n")


def embed_texts(texts: list[str], env_file: Path, cache_file: Path) -> list[list[float]]:
    key = load_key(env_file)
    cache = load_cache(cache_file)
    missing = [t for t in texts if t not in cache]
    print(f"semantic embedding cache hit {len(texts) - len(missing)} / {len(texts)}; missing {len(missing)}")

    url = f"https://generativelanguage.googleapis.com/v1beta/models/{MODEL}:batchEmbedContents?key={key}"
    for start in range(0, len(missing), BATCH_SIZE):
        batch = missing[start : start + BATCH_SIZE]
        payload = {
            "requests": [
                {
                    "model": f"models/{MODEL}",
                    "content": {"parts": [{"text": text}]},
                    "taskType": "CLUSTERING",
                }
                for text in batch
            ]
        }
        req = urllib.request.Request(
            url,
            data=json.dumps(payload).encode("utf-8"),
            headers={"Content-Type": "application/json"},
            method="POST",
        )
        last_error = None
        for attempt in range(1, API_RETRIES + 1):
            try:
                with urllib.request.urlopen(req, timeout=90) as resp:
                    data = json.loads(resp.read().decode("utf-8"))
                break
            except urllib.error.HTTPError as exc:
                detail = exc.read().decode("utf-8", errors="replace")[:500]
                last_error = RuntimeError(f"HTTP {exc.code}: {detail}")
                if exc.code < 500 and exc.code not in {408, 429}:
                    raise last_error from exc
                print(f"semantic batch retry {attempt}/{API_RETRIES}: HTTP {exc.code}")
            except (urllib.error.URLError, TimeoutError) as exc:
                last_error = exc
                print(f"semantic batch retry {attempt}/{API_RETRIES}: {exc}")
            if attempt < API_RETRIES:
                time.sleep(2 * attempt)
        else:
            raise RuntimeError(f"Google embedding API failed after {API_RETRIES} retries: {last_error}") from last_error

        embs = [e.get("values", []) for e in data.get("embeddings", [])]
        if len(embs) != len(batch):
            raise RuntimeError(f"Embedding count mismatch: got {len(embs)} for {len(batch)}")
        append_cache(cache_file, zip(batch, embs))
        for text, emb in zip(batch, embs):
            cache[text] = emb
        print(f"embedded {min(start + len(batch), len(missing))} / {len(missing)}")
        time.sleep(0.2)
    return [cache[t] for t in texts]


def headers(ws) -> dict[str, int]:
    return {cell.value: i + 1 for i, cell in enumerate(ws[1])}


def row_dict(ws, row_idx: int, h: dict[str, int]) -> dict:
    return {name: ws.cell(row_idx, col).value for name, col in h.items()}


def semantic_text(row: dict) -> str:
    parts = [
        row.get("Primary Keyword"),
        row.get("Expanded Page Candidate") or row.get("Page / Article Candidate"),
        row.get("Opportunity Topic"),
        row.get("Opportunity Type"),
        row.get("Search Intent Class"),
        row.get("Page Role"),
        row.get("Supporting Keywords"),
    ]
    return " | ".join(str(p or "") for p in parts)


def keyword_semantic_text(row: dict) -> str:
    parts = [
        row.get("Keyword"),
        row.get("Search Intent Class"),
        row.get("Page Role"),
        row.get("Functional Filter"),
        row.get("Seller Profile"),
    ]
    return " | ".join(str(p or "") for p in parts)


def l2_normalize(arr: np.ndarray) -> np.ndarray:
    arr = np.asarray(arr, dtype=np.float32)
    norms = np.linalg.norm(arr, axis=1, keepdims=True)
    norms[norms == 0] = 1
    return arr / norms


def centroid(vectors: np.ndarray) -> np.ndarray:
    c = np.mean(vectors, axis=0)
    n = np.linalg.norm(c)
    if n == 0:
        return c
    return c / n


def group_purity(vectors: np.ndarray) -> float:
    if len(vectors) <= 1:
        return 1.0
    c = centroid(vectors)
    return float(np.mean(vectors @ c))


def semantic_decision(cluster_sim: float, cluster_purity: float, topic_purity: float, group_size: int, production_decision: str, page_role: str):
    if group_size < GROUP_MIN_SIZE:
        return "Semantic Pass", "Small group; semantic gate is supportive only."
    if page_role in {"review", "FAQ/section"}:
        if cluster_sim < SEMANTIC_REVIEW:
            return "Semantic Review", "Review/FAQ candidate has weak semantic fit with its cluster."
        return "Semantic Pass", "Review/FAQ candidate is coherent enough for its workflow."
    if cluster_sim >= SEMANTIC_PASS and cluster_purity >= SEMANTIC_PASS and topic_purity >= SEMANTIC_REVIEW:
        return "Semantic Pass", "Embedding similarity supports the assigned Topic/Cluster."
    if cluster_sim >= SEMANTIC_REVIEW and cluster_purity >= SEMANTIC_REVIEW:
        return "Semantic Review", "Embedding similarity is borderline; keep for manual review before production."
    if production_decision == "Build now":
        return "Semantic Block", "Build-now candidate has weak embedding fit; downgrade to manual review."
    return "Semantic Review", "Embedding fit is weak or mixed; do not treat as automatic production."


def semantic_production_decision(original: str, gate: str) -> str:
    if gate == "Semantic Block" and original == "Build now":
        return "Manual review before build"
    return original


def style_sheet(ws) -> None:
    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    for col_idx, col in enumerate(ws.iter_cols(min_row=1, max_row=min(ws.max_row, 200)), 1):
        max_len = 10
        for cell in col:
            max_len = max(max_len, min(70, len(str(cell.value or ""))))
        ws.column_dimensions[get_column_letter(col_idx)].width = max_len + 2


def append_columns(ws, new_headers: list[str], values_by_row: dict[int, list]) -> None:
    existing = headers(ws)
    for name in new_headers:
        if name in existing:
            raise RuntimeError(f"Sheet {ws.title} already has semantic column: {name}")
    start = ws.max_column + 1
    for offset, name in enumerate(new_headers):
        ws.cell(1, start + offset).value = name
    for row_idx, vals in values_by_row.items():
        for offset, val in enumerate(vals):
            ws.cell(row_idx, start + offset).value = val
    style_sheet(ws)


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("--source", default=str(DEFAULT_SOURCE))
    parser.add_argument("--output", default=str(DEFAULT_OUTPUT))
    parser.add_argument("--cache", default=str(DEFAULT_CACHE))
    parser.add_argument("--env", default=str(DEFAULT_ENV))
    args = parser.parse_args()

    source = Path(args.source)
    output = Path(args.output)
    cache_file = Path(args.cache)
    env_file = Path(args.env)

    if not source.exists():
        raise FileNotFoundError(f"Missing source workbook: {source}")
    if not env_file.exists():
        raise FileNotFoundError(f"Missing env file with Google/Gemini API key: {env_file}")

    output.parent.mkdir(parents=True, exist_ok=True)
    shutil.copy2(source, output)
    wb = load_workbook(output)
    keyword_master = wb["Keyword_Master"]
    kh = headers(keyword_master)
    qa = wb["QA_Scored_Articles"]
    h = headers(qa)

    required = ["Hub", "Macro Pillar", "Topic", "Cluster", "Primary Keyword", "Production Decision", "Page Role"]
    missing = [col for col in required if col not in h]
    if missing:
        raise RuntimeError(f"QA_Scored_Articles missing required columns: {missing}")
    page_col = "Expanded Page Candidate" if "Expanded Page Candidate" in h else "Page / Article Candidate"
    keyword_required = ["Keyword", "Hub", "Macro Pillar", "Topic", "Cluster", "Production Decision"]
    keyword_missing = [col for col in keyword_required if col not in kh]
    if keyword_missing:
        raise RuntimeError(f"Keyword_Master missing required columns: {keyword_missing}")

    rows = []
    texts = []
    for row_idx in range(2, qa.max_row + 1):
        row = row_dict(qa, row_idx, h)
        row["_row_idx"] = row_idx
        rows.append(row)
        texts.append(semantic_text(row))

    if not rows:
        raise RuntimeError("No article candidates found for semantic QA")

    keyword_rows = []
    keyword_texts = []
    for row_idx in range(2, keyword_master.max_row + 1):
        row = row_dict(keyword_master, row_idx, kh)
        if row.get("Production Decision") == "Reject/deprioritize":
            continue
        row["_row_idx"] = row_idx
        keyword_rows.append(row)
        keyword_texts.append(keyword_semantic_text(row))

    if not keyword_rows:
        raise RuntimeError("No non-rejected keyword rows found for semantic QA")

    all_texts = keyword_texts + texts
    all_vectors = l2_normalize(np.array(embed_texts(all_texts, env_file, cache_file), dtype=np.float32))
    keyword_vectors = all_vectors[: len(keyword_rows)]
    vectors = all_vectors[len(keyword_rows) :]

    topic_groups = defaultdict(list)
    cluster_groups = defaultdict(list)
    for i, row in enumerate(keyword_rows):
        topic_groups[(row.get("Hub"), row.get("Macro Pillar"), row.get("Topic"))].append(i)
        cluster_groups[(row.get("Hub"), row.get("Macro Pillar"), row.get("Topic"), row.get("Cluster"))].append(i)

    topic_purity = {}
    cluster_purity = {}
    topic_centroids = {}
    cluster_centroids = {}
    for key, idxs in topic_groups.items():
        vecs = keyword_vectors[idxs]
        topic_purity[key] = group_purity(vecs)
        topic_centroids[key] = centroid(vecs)
    for key, idxs in cluster_groups.items():
        vecs = keyword_vectors[idxs]
        cluster_purity[key] = group_purity(vecs)
        cluster_centroids[key] = centroid(vecs)

    semantic_headers = [
        "Semantic Enabled",
        "Semantic Model",
        "Semantic Topic Similarity",
        "Semantic Cluster Similarity",
        "Semantic Topic Purity",
        "Semantic Cluster Purity",
        "Semantic Cluster Size",
        "Semantic QA Decision",
        "Semantic Production Decision",
        "Semantic QA Reason",
    ]

    qa_values = {}
    semantic_rows = []
    decision_counts = Counter()
    adjusted_counts = Counter()
    values_by_candidate = {}
    keyword_values = {}
    keyword_semantic_rows = []
    keyword_decision_counts = Counter()
    keyword_adjusted_counts = Counter()

    for i, row in enumerate(keyword_rows):
        topic_key = (row.get("Hub"), row.get("Macro Pillar"), row.get("Topic"))
        cluster_key = (row.get("Hub"), row.get("Macro Pillar"), row.get("Topic"), row.get("Cluster"))
        topic_sim = float(keyword_vectors[i] @ topic_centroids[topic_key])
        cluster_sim = float(keyword_vectors[i] @ cluster_centroids[cluster_key])
        tp = topic_purity[topic_key]
        cp = cluster_purity[cluster_key]
        group_size = len(cluster_groups[cluster_key])
        gate, reason = semantic_decision(
            cluster_sim,
            cp,
            tp,
            group_size,
            row.get("Production Decision"),
            row.get("Page Role"),
        )
        adjusted = semantic_production_decision(row.get("Production Decision"), gate)
        keyword_decision_counts[gate] += 1
        keyword_adjusted_counts[adjusted] += 1
        vals = [
            "Yes",
            MODEL,
            round(topic_sim, 3),
            round(cluster_sim, 3),
            round(tp, 3),
            round(cp, 3),
            group_size,
            gate,
            adjusted,
            reason,
        ]
        keyword_values[row["_row_idx"]] = vals
        keyword_semantic_rows.append(
            [
                row.get("Keyword"),
                row.get("Hub"),
                row.get("Macro Pillar"),
                row.get("Topic"),
                row.get("Cluster"),
                row.get("Production Decision"),
                adjusted,
                round(topic_sim, 3),
                round(cluster_sim, 3),
                round(tp, 3),
                round(cp, 3),
                group_size,
                gate,
                reason,
            ]
        )

    for i, row in enumerate(rows):
        topic_key = (row.get("Hub"), row.get("Macro Pillar"), row.get("Topic"))
        cluster_key = (row.get("Hub"), row.get("Macro Pillar"), row.get("Topic"), row.get("Cluster"))
        if topic_key not in topic_centroids or cluster_key not in cluster_centroids:
            topic_sim = cluster_sim = tp = cp = 1.0
            group_size = 1
        else:
            topic_sim = float(vectors[i] @ topic_centroids[topic_key])
            cluster_sim = float(vectors[i] @ cluster_centroids[cluster_key])
            tp = topic_purity[topic_key]
            cp = cluster_purity[cluster_key]
            group_size = len(cluster_groups[cluster_key])
        gate, reason = semantic_decision(
            cluster_sim,
            cp,
            tp,
            group_size,
            row.get("Production Decision"),
            row.get("Page Role"),
        )
        adjusted = semantic_production_decision(row.get("Production Decision"), gate)
        decision_counts[gate] += 1
        adjusted_counts[adjusted] += 1
        vals = [
            "Yes",
            MODEL,
            round(topic_sim, 3),
            round(cluster_sim, 3),
            round(tp, 3),
            round(cp, 3),
            group_size,
            gate,
            adjusted,
            reason,
        ]
        qa_values[row["_row_idx"]] = vals
        key = (
            norm(row.get(page_col)),
            norm(row.get("Primary Keyword")),
            row.get("Total Search Volume"),
        )
        values_by_candidate[key] = vals
        semantic_rows.append(
            [
                row.get("Hub"),
                row.get("Macro Pillar"),
                row.get("Topic"),
                row.get("Cluster"),
                row.get(page_col),
                row.get("Primary Keyword"),
                row.get("Production Decision"),
                adjusted,
                round(topic_sim, 3),
                round(cluster_sim, 3),
                round(tp, 3),
                round(cp, 3),
                group_size,
                gate,
                reason,
            ]
        )

    append_columns(keyword_master, semantic_headers, keyword_values)
    append_columns(qa, semantic_headers, qa_values)

    if "Production_Batches" in wb.sheetnames:
        prod = wb["Production_Batches"]
        ph = headers(prod)
        prod_page_col = "Expanded Page Candidate" if "Expanded Page Candidate" in ph else "Page / Article Candidate"
        prod_values = {}
        for row_idx in range(2, prod.max_row + 1):
            key = (
                norm(prod.cell(row_idx, ph[prod_page_col]).value),
                norm(prod.cell(row_idx, ph["Primary Keyword"]).value),
                prod.cell(row_idx, ph["Total Search Volume"]).value,
            )
            vals = values_by_candidate.get(key)
            if vals:
                prod_values[row_idx] = vals
        append_columns(prod, semantic_headers, prod_values)

    for sheet_name in ["Keyword_Semantic_QA", "Semantic_QA", "Semantic_Group_Summary", "Feature_Consistency_Audit"]:
        if sheet_name in wb.sheetnames:
            del wb[sheet_name]

    ws = wb.create_sheet("Keyword_Semantic_QA")
    ws.append(
        [
            "Keyword",
            "Hub",
            "Macro Pillar",
            "Topic",
            "Cluster",
            "Original Production Decision",
            "Semantic Production Decision",
            "Semantic Topic Similarity",
            "Semantic Cluster Similarity",
            "Semantic Topic Purity",
            "Semantic Cluster Purity",
            "Semantic Cluster Size",
            "Semantic QA Decision",
            "Semantic QA Reason",
        ]
    )
    for row in sorted(keyword_semantic_rows, key=lambda r: ({"Semantic Block": 0, "Semantic Review": 1, "Semantic Pass": 2}.get(r[12], 9), r[9], r[10], r[0])):
        ws.append(row)
    style_sheet(ws)

    ws = wb.create_sheet("Semantic_QA")
    ws.append(
        [
            "Hub",
            "Macro Pillar",
            "Topic",
            "Cluster",
            page_col,
            "Primary Keyword",
            "Original Production Decision",
            "Semantic Production Decision",
            "Semantic Topic Similarity",
            "Semantic Cluster Similarity",
            "Semantic Topic Purity",
            "Semantic Cluster Purity",
            "Semantic Cluster Size",
            "Semantic QA Decision",
            "Semantic QA Reason",
        ]
    )
    for row in sorted(semantic_rows, key=lambda r: ({"Semantic Block": 0, "Semantic Review": 1, "Semantic Pass": 2}.get(r[13], 9), r[10], r[11])):
        ws.append(row)
    style_sheet(ws)

    ws = wb.create_sheet("Semantic_Group_Summary")
    ws.append(
        [
            "Level",
            "Hub",
            "Macro Pillar",
            "Topic",
            "Cluster",
            "Item Count",
            "Semantic Purity",
            "Low Similarity Count",
            "Build Now Count",
            "Recommendation",
        ]
    )
    for key, idxs in topic_groups.items():
        low = sum(1 for i in idxs if float(keyword_vectors[i] @ topic_centroids[key]) < SEMANTIC_REVIEW)
        build_now = sum(1 for i in idxs if keyword_rows[i].get("Production Decision") == "Build now")
        rec = "Review / split topic" if low or topic_purity[key] < SEMANTIC_REVIEW else "OK"
        ws.append(["Topic", key[0], key[1], key[2], "", len(idxs), round(topic_purity[key], 3), low, build_now, rec])
    for key, idxs in cluster_groups.items():
        low = sum(1 for i in idxs if float(keyword_vectors[i] @ cluster_centroids[key]) < SEMANTIC_REVIEW)
        build_now = sum(1 for i in idxs if keyword_rows[i].get("Production Decision") == "Build now")
        rec = "Review / split cluster" if len(idxs) >= GROUP_MIN_SIZE and (low or cluster_purity[key] < SEMANTIC_REVIEW) else "OK"
        ws.append(["Cluster", key[0], key[1], key[2], key[3], len(idxs), round(cluster_purity[key], 3), low, build_now, rec])
    style_sheet(ws)

    ws = wb.create_sheet("Feature_Consistency_Audit")
    ws.append(["Feature", "Status", "Evidence", "Notes"])
    audit_rows = [
        ["Semantic embedding layer", "Enabled", MODEL, f"{len(keyword_rows)} keyword rows and {len(rows)} article/page candidates embedded and scored"],
        ["Semantic cache", "Enabled" if cache_file.exists() else "Missing", str(cache_file), "Local Electric Dirt Bike cache, separate from ERP cache"],
        ["Semantic QA fields", "Enabled", ", ".join(semantic_headers), "Fields appended to Keyword_Master, QA_Scored_Articles and Production_Batches"],
        ["Keyword_Semantic_QA sheet", "Enabled", len(keyword_semantic_rows), "One row per non-rejected keyword"],
        ["Semantic_QA sheet", "Enabled", len(semantic_rows), "One row per article/page candidate"],
        ["Semantic_Group_Summary sheet", "Enabled", len(topic_groups) + len(cluster_groups), "Topic and cluster purity summary"],
        ["Article semantic block count", decision_counts.get("Semantic Block", 0), "", "Article/page candidates blocked by semantic QA"],
        ["Article semantic review count", decision_counts.get("Semantic Review", 0), "", "Article/page candidates needing manual semantic review"],
        ["Article semantic pass count", decision_counts.get("Semantic Pass", 0), "", "Article/page candidates passing semantic coherence gate"],
        ["Keyword semantic block count", keyword_decision_counts.get("Semantic Block", 0), "", "Keyword rows blocked by semantic QA"],
        ["Keyword semantic review count", keyword_decision_counts.get("Semantic Review", 0), "", "Keyword rows needing manual semantic review"],
        ["Keyword semantic pass count", keyword_decision_counts.get("Semantic Pass", 0), "", "Keyword rows passing semantic coherence gate"],
    ]
    for row in audit_rows:
        ws.append(row)
    style_sheet(ws)

    wb.save(output)
    print(f"wrote {output.resolve()}")
    print(f"embedded keyword rows: {len(keyword_rows)}")
    print(f"embedded candidates: {len(rows)}")
    print(f"semantic decisions: {decision_counts}")
    print(f"semantic production decisions: {adjusted_counts}")
    print(f"keyword semantic decisions: {keyword_decision_counts}")
    print(f"keyword semantic production decisions: {keyword_adjusted_counts}")


if __name__ == "__main__":
    main()
