from __future__ import annotations

import argparse
import json
import math
import os
import re
import shutil
import time
import unicodedata
import urllib.error
import urllib.parse
import urllib.request
from collections import Counter, defaultdict
from pathlib import Path
from urllib.parse import urlparse

from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


DEFAULT_SOURCE = Path("outputs") / "electric_dirt_bike_keyword_architecture_semantic.xlsx"
DEFAULT_OUTPUT = Path("outputs") / "electric_dirt_bike_keyword_architecture_serp.xlsx"
DEFAULT_CACHE = Path("outputs") / "electric_dirt_bike_serp_cache.jsonl"
DEFAULT_ENV = Path(r"D:\obsidian-erp\erp.env")

SERPAPI_ENDPOINT = "https://serpapi.com/search.json"
SERP_RETRIES = 3
REQUEST_SLEEP_SECONDS = 1.0

BUSINESS_TERMS = {
    "electric", "dirt", "bike", "bikes", "motocross", "offroad", "off-road",
    "adult", "adults", "kids", "youth", "teen", "street", "legal", "dual",
    "sport", "72v", "60v", "48v", "battery", "range", "charger", "motor",
    "speed", "mph", "surron", "razor", "talaria", "alternative", "compare",
    "best", "price", "budget", "dealer", "wholesale", "parts", "suspension",
    "brakes", "tires", "helmet", "maintenance", "warranty",
}
CONSUMER_RISK_TERMS = {
    "coupon", "promo", "discount code", "login", "tracking", "app", "download",
    "game", "movie", "toy only", "mobility scooter", "atv", "gas only",
}
PORTUGUESE_STOPWORDS = {
    "a", "an", "the", "and", "or", "of", "for", "to", "in", "on", "with", "without",
    "near", "me", "best", "buy", "sale", "online", "bike", "bikes", "dirt", "electric",
}


def norm(text) -> str:
    value = "" if text is None else str(text)
    value = unicodedata.normalize("NFKD", value.lower())
    value = "".join(ch for ch in value if not unicodedata.combining(ch))
    value = re.sub(r"[^a-z0-9\s\-\/\.]", " ", value)
    value = re.sub(r"\s+", " ", value).strip()
    return value


def tokens(text) -> set[str]:
    return {t for t in re.findall(r"[a-z0-9]+", norm(text)) if len(t) > 2 and t not in PORTUGUESE_STOPWORDS}


def load_env_key(env_file: Path | None, explicit_key: str | None) -> str:
    if explicit_key:
        return explicit_key.strip()
    for name in ["SERPAPI_API_KEY", "SERPAPI_KEY"]:
        if os.environ.get(name):
            return os.environ[name].strip()
    if env_file and env_file.exists():
        text = env_file.read_text(encoding="utf-8-sig", errors="ignore")
        for line in text.splitlines():
            s = line.strip()
            if not s or s.startswith("#"):
                continue
            m = re.match(r"(?:\$env:)?(SERPAPI_API_KEY|SERPAPI_KEY)\s*=\s*(.*)", s)
            if m:
                return m.group(2).strip().strip('"').strip("'")
    raise RuntimeError(
        "No SerpAPI key found. Pass --serpapi-key, set SERPAPI_API_KEY, or add SERPAPI_API_KEY to the env file."
    )


def load_cache(cache_file: Path) -> dict[str, dict]:
    cache = {}
    if not cache_file.exists():
        return cache
    with cache_file.open("r", encoding="utf-8") as f:
        for line in f:
            if not line.strip():
                continue
            obj = json.loads(line)
            cache[obj["query"]] = obj["result"]
    return cache


def append_cache(cache_file: Path, query: str, result: dict) -> None:
    cache_file.parent.mkdir(parents=True, exist_ok=True)
    with cache_file.open("a", encoding="utf-8") as f:
        f.write(json.dumps({"query": query, "result": result}, ensure_ascii=False) + "\n")


def fetch_serp(query: str, api_key: str, cache: dict[str, dict], cache_file: Path, gl: str, hl: str, num: int) -> dict:
    query_key = norm(query)
    if query_key in cache:
        return cache[query_key]

    params = {
        "engine": "google",
        "q": query,
        "api_key": api_key,
        "gl": gl,
        "hl": hl,
        "num": str(num),
    }
    url = SERPAPI_ENDPOINT + "?" + urllib.parse.urlencode(params)
    last_error = None
    for attempt in range(1, SERP_RETRIES + 1):
        try:
            with urllib.request.urlopen(url, timeout=60) as resp:
                data = json.loads(resp.read().decode("utf-8"))
            result = parse_serp_response(data)
            cache[query_key] = result
            append_cache(cache_file, query_key, result)
            time.sleep(REQUEST_SLEEP_SECONDS)
            return result
        except urllib.error.HTTPError as exc:
            detail = exc.read().decode("utf-8", errors="replace")[:500]
            last_error = f"HTTP {exc.code}: {detail}"
            if exc.code < 500 and exc.code not in {408, 429}:
                break
        except (urllib.error.URLError, TimeoutError, json.JSONDecodeError) as exc:
            last_error = str(exc)
        if attempt < SERP_RETRIES:
            time.sleep(2 * attempt)
    result = {"status": "error", "message": str(last_error), "top_urls": [], "titles": [], "snippets": [], "domains": []}
    cache[query_key] = result
    append_cache(cache_file, query_key, result)
    return result


def parse_serp_response(data: dict) -> dict:
    organic = data.get("organic_results") or []
    rows = []
    for item in organic[:10]:
        link = item.get("link") or item.get("displayed_link") or ""
        title = item.get("title") or ""
        snippet = item.get("snippet") or ""
        domain = domain_of(link)
        rows.append(
            {
                "position": item.get("position"),
                "url": link,
                "domain": domain,
                "title": title,
                "snippet": snippet,
            }
        )
    return {
        "status": "ok" if rows else "empty",
        "message": "",
        "top_urls": [r["url"] for r in rows if r["url"]],
        "domains": [r["domain"] for r in rows if r["domain"]],
        "titles": [r["title"] for r in rows if r["title"]],
        "snippets": [r["snippet"] for r in rows if r["snippet"]],
        "organic_results": rows,
    }


def domain_of(url: str) -> str:
    if not url:
        return ""
    parsed = urlparse(url if "://" in url else "https://" + url)
    host = parsed.netloc.lower().replace("www.", "")
    return host.split(":")[0]


def serp_overlap(a: dict, b: dict) -> tuple[float, float, float]:
    urls_a = set(a.get("top_urls") or [])
    urls_b = set(b.get("top_urls") or [])
    dom_a = set(a.get("domains") or [])
    dom_b = set(b.get("domains") or [])
    text_a = tokens(" ".join((a.get("titles") or []) + (a.get("snippets") or [])))
    text_b = tokens(" ".join((b.get("titles") or []) + (b.get("snippets") or [])))
    url_score = jaccard(urls_a, urls_b)
    domain_score = jaccard(dom_a, dom_b)
    text_score = jaccard(text_a, text_b)
    blended = round((url_score * 0.5) + (domain_score * 0.3) + (text_score * 0.2), 3)
    return blended, round(domain_score, 3), round(text_score, 3)


def jaccard(a: set, b: set) -> float:
    if not a or not b:
        return 0.0
    return len(a & b) / len(a | b)


def serp_intent_match(serp: dict) -> tuple[str, str, str]:
    haystack = norm(" ".join((serp.get("titles") or []) + (serp.get("snippets") or [])))
    found_business = sorted(t for t in BUSINESS_TERMS if t in haystack)
    found_risk = sorted(t for t in CONSUMER_RISK_TERMS if t in haystack)
    if found_risk and len(found_risk) >= len(found_business):
        return "Weak", ", ".join(found_risk[:8]), "SERP leans consumer/navigation; review before production."
    if len(found_business) >= 4:
        return "Strong", ", ".join(found_business[:10]), "SERP reinforces electric dirt bike buyer/product/support intent."
    if len(found_business) >= 2:
        return "Medium", ", ".join(found_business[:10]), "SERP has partial electric dirt bike business fit; keep as validation evidence."
    return "Weak", ", ".join(found_business[:10]), "SERP does not clearly reinforce electric dirt bike business intent."


def frequent_terms(serp: dict, supporting_keywords: str, primary_keyword: str, max_terms: int = 12) -> list[str]:
    counter = Counter()
    text = " ".join((serp.get("titles") or []) + (serp.get("snippets") or []))
    for t in tokens(text):
        if len(t) > 3:
            counter[t] += 2
    for kw in str(supporting_keywords or "").split(";"):
        for t in tokens(kw):
            if len(t) > 3:
                counter[t] += 3
    for t in tokens(primary_keyword):
        counter[t] += 4
    return [term for term, _ in counter.most_common(max_terms)]


def recommended_h1(page: dict) -> str:
    candidate = str(page.get("Expanded Page Candidate") or page.get("Primary Keyword") or "").strip()
    primary = str(page.get("Primary Keyword") or "").strip()
    if candidate and len(candidate) <= 90:
        return sentence_case(candidate)
    return sentence_case(primary)


def sentence_case(text: str) -> str:
    text = re.sub(r"\s+", " ", str(text or "").strip())
    if not text:
        return ""
    return text[:1].upper() + text[1:]


def recommended_headings(page: dict, serp: dict) -> tuple[list[str], list[str]]:
    terms = frequent_terms(serp, page.get("Supporting Keywords"), page.get("Primary Keyword"), max_terms=10)
    role = str(page.get("Page Role") or "").lower()
    functional = str(page.get("Functional Filter") or "").lower()
    primary = str(page.get("Primary Keyword") or "").strip()
    h2 = []
    if "adult" in functional or "performance" in functional:
        h2 = [
            f"Who should buy {primary}",
            "Speed, voltage, range, suspension and terrain fit",
            "How to compare performance, price and support",
        ]
    elif "kids" in functional or "youth" in functional:
        h2 = [
            f"Is {primary} the right fit for your child",
            "Age, size, voltage and speed safety checks",
            "Parent buying checklist before checkout",
        ]
    elif "street legal" in functional:
        h2 = [
            f"What {primary} means for US riders",
            "Dual-sport, registration and compliance questions",
            "What to verify before riding on public roads",
        ]
    elif "comparison" in functional:
        h2 = [
            f"How {primary} compares on price, power and use case",
            "When to choose a budget alternative",
            "Support, parts and ownership tradeoffs",
        ]
    elif "battery" in functional or "power" in functional:
        h2 = [
            f"What to know about {primary}",
            "Voltage, range, charging and battery life",
            "How to choose power for rider size and terrain",
        ]
    elif "parts" in functional or "support" in functional:
        h2 = [
            f"What to check before buying {primary}",
            "Compatibility, maintenance and replacement timing",
            "When to contact support or order parts",
        ]
    else:
        h2 = [
            f"What to consider about {primary}",
            "Buyer fit, terrain fit and ownership cost",
            "Frequently asked questions before buying",
        ]
    h3 = [f"Related topic: {term}" for term in terms[:8]]
    return h2[:5], h3[:8]


def url_slug(page: dict) -> str:
    existing = page.get("URL Slug") or page.get("url_slug")
    if existing:
        return str(existing)
    base = page.get("Primary Keyword") or page.get("Expanded Page Candidate") or "pagina"
    slug = norm(base).replace("/", " ")
    slug = re.sub(r"[^a-z0-9\s-]", "", slug)
    slug = re.sub(r"\s+", "-", slug).strip("-")[:80]
    section = norm(page.get("Site Section") or "resources").replace(" ", "-")
    return f"/{section}/{slug}" if slug else f"/{section}/pagina"


def page_type(page: dict) -> str:
    role = str(page.get("Page Role") or "").lower()
    functional = str(page.get("Functional Filter") or "").lower()
    if "adult" in functional or "kids" in functional or "shop" in functional:
        return "Collection"
    if "street legal" in functional:
        return "Street Legal Guide"
    if "comparison" in functional:
        return "Comparison Guide"
    if "battery" in functional or "power" in functional:
        return "Buying Guide"
    if "support" in functional or "parts" in functional:
        return "Support Guide"
    if "product" in role:
        return "Product Guide"
    if "faq" in role:
        return "FAQ / Parent Section"
    return "SEO Article"


def priority_score(page: dict, intent_match: str, overlap_conflict: str) -> tuple[int, str]:
    decision = page.get("Semantic Production Decision") or page.get("Production Decision") or ""
    volume = safe_float(page.get("Total Search Volume"))
    qa = safe_float(page.get("QA Score"))
    semantic = str(page.get("Semantic QA Decision") or "")
    risk = str(page.get("Risk Flags") or "")
    score = 0
    score += 35 if decision == "Build now" else 22 if decision == "Build later" else 12 if decision == "Manual review before build" else 8
    score += min(20, int(math.log10(max(volume, 1)) * 7))
    score += min(20, int(qa / 5))
    score += 15 if intent_match == "Strong" else 8 if intent_match == "Medium" else -8
    score += 8 if semantic == "Semantic Pass" else -8 if semantic == "Semantic Block" else 0
    if risk:
        score -= min(15, 5 + len([x for x in risk.split(",") if x.strip()]) * 3)
    if overlap_conflict == "Potential duplicate / merge review":
        score -= 5
    if overlap_conflict == "Potential split / intent mismatch":
        score -= 8
    score = max(0, min(100, score))
    reason = f"decision={decision}; volume={int(volume)}; qa={int(qa)}; serp={intent_match}; semantic={semantic or 'n/a'}; overlap={overlap_conflict or 'n/a'}"
    return score, reason


def safe_float(value) -> float:
    try:
        if value is None or value == "":
            return 0.0
        return float(value)
    except (TypeError, ValueError):
        return 0.0


def headers(ws) -> dict[str, int]:
    return {cell.value: i + 1 for i, cell in enumerate(ws[1])}


def row_dict(ws, row_idx: int, h: dict[str, int]) -> dict:
    return {name: ws.cell(row_idx, col).value for name, col in h.items()}


def append_columns(ws, new_headers: list[str], values_by_row: dict[int, list]) -> None:
    existing = headers(ws)
    start = ws.max_column + 1
    effective = []
    for name in new_headers:
        if name in existing:
            effective.append((name, existing[name]))
        else:
            ws.cell(1, start).value = name
            effective.append((name, start))
            start += 1
    for row_idx, vals in values_by_row.items():
        for offset, val in enumerate(vals):
            ws.cell(row_idx, effective[offset][1]).value = val
    style_sheet(ws)


def style_sheet(ws) -> None:
    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    for col_idx, col in enumerate(ws.iter_cols(min_row=1, max_row=min(ws.max_row, 200)), 1):
        max_len = 10
        for cell in col:
            max_len = max(max_len, min(70, len(str(cell.value or ""))))
        ws.column_dimensions[get_column_letter(col_idx)].width = max_len + 2


def delete_if_exists(wb, sheet_names: list[str]) -> None:
    for sheet_name in sheet_names:
        if sheet_name in wb.sheetnames:
            del wb[sheet_name]


def select_pages(rows: list[dict], max_pages: int) -> list[dict]:
    def rank(row):
        semantic_decision = row.get("Semantic Production Decision") or row.get("Production Decision") or ""
        semantic_gate = row.get("Semantic QA Decision") or ""
        decision_rank = {"Build now": 0, "Manual review before build": 1, "Build later": 2, "Merge into parent/FAQ": 3}.get(semantic_decision, 4)
        gate_rank = {"Semantic Block": 0, "Semantic Review": 1, "Semantic Pass": 2}.get(semantic_gate, 3)
        return (decision_rank, gate_rank, -safe_float(row.get("Total Search Volume")), str(row.get("Primary Keyword") or ""))
    selected = sorted(rows, key=rank)
    if max_pages and max_pages > 0:
        return selected[:max_pages]
    return selected


def build_pair_reviews(pages: list[dict], serp_by_key: dict[str, dict]) -> tuple[list[list], dict[str, str]]:
    rows = []
    conflict_by_page = defaultdict(str)
    for i in range(len(pages)):
        a = pages[i]
        for j in range(i + 1, len(pages)):
            b = pages[j]
            same_topic = (a.get("Hub"), a.get("Macro Pillar"), a.get("Topic")) == (b.get("Hub"), b.get("Macro Pillar"), b.get("Topic"))
            same_cluster = same_topic and a.get("Cluster") == b.get("Cluster")
            if not same_topic:
                continue
            sa = serp_by_key.get(a["_key"], {})
            sb = serp_by_key.get(b["_key"], {})
            if sa.get("status") != "ok" or sb.get("status") != "ok":
                continue
            overlap, domain_overlap, text_overlap = serp_overlap(sa, sb)
            recommendation = "Keep separate"
            reason = "SERP overlap is moderate or weak."
            if overlap >= 0.55:
                recommendation = "Merge review"
                reason = "High SERP overlap suggests these candidates may satisfy the same search intent."
                conflict_by_page[a["_key"]] = "Potential duplicate / merge review"
                conflict_by_page[b["_key"]] = "Potential duplicate / merge review"
            elif same_cluster and overlap <= 0.12 and text_overlap <= 0.08:
                recommendation = "Split review"
                reason = "Same assigned cluster but low SERP overlap suggests mixed intent."
                conflict_by_page.setdefault(a["_key"], "Potential split / intent mismatch")
                conflict_by_page.setdefault(b["_key"], "Potential split / intent mismatch")
            if recommendation != "Keep separate":
                rows.append([
                    a.get("Primary Keyword"),
                    b.get("Primary Keyword"),
                    a.get("Expanded Page Candidate"),
                    b.get("Expanded Page Candidate"),
                    a.get("Hub"),
                    a.get("Macro Pillar"),
                    a.get("Topic"),
                    a.get("Cluster"),
                    overlap,
                    domain_overlap,
                    text_overlap,
                    recommendation,
                    reason,
                ])
    return rows, conflict_by_page


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("--source", default=str(DEFAULT_SOURCE))
    parser.add_argument("--output", default=str(DEFAULT_OUTPUT))
    parser.add_argument("--cache", default=str(DEFAULT_CACHE))
    parser.add_argument("--env", default=str(DEFAULT_ENV))
    parser.add_argument("--serpapi-key", default="")
    parser.add_argument("--gl", default="us")
    parser.add_argument("--hl", default="en")
    parser.add_argument("--num", type=int, default=10)
    parser.add_argument("--max-pages", type=int, default=0, help="0 means all Production_Batches rows")
    parser.add_argument("--dry-run", action="store_true", help="Do not call SerpAPI; use existing cache only")
    args = parser.parse_args()

    source = Path(args.source)
    output = Path(args.output)
    cache_file = Path(args.cache)
    env_file = Path(args.env)

    if not source.exists():
        raise FileNotFoundError(f"Missing source workbook: {source}")
    api_key = ""
    if not args.dry_run:
        api_key = load_env_key(env_file, args.serpapi_key)

    output.parent.mkdir(parents=True, exist_ok=True)
    shutil.copy2(source, output)
    wb = load_workbook(output)
    if "Production_Batches" not in wb.sheetnames:
        raise RuntimeError("Source workbook must contain Production_Batches")

    prod = wb["Production_Batches"]
    h = headers(prod)
    required = ["Hub", "Macro Pillar", "Topic", "Cluster", "Expanded Page Candidate", "Primary Keyword", "Production Decision", "Total Search Volume", "Supporting Keywords"]
    missing = [col for col in required if col not in h]
    if missing:
        raise RuntimeError(f"Production_Batches missing required columns: {missing}")

    all_rows = []
    for row_idx in range(2, prod.max_row + 1):
        row = row_dict(prod, row_idx, h)
        row["_row_idx"] = row_idx
        row["_key"] = norm(row.get("Primary Keyword") or row.get("Expanded Page Candidate") or f"row-{row_idx}")
        all_rows.append(row)

    selected = select_pages(all_rows, args.max_pages)
    cache = load_cache(cache_file)
    serp_by_key = {}
    serp_rows = []
    for idx, row in enumerate(selected, 1):
        query = str(row.get("Primary Keyword") or row.get("Expanded Page Candidate") or "").strip()
        if not query:
            continue
        query_key = norm(query)
        if args.dry_run and query_key not in cache:
            serp = {"status": "missing-cache", "message": "dry-run mode and query not cached", "top_urls": [], "titles": [], "snippets": [], "domains": []}
        else:
            serp = fetch_serp(query, api_key, cache, cache_file, args.gl, args.hl, args.num)
        serp_by_key[row["_key"]] = serp
        intent_match, matched_terms, intent_reason = serp_intent_match(serp)
        h2, h3 = recommended_headings(row, serp)
        serp_rows.append([
            row.get("Hub"), row.get("Macro Pillar"), row.get("Topic"), row.get("Cluster"),
            row.get("Expanded Page Candidate"), query, serp.get("status"), serp.get("message"),
            intent_match, matched_terms, intent_reason,
            " | ".join(serp.get("top_urls") or []),
            " | ".join(serp.get("titles") or []),
            " | ".join((serp.get("snippets") or [])[:5]),
            recommended_h1(row), " | ".join(h2), " | ".join(h3), url_slug(row), page_type(row),
        ])
        if idx % 25 == 0:
            print(f"processed SERP {idx}/{len(selected)}")

    pair_rows, conflict_by_page = build_pair_reviews(selected, serp_by_key)

    serp_values = {}
    blueprint_rows = []
    internal_link_rows = []
    for row in all_rows:
        serp = serp_by_key.get(row["_key"], {})
        intent_match, matched_terms, intent_reason = serp_intent_match(serp)
        conflict = conflict_by_page.get(row["_key"], "")
        h2, h3 = recommended_headings(row, serp)
        priority, priority_reason = priority_score(row, intent_match, conflict)
        decision = row.get("Semantic Production Decision") or row.get("Production Decision")
        serp_recommended_decision = decision
        if intent_match == "Weak" and decision == "Build now":
            serp_recommended_decision = "Manual review before build"
        if conflict == "Potential duplicate / merge review" and decision in {"Build now", "Build later"}:
            serp_recommended_decision = "Merge review before build"
        # Split flags are evidence for editorial QA, not enough by themselves to downgrade production.
        vals = [
            "Yes" if serp else "No",
            serp.get("status", "not-selected"),
            intent_match if serp else "Not checked",
            matched_terms,
            conflict,
            serp_recommended_decision,
            priority,
            priority_reason,
            url_slug(row),
            recommended_h1(row),
            page_type(row),
            " | ".join(h2),
            " | ".join(h3),
        ]
        serp_values[row["_row_idx"]] = vals
        blueprint_rows.append([
            row.get("Hub"), row.get("Macro Pillar"), row.get("Topic"), row.get("Cluster"),
            row.get("Expanded Page Candidate"), row.get("Primary Keyword"), url_slug(row), recommended_h1(row),
            page_type(row), row.get("Search Intent Class") or row.get("Opportunity Type"),
            row.get("Hub"), row.get("Macro Pillar"), row.get("Topic"), row.get("Cluster"),
            " | ".join(h2), " | ".join(h3), priority, serp_recommended_decision, priority_reason,
        ])

    link_candidates = sorted(all_rows, key=lambda r: -safe_float(r.get("Total Search Volume")))
    for row in all_rows:
        sources = []
        row_tokens = tokens(str(row.get("Expanded Page Candidate") or "") + " " + str(row.get("Primary Keyword") or ""))
        for other in link_candidates:
            if other is row:
                continue
            same_hub = row.get("Hub") == other.get("Hub")
            same_topic = same_hub and row.get("Topic") == other.get("Topic")
            if not same_hub:
                continue
            other_tokens = tokens(str(other.get("Expanded Page Candidate") or "") + " " + str(other.get("Primary Keyword") or ""))
            sim = jaccard(row_tokens, other_tokens)
            structural = 0.25 if same_topic else 0.1
            score = round(sim + structural + min(0.2, safe_float(other.get("Total Search Volume")) / 50000), 3)
            if score >= 0.22:
                sources.append((score, other))
        for score, target in sorted(sources, key=lambda x: -x[0])[:8]:
            internal_link_rows.append([
                row.get("Primary Keyword"), url_slug(row), target.get("Primary Keyword"), url_slug(target),
                row.get("Hub"), row.get("Topic"), target.get("Topic"), score,
                target.get("Primary Keyword"),
            ])

    serp_headers = [
        "SERP Checked", "SERP Status", "SERP Intent Match", "SERP Matched Terms", "SERP Merge/Split Flag",
        "SERP Recommended Decision", "Page Priority Score", "Page Priority Reason", "Recommended URL Slug",
        "Recommended H1", "Recommended Page Type", "Recommended H2", "Recommended H3",
    ]
    append_columns(prod, serp_headers, serp_values)

    delete_if_exists(wb, ["SERP_Evidence", "SERP_Merge_Review", "SERP_Content_Blueprint", "Internal_Links", "SERP_Feature_Audit"])

    ws = wb.create_sheet("SERP_Evidence")
    ws.append([
        "Hub", "Macro Pillar", "Topic", "Cluster", "Expanded Page Candidate", "Primary Keyword", "SERP Status", "SERP Message",
        "SERP Intent Match", "SERP Matched Terms", "SERP Intent Reason", "Top URLs", "Titles", "Snippets", "Recommended H1",
        "Recommended H2", "Recommended H3", "Recommended URL Slug", "Recommended Page Type",
    ])
    for row in serp_rows:
        ws.append(row)
    style_sheet(ws)

    ws = wb.create_sheet("SERP_Merge_Review")
    ws.append([
        "Keyword A", "Keyword B", "Page Candidate A", "Page Candidate B", "Hub", "Macro Pillar", "Topic", "Cluster",
        "SERP Overlap Score", "Domain Overlap Score", "Title/Snippet Token Overlap", "Recommendation", "Reason",
    ])
    for row in sorted(pair_rows, key=lambda r: (-safe_float(r[8]), r[0], r[1])):
        ws.append(row)
    style_sheet(ws)

    ws = wb.create_sheet("SERP_Content_Blueprint")
    ws.append([
        "Hub", "Macro Pillar", "Topic", "Cluster", "Expanded Page Candidate", "Primary Keyword", "url_slug", "H1", "Page Type",
        "Search Intent", "Level 1 Category", "Level 2 Subcategory", "Level 3 Page", "Cluster Label", "H2 Keywords", "H3 Keywords",
        "Page Priority Score", "SERP Recommended Decision", "Page Priority Reason",
    ])
    for row in sorted(blueprint_rows, key=lambda r: (-safe_float(r[16]), r[0], r[1], r[5])):
        ws.append(row)
    style_sheet(ws)

    ws = wb.create_sheet("Internal_Links")
    ws.append([
        "Source Primary Keyword", "Source URL", "Target Primary Keyword", "Target URL", "Hub", "Source Topic", "Target Topic", "Link Score", "Anchor Text",
    ])
    for row in sorted(internal_link_rows, key=lambda r: (r[4], r[0], -safe_float(r[7])))[:3000]:
        ws.append(row)
    style_sheet(ws)

    ws = wb.create_sheet("SERP_Feature_Audit")
    ws.append(["Feature", "Status", "Evidence", "Notes"])
    checked = sum(1 for row in serp_rows if row[6] == "ok")
    ws.append(["SERP validation layer", "Enabled", f"{checked}/{len(serp_rows)} selected candidates returned organic results", "Reads Production_Batches and appends SERP fields without replacing semantic decisions"])
    ws.append(["SERP cache", "Enabled" if cache_file.exists() else "Missing", str(cache_file), "One JSONL row per normalized primary keyword query"])
    ws.append(["SERP market", "Enabled", f"gl={args.gl}; hl={args.hl}; num={args.num}", "Defaults target United States English SERPs"])
    ws.append(["SERP merge review", "Enabled", len(pair_rows), "Only same-topic candidate pairs are flagged for merge/split review"])
    ws.append(["SERP content blueprint", "Enabled", len(blueprint_rows), "Adds url_slug, H1, page type, H2/H3, priority score and recommended decision"])
    ws.append(["Internal links", "Enabled", len(internal_link_rows), "Structure and token-similarity based suggestions; SERP is not required for links"])
    style_sheet(ws)

    wb.save(output)
    print(f"wrote {output.resolve()}")
    print(f"selected pages: {len(selected)}")
    print(f"serp evidence rows: {len(serp_rows)}")
    print(f"merge/split review rows: {len(pair_rows)}")
    print(f"internal link rows: {len(internal_link_rows)}")


if __name__ == "__main__":
    main()
